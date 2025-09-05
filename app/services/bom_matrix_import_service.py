#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
from typing import List, Dict, Tuple, Optional
from datetime import datetime
from app.db import query_one, query_all, execute, get_last_id
from app.services.bom_service import BomService
from app.services.item_service import ItemService
from app.services.bom_history_service import BomHistoryService


class BomMatrixImportService:
    """BOM矩阵导入服务 - 处理导出的Excel矩阵格式"""
    
    @staticmethod
    def normalize_string(text: str) -> str:
        """
        标准化字符串，用于匹配
        去除空格、连字符、下划线等，转换为小写
        
        Args:
            text: 原始字符串
            
        Returns:
            str: 标准化后的字符串
        """
        if not text:
            return ""
        
        # 去除首尾空格
        normalized = str(text).strip()
        
        # 将多个空格替换为单个空格
        normalized = " ".join(normalized.split())
        
        # 将连字符、下划线、点号都替换为空格
        normalized = normalized.replace("-", " ").replace("_", " ").replace(".", " ")
        
        # 去除所有空格
        normalized = normalized.replace(" ", "")
        
        # 转换为小写
        normalized = normalized.lower()
        
        return normalized
    
    @staticmethod
    def parse_matrix_excel(file_path: str) -> Tuple[Dict, List[str]]:
        """
        解析矩阵格式的Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Tuple[Dict, List[str]]: (解析结果, 错误信息列表)
        """
        try:
            print(f"开始解析矩阵Excel文件: {file_path}")
            
            # 使用openpyxl读取Excel文件
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 获取数据范围
            max_row = ws.max_row
            max_col = ws.max_column
            
            print(f"Excel文件尺寸: {max_row}行 x {max_col}列")
            
            if max_row < 5 or max_col < 4:
                return {}, ["文件格式错误：至少需要5行4列数据"]
            
            # 解析成品信息（第1-4行，从D列开始）
            products = []
            product_start_col = 4  # D列
            
            for col in range(product_start_col, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                
                # 检查是否有成品数据
                code = ws[f'{col_letter}1'].value
                name = ws[f'{col_letter}2'].value
                spec = ws[f'{col_letter}3'].value
                brand = ws[f'{col_letter}4'].value
                
                if code and name and spec and brand:
                    product = {
                        'ItemCode': str(code).strip(),
                        'CnName': str(name).strip(),
                        'ItemSpec': str(spec).strip(),
                        'Brand': str(brand).strip(),
                        'Column': col,
                        'ColumnLetter': col_letter
                    }
                    products.append(product)
                    print(f"成品: {product['Brand']} - {product['CnName']} ({product['ItemCode']})")
            
            if not products:
                return {}, ["没有找到有效的成品数据"]
            
            # 解析零部件信息（第5行开始，A-C列）
            components = []
            
            for row in range(5, max_row + 1):
                code = ws[f'A{row}'].value
                name = ws[f'B{row}'].value
                spec = ws[f'C{row}'].value
                
                if code and name and spec:
                    component = {
                        'ItemCode': str(code).strip(),
                        'CnName': str(name).strip(),
                        'ItemSpec': str(spec).strip(),
                        'Row': row
                    }
                    components.append(component)
                    print(f"零部件: {component['CnName']} ({component['ItemCode']})")
            
            if not components:
                return {}, ["没有找到有效的零部件数据"]
            
            # 解析数量矩阵（包括0值）
            quantity_matrix = {}
            
            for component in components:
                component_quantities = {}
                
                for product in products:
                    quantity_cell = ws[f"{product['ColumnLetter']}{component['Row']}"].value
                    
                    try:
                        quantity = float(quantity_cell) if quantity_cell is not None else 0
                    except (ValueError, TypeError):
                        quantity = 0
                    
                    # 记录所有数量，包括0值
                    component_quantities[product['Brand']] = quantity
                
                # 记录所有零部件，即使数量都为0
                quantity_matrix[component['ItemCode']] = {
                    'component': component,
                    'quantities': component_quantities
                }
            
            result = {
                'products': products,
                'components': components,
                'quantity_matrix': quantity_matrix
            }
            
            print(f"解析完成: {len(products)}个成品, {len(components)}个零部件")
            return result, []
            
        except Exception as e:
            return {}, [f"解析Excel文件失败: {str(e)}"]
    
    @staticmethod
    def find_product_by_brand_and_info(brand: str, code: str, name: str, spec: str) -> Tuple[Optional[int], List[str]]:
        """
        通过品牌和其他信息查找成品
        
        Args:
            brand: 品牌
            code: 编码
            name: 名称
            spec: 规格
            
        Returns:
            Tuple[Optional[int], List[str]]: (物料ID, 错误信息列表)
        """
        try:
            # 首先尝试通过品牌精确匹配
            sql = """
                SELECT ItemId, ItemCode, CnName, ItemSpec, Brand FROM Items 
                WHERE Brand = ? AND ItemType = 'FG'
            """
            items = query_all(sql, (brand,))
            
            if not items:
                return None, [f"未找到品牌为 '{brand}' 的成品"]
            
            # 如果有多个匹配，尝试通过其他字段进一步匹配
            if len(items) == 1:
                item = items[0]
                print(f"找到唯一成品: {item['Brand']} - {item['CnName']} ({item['ItemCode']})")
                return item['ItemId'], []
            
            # 多个匹配时，尝试通过编码、名称、规格进一步匹配
            normalized_code = BomMatrixImportService.normalize_string(code)
            normalized_name = BomMatrixImportService.normalize_string(name)
            normalized_spec = BomMatrixImportService.normalize_string(spec)
            
            best_match = None
            best_score = 0
            
            for item in items:
                score = 0
                
                # 编码匹配
                if BomMatrixImportService.normalize_string(item['ItemCode']) == normalized_code:
                    score += 3
                
                # 名称匹配
                if BomMatrixImportService.normalize_string(item['CnName']) == normalized_name:
                    score += 2
                
                # 规格匹配
                if BomMatrixImportService.normalize_string(item['ItemSpec']) == normalized_spec:
                    score += 1
                
                if score > best_score:
                    best_score = score
                    best_match = item
            
            if best_match and best_score > 0:
                print(f"找到最佳匹配成品: {best_match['Brand']} - {best_match['CnName']} ({best_match['ItemCode']}) 得分: {best_score}")
                return best_match['ItemId'], []
            
            # 如果都没有匹配，返回第一个
            print(f"使用第一个匹配的成品: {items[0]['Brand']} - {items[0]['CnName']} ({items[0]['ItemCode']})")
            return items[0]['ItemId'], []
            
        except Exception as e:
            return None, [f"查找成品失败 {brand}: {str(e)}"]
    
    @staticmethod
    def find_component_by_code_and_spec(code: str, spec: str) -> Tuple[Optional[int], List[str]]:
        """
        通过编码和规格查找零部件
        
        Args:
            code: 编码
            spec: 规格
            
        Returns:
            Tuple[Optional[int], List[str]]: (物料ID, 错误信息列表)
        """
        try:
            # 获取所有类型的物料（包括成品作为组件的情况）
            sql = """
                SELECT ItemId, ItemCode, CnName, ItemSpec FROM Items 
                WHERE ItemType IN ('RM', 'SFG', 'FG')
            """
            items = query_all(sql)
            
            if not items:
                return None, ["没有找到任何零部件物料"]
            
            # 标准化搜索条件
            normalized_code = BomMatrixImportService.normalize_string(code)
            normalized_spec = BomMatrixImportService.normalize_string(spec)
            
            best_match = None
            best_score = 0
            
            for item in items:
                score = 0
                
                # 编码匹配（权重更高）
                if BomMatrixImportService.normalize_string(item['ItemCode']) == normalized_code:
                    score += 3
                
                # 规格匹配
                if BomMatrixImportService.normalize_string(item['ItemSpec']) == normalized_spec:
                    score += 2
                
                if score > best_score:
                    best_score = score
                    best_match = item
            
            if best_match and best_score >= 2:  # 至少要有编码或规格匹配
                print(f"找到零部件: {best_match['CnName']} ({best_match['ItemCode']}) 得分: {best_score}")
                return best_match['ItemId'], []
            
            return None, [f"未找到零部件: {code} ({spec})"]
            
        except Exception as e:
            return None, [f"查找零部件失败 {code}: {str(e)}"]
    
    @staticmethod
    def find_or_create_bom_by_brand(brand: str, product_id: int) -> Tuple[Optional[int], List[str]]:
        """
        通过品牌查找或创建BOM
        
        Args:
            brand: 品牌（作为BOM名称）
            product_id: 父产品ID
            
        Returns:
            Tuple[Optional[int], List[str]]: (BOM ID, 错误信息列表)
        """
        try:
            # 查找现有BOM
            sql = """
                SELECT BomId FROM BomHeaders 
                WHERE BomName = ? AND ParentItemId = ?
            """
            existing_bom = query_one(sql, (brand, product_id))
            
            if existing_bom:
                print(f"找到现有BOM: {brand} -> ID: {existing_bom['BomId']}")
                return existing_bom['BomId'], []
            
            # 创建新BOM
            bom_data = {
                'BomName': brand,
                'ParentItemId': product_id,
                'Rev': 'A',
                'EffectiveDate': datetime.now().strftime("%Y-%m-%d"),
                'ExpireDate': '2035-12-31',
                'Remark': f"从矩阵Excel导入的BOM - {brand}"
            }
            
            bom_id = BomService.create_bom_header(bom_data)
            print(f"创建新BOM: {brand} -> ID: {bom_id}")
            return bom_id, []
            
        except Exception as e:
            return None, [f"查找或创建BOM失败 {brand}: {str(e)}"]
    
    @staticmethod
    def update_bom_quantities(bom_id: int, component_id: int, quantity: float) -> Tuple[bool, List[str]]:
        """
        更新BOM中的零部件数量
        
        Args:
            bom_id: BOM ID
            component_id: 零部件ID
            quantity: 数量
            
        Returns:
            Tuple[bool, List[str]]: (是否成功, 错误信息列表)
        """
        try:
            # 检查是否已存在该BOM行
            sql = """
                SELECT LineId FROM BomLines 
                WHERE BomId = ? AND ChildItemId = ?
            """
            existing_line = query_one(sql, (bom_id, component_id))
            
            if existing_line:
                if quantity > 0:
                    # 获取旧数据用于历史记录
                    old_line = query_one("SELECT * FROM BomLines WHERE LineId = ?", (existing_line['LineId'],))
                    
                    # 更新现有行
                    update_sql = """
                        UPDATE BomLines 
                        SET QtyPer = ?, CreatedDate = ?
                        WHERE LineId = ?
                    """
                    execute(update_sql, (quantity, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), existing_line['LineId']))
                    
                    # 记录更新历史
                    BomHistoryService.log_operation(
                        bom_id=bom_id,
                        operation_type='UPDATE',
                        operation_target='LINE',
                        target_id=existing_line['LineId'],
                        old_data=dict(old_line) if old_line else None,
                        new_data={'ChildItemId': component_id, 'QtyPer': quantity, 'ScrapFactor': 0},
                        operation_user='系统',
                        operation_source='IMPORT',
                        remark=f"导入更新数量: {quantity}"
                    )
                    
                    print(f"更新BOM行: BOM {bom_id} -> 零部件 {component_id} = {quantity}")
                else:
                    # 获取旧数据用于历史记录
                    old_line = query_one("SELECT * FROM BomLines WHERE LineId = ?", (existing_line['LineId'],))
                    
                    # 删除现有行（数量为0或负数）
                    delete_sql = """
                        DELETE FROM BomLines 
                        WHERE LineId = ?
                    """
                    execute(delete_sql, (existing_line['LineId'],))
                    
                    # 记录删除历史
                    BomHistoryService.log_operation(
                        bom_id=bom_id,
                        operation_type='DELETE',
                        operation_target='LINE',
                        target_id=existing_line['LineId'],
                        old_data=dict(old_line) if old_line else None,
                        operation_user='系统',
                        operation_source='IMPORT',
                        remark=f"导入删除零部件 (数量={quantity})"
                    )
                    
                    print(f"删除BOM行: BOM {bom_id} -> 零部件 {component_id} (数量={quantity})")
            else:
                if quantity > 0:
                    # 创建新行
                    BomService.create_bom_line(bom_id, {
                        'ChildItemId': component_id,
                        'QtyPer': quantity,
                        'ScrapFactor': 0
                    })
                    print(f"创建BOM行: BOM {bom_id} -> 零部件 {component_id} = {quantity}")
                else:
                    # 数量为0，不需要创建
                    print(f"跳过创建BOM行: BOM {bom_id} -> 零部件 {component_id} (数量={quantity})")
            
            return True, []
            
        except Exception as e:
            return False, [f"更新BOM数量失败: {str(e)}"]
    
    @staticmethod
    def import_matrix_excel(file_path: str) -> Tuple[int, List[str], List[str]]:
        """
        导入矩阵格式的Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Tuple[int, List[str], List[str]]: (成功数量, 错误信息列表, 警告信息列表)
        """
        try:
            print("=== 开始导入矩阵Excel文件 ===")
            print(f"文件路径: {file_path}")
            
            # 解析Excel文件
            result, parse_errors = BomMatrixImportService.parse_matrix_excel(file_path)
            if parse_errors:
                print(f"解析错误: {parse_errors}")
                return 0, parse_errors, []
            
            products = result['products']
            components = result['components']
            quantity_matrix = result['quantity_matrix']
            
            print(f"解析结果: {len(products)}个成品, {len(components)}个零部件, {len(quantity_matrix)}个数量关系")
            
            success_count = 0
            errors = []
            warnings = []
            
            # 处理每个成品
            for product in products:
                try:
                    print(f"\n处理成品: {product['Brand']} - {product['CnName']}")
                    
                    # 查找成品物料
                    print(f"  查找成品: 品牌={product['Brand']}, 编码={product['ItemCode']}, 名称={product['CnName']}, 规格={product['ItemSpec']}")
                    product_id, product_errors = BomMatrixImportService.find_product_by_brand_and_info(
                        product['Brand'], product['ItemCode'], product['CnName'], product['ItemSpec']
                    )
                    
                    if product_errors:
                        print(f"  成品查找错误: {product_errors}")
                        errors.extend(product_errors)
                        continue
                    
                    if not product_id:
                        print(f"  未找到成品: {product['Brand']} - {product['CnName']}")
                        errors.append(f"未找到成品: {product['Brand']} - {product['CnName']}")
                        continue
                    
                    print(f"  找到成品ID: {product_id}")
                    
                    # 查找或创建BOM
                    bom_id, bom_errors = BomMatrixImportService.find_or_create_bom_by_brand(
                        product['Brand'], product_id
                    )
                    
                    if bom_errors:
                        errors.extend(bom_errors)
                        continue
                    
                    if not bom_id:
                        errors.append(f"无法创建BOM: {product['Brand']}")
                        continue
                    
                    # 获取现有BOM结构
                    existing_structure = BomMatrixImportService.get_existing_bom_structure(bom_id)
                    
                    # 处理Excel中的所有零部件
                    product_success_count = 0
                    processed_components = set()  # 记录已处理的零部件
                    
                    for component_code, component_data in quantity_matrix.items():
                        try:
                            component = component_data['component']
                            quantities = component_data['quantities']
                            
                            # 获取该零部件在当前成品中的数量（如果没有则为0）
                            quantity = quantities.get(product['Brand'], 0)
                            
                            print(f"  处理零部件: {component['CnName']} -> 数量: {quantity}")
                            
                            # 查找零部件
                            component_id, component_errors = BomMatrixImportService.find_component_by_code_and_spec(
                                component['ItemCode'], component['ItemSpec']
                            )
                            
                            if component_errors:
                                errors.extend(component_errors)
                                continue
                            
                            if not component_id:
                                errors.append(f"未找到零部件: {component['ItemCode']} ({component['ItemSpec']})")
                                continue
                            
                            # 记录已处理的零部件
                            processed_components.add(component_id)
                            
                            # 更新BOM数量（包括删除和清零）
                            success, update_errors = BomMatrixImportService.update_bom_quantities(
                                bom_id, component_id, quantity
                            )
                            
                            if success:
                                product_success_count += 1
                                success_count += 1
                                print(f"  处理完成: {component['CnName']} = {quantity}")
                            else:
                                errors.extend(update_errors)
                        
                        except Exception as e:
                            error_msg = f"处理零部件失败 {component_code}: {str(e)}"
                            errors.append(error_msg)
                            print(f"  错误: {error_msg}")
                    
                    # 删除Excel中不存在的零部件关系
                    for existing_component_id in existing_structure:
                        if existing_component_id not in processed_components:
                            # 这个零部件在Excel中不存在，应该删除
                            success, delete_errors = BomMatrixImportService.update_bom_quantities(
                                bom_id, existing_component_id, 0
                            )
                            if success:
                                product_success_count += 1
                                success_count += 1
                                print(f"  删除多余零部件: ID {existing_component_id}")
                            else:
                                errors.extend(delete_errors)
                    
                    if product_success_count > 0:
                        print(f"成品 {product['Brand']} 处理完成: {product_success_count} 个零部件")
                    else:
                        warnings.append(f"成品 {product['Brand']} 没有找到任何有效的零部件关系")
                
                except Exception as e:
                    error_msg = f"处理成品失败 {product['Brand']}: {str(e)}"
                    errors.append(error_msg)
                    print(f"错误: {error_msg}")
            
            print(f"\n=== 导入完成 ===")
            print(f"成功更新: {success_count} 个BOM关系")
            print(f"错误数量: {len(errors)}")
            print(f"警告数量: {len(warnings)}")
            
            return success_count, errors, warnings
            
        except Exception as e:
            return 0, [f"导入矩阵Excel失败: {str(e)}"], []
    
    @staticmethod
    def get_existing_bom_structure(bom_id: int) -> Dict[int, float]:
        """
        获取现有BOM结构
        
        Args:
            bom_id: BOM ID
            
        Returns:
            Dict[int, float]: {零部件ID: 数量}
        """
        try:
            sql = """
                SELECT ChildItemId, QtyPer FROM BomLines 
                WHERE BomId = ?
            """
            lines = query_all(sql, (bom_id,))
            
            existing_structure = {}
            for line in lines:
                existing_structure[line['ChildItemId']] = line['QtyPer']
            
            print(f"现有BOM结构: {len(existing_structure)}个零部件")
            return existing_structure
            
        except Exception as e:
            print(f"获取BOM结构失败: {str(e)}")
            return {}


if __name__ == "__main__":
    # 测试导入
    file_path = "bom_matrix.xlsx"
    success_count, errors, warnings = BomMatrixImportService.import_matrix_excel(file_path)
    
    print(f"\n导入结果:")
    print(f"成功: {success_count}")
    print(f"错误: {errors}")
    print(f"警告: {warnings}")
