# app/ui/mrp_viewer.py
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDateEdit, QLabel, QComboBox, QGroupBox,
    QMessageBox, QHeaderView, QTabWidget, QLineEdit, QCheckBox,
    QFileDialog, QProgressBar
)
from PySide6.QtCore import Qt, QDate, QThread, Signal, QTimer
from PySide6.QtGui import QFont, QColor, QBrush

from app.services.mrp_service import MRPService
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MRPCalcThread(QThread):
    finished = Signal(dict)
    failed = Signal(str)
    progress = Signal(int, str)  # 进度百分比和状态文本

    def __init__(self, start_date: str, end_date: str, import_id: Optional[int] = None, 
                  search_filter: Optional[str] = None, calc_type: str = "comprehensive"):
        super().__init__()
        self.start_date = start_date
        self.end_date = end_date
        self.import_id = import_id
        self.search_filter = search_filter
        self.calc_type = calc_type  # "child", "parent", 或 "comprehensive"

    def run(self):
        try:
            print(f"🔄 [MRPCalcThread] 开始MRP计算")
            print(f"🔄 [MRPCalcThread] 参数：start_date={self.start_date}, end_date={self.end_date}")
            print(f"🔄 [MRPCalcThread] 参数：import_id={self.import_id}, search_filter={self.search_filter}")
            print(f"🔄 [MRPCalcThread] 计算类型：{self.calc_type}")
            
            self.progress.emit(10, "正在初始化计算参数...")
            
            if self.calc_type == "child":
                # 计算零部件MRP
                print(f"🔄 [MRPCalcThread] 调用 calculate_mrp_kanban")
                self.progress.emit(30, "正在计算零部件MRP...")
                data = MRPService.calculate_mrp_kanban(
                    self.start_date, self.end_date, 
                    self.import_id, self.search_filter
                )
            elif self.calc_type == "parent":
                # 计算成品MRP
                print(f"[MRPCalcThread] 调用 calculate_parent_mrp_kanban")
                self.progress.emit(30, "正在计算成品MRP...")
                data = MRPService.calculate_parent_mrp_kanban(
                    self.start_date, self.end_date, 
                    self.import_id, self.search_filter
                )
            else:
                # 计算综合MRP
                print(f"[MRPCalcThread] 调用 calculate_comprehensive_mrp_kanban")
                self.progress.emit(30, "正在计算综合MRP...")
                data = MRPService.calculate_comprehensive_mrp_kanban(
                    self.start_date, self.end_date, 
                    self.import_id, self.search_filter
                )
            
            self.progress.emit(80, "正在处理计算结果...")
            print(f"[MRPCalcThread] 计算完成，返回数据：weeks={len(data.get('weeks', []))}, rows={len(data.get('rows', []))}")
            
            self.progress.emit(100, "计算完成！")
            self.finished.emit(data)
        except Exception as e:
            print(f"[MRPCalcThread] 计算失败：{str(e)}")
            self.failed.emit(str(e))

class MRPViewer(QWidget):
    """MRP 看板（支持零部件和成品两种计算模式）"""
    def __init__(self):
        super().__init__()
        self._build_ui()
        self._thread = None
        self._signal_connected = False  # 跟踪信号连接状态
        self._load_available_data()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(12, 12, 12, 12)

        # 标题
        title = QLabel("MRP 看板（周）")
        title.setAlignment(Qt.AlignCenter)
        title.setFont(QFont("Arial", 16, QFont.Bold))
        layout.addWidget(title)

        # 控制区
        ctrl = QGroupBox("计算参数")
        ctrl.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #cccccc;
                border-radius: 5px;
                margin-top: 1ex;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        cly = QVBoxLayout(ctrl)
        
        # 第一行：日期范围
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("开始日期"))
        self.dt_start = QDateEdit()
        self.dt_start.setCalendarPopup(True)
        self.dt_start.setDate(QDate.currentDate())
        date_layout.addWidget(self.dt_start)

        date_layout.addWidget(QLabel("结束日期"))
        self.dt_end = QDateEdit()
        self.dt_end.setCalendarPopup(True)
        self.dt_end.setDate(QDate.currentDate().addDays(56))
        date_layout.addWidget(self.dt_end)
        
        date_layout.addStretch()
        cly.addLayout(date_layout)
        
        # 第二行：客户订单版本选择
        order_layout = QHBoxLayout()
        order_layout.addWidget(QLabel("客户订单版本:"))
        self.order_version_combo = QComboBox()
        self.order_version_combo.addItem("全部订单汇总", None)
        self.order_version_combo.setMinimumWidth(250)
        self.order_version_combo.setMinimumHeight(12)   # 设置最小高度
        self.order_version_combo.setStyleSheet("""
            QComboBox {
                padding: 6px 12px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background: white;
                min-width: 250px;
                max-width: 300px;
                min-height: 12px;
                max-height: 12px;
            }
            QComboBox:focus {
                border-color: #007bff;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #6c757d;
                margin-right: 5px;
            }
        """)
        order_layout.addWidget(self.order_version_combo)
        
        order_layout.addWidget(QLabel("搜索:"))
        search_layout = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入物料名称或规格进行实时搜索")
        self.search_edit.setMinimumWidth(300)
        self.search_edit.textChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.search_edit)
        
        # 添加重置按钮
        reset_btn = QPushButton("重置")
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 60px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        reset_btn.clicked.connect(self.on_reset_search)
        search_layout.addWidget(reset_btn)
        
        order_layout.addLayout(search_layout)
        
        # 添加刷新按钮
        refresh_btn = QPushButton("刷新订单版本")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_order_versions)
        order_layout.addWidget(refresh_btn)
        
        order_layout.addStretch()
        cly.addLayout(order_layout)
        
        # 第三行：计算类型和按钮
        calc_layout = QHBoxLayout()
        
        # 计算类型选择
        calc_layout.addWidget(QLabel("计算类型:"))
        self.calc_type_combo = QComboBox()
        self.calc_type_combo.addItems(["综合MRP", "零部件MRP", "成品MRP"])
        self.calc_type_combo.setCurrentText("综合MRP")
        calc_layout.addWidget(self.calc_type_combo)
        
        # 说明标签
        type_desc_label = QLabel("综合MRP：结合成品库存和零部件库存计算；零部件MRP：展开BOM计算原材料需求；成品MRP：直接显示成品需求")
        type_desc_label.setStyleSheet("color: #666; font-size: 11px;")
        calc_layout.addWidget(type_desc_label)
        
        calc_layout.addStretch()
        
        # 生成看板按钮
        self.btn_calc = QPushButton("生成看板")
        self.btn_calc.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:disabled {
                background-color: #6c757d;
            }
        """)
        self.btn_calc.clicked.connect(self.on_calc)
        calc_layout.addWidget(self.btn_calc)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #dee2e6;
                border-radius: 5px;
                text-align: center;
                background-color: #f8f9fa;
            }
            QProgressBar::chunk {
                background-color: #007bff;
                border-radius: 3px;
            }
        """)
        calc_layout.addWidget(self.progress_bar)
        
        # 导出Excel按钮
        self.btn_export = QPushButton("导出Excel")
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:disabled {
                background-color: #6c757d;
            }
        """)
        self.btn_export.clicked.connect(self.on_export)
        self.btn_export.setEnabled(False)  # 初始状态禁用
        calc_layout.addWidget(self.btn_export)
        
        cly.addLayout(calc_layout)
        layout.addWidget(ctrl)

        # 表格
        self.tbl = QTableWidget()
        self.tbl.setAlternatingRowColors(True)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setSelectionMode(QTableWidget.NoSelection)
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setStretchLastSection(True)
        
        # 设置表格样式
        self.tbl.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: transparent;
                alternate-background-color: transparent;
                selection-background-color: transparent;
            }
            QTableWidget::item:selected {
                background-color: transparent !important;
                border: 2px solid #007bff;
                border-radius: 2px;
            }
            QTableWidget::item:selected:focus {
                background-color: transparent !important;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                color: #495057;
                padding: 8px 6px;
                border: none;
                border-bottom: 2px solid #dee2e6;
                font-weight: 600;
                font-size: 12px;
            }
        """)
        
        layout.addWidget(self.tbl)

    def _load_available_data(self):
        """加载可用的客户订单版本和成品信息"""
        self.refresh_order_versions()
        
    def refresh_order_versions(self):
        """刷新订单版本列表"""
        try:
            # 保存当前选中的版本
            current_import_id = self.order_version_combo.currentData()
            
            # 加载客户订单版本
            versions = MRPService.get_available_import_versions()
            self.order_version_combo.clear()
            self.order_version_combo.addItem("全部订单汇总", None)
            
            for version in versions:
                display_text = f"{version['ImportId']} - {version['FileName']} ({version['ImportDate']})"
                self.order_version_combo.addItem(display_text, version['ImportId'])
            
            # 尝试恢复之前选中的版本
            if current_import_id is not None:
                for i in range(self.order_version_combo.count()):
                    if self.order_version_combo.itemData(i) == current_import_id:
                        self.order_version_combo.setCurrentIndex(i)
                        break
            
            # 连接订单版本选择变化事件（避免重复连接）
            if self._signal_connected:
                try:
                    self.order_version_combo.currentIndexChanged.disconnect()
                except (TypeError, RuntimeError):
                    pass
                self._signal_connected = False
            
            self.order_version_combo.currentIndexChanged.connect(self.on_order_version_changed)
            self._signal_connected = True
                
        except Exception as e:
            print(f"加载客户订单版本失败: {e}")

    def on_order_version_changed(self):
        """当选择的客户订单版本变化时，自动调整日期范围"""
        import_id = self.order_version_combo.currentData()
        if import_id is None:
            # 选择"全部订单汇总"时，使用默认日期范围
            self.dt_start.setDate(QDate.currentDate())
            self.dt_end.setDate(QDate.currentDate().addDays(56))
            return
        
        try:
            # 获取指定订单版本的时间范围
            order_range = MRPService.get_order_version_date_range(import_id)
            if order_range:
                start_date = order_range.get("earliest_date")
                end_date = order_range.get("latest_date")
                
                if start_date and end_date:
                    # 转换为QDate对象
                    q_start = QDate.fromString(start_date, "yyyy-MM-dd")
                    q_end = QDate.fromString(end_date, "yyyy-MM-dd")
                    
                    if q_start.isValid() and q_end.isValid():
                        # 直接设置为订单的实际时间范围
                        self.dt_start.setDate(q_start)
                        self.dt_end.setDate(q_end)
        except Exception as e:
            print(f"自动调整日期范围失败: {e}")

    # ---- 交互 ----
    def on_calc(self):
        print(f"🔘 [on_calc] 用户点击计算按钮")
        
        s = self.dt_start.date().toString("yyyy-MM-dd")
        e = self.dt_end.date().toString("yyyy-MM-dd")
        print(f"🔘 [on_calc] 日期范围：{s} 到 {e}")
        
        if self.dt_start.date() >= self.dt_end.date():
            print(f"❌ [on_calc] 日期范围错误")
            QMessageBox.warning(self, "提示", "结束日期必须大于开始日期")
            return
            
        # 获取选择的客户订单版本ID
        import_id = self.order_version_combo.currentData()
        print(f"🔘 [on_calc] 客户订单版本ID：{import_id}")
        
        # 获取搜索条件
        search_filter = self.search_edit.text().strip() or None
        print(f"🔘 [on_calc] 搜索条件：{search_filter}")
        
        # 获取计算类型
        calc_type_text = self.calc_type_combo.currentText()
        if calc_type_text == "零部件MRP":
            calc_type = "child"
        elif calc_type_text == "成品MRP":
            calc_type = "parent"
        else:  # 综合MRP
            calc_type = "comprehensive"
        print(f"🔘 [on_calc] 计算类型：{calc_type}")
        
        self.btn_calc.setEnabled(False)
        self.btn_export.setEnabled(False)  # 禁用导出按钮
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("准备开始计算...")
        
        self.tbl.clear()
        
        # 显示计算状态
        self.tbl.setRowCount(1)
        self.tbl.setColumnCount(1)
        self.tbl.setHorizontalHeaderLabels(["计算中..."])
        self.tbl.setItem(0, 0, QTableWidgetItem("正在计算MRP，请稍候..."))
        
        self._thread = MRPCalcThread(s, e, import_id, search_filter, calc_type)
        self._thread.finished.connect(self.render_board)
        self._thread.failed.connect(self.show_error)
        self._thread.progress.connect(self.on_progress_update)
        self._thread.start()

    def on_progress_update(self, progress: int, status: str):
        """更新进度条"""
        self.progress_bar.setValue(progress)
        self.progress_bar.setFormat(f"{status} ({progress}%)")
        if progress == 100:
            # 计算完成后隐藏进度条
            QTimer.singleShot(1000, lambda: self.progress_bar.setVisible(False))

    def show_error(self, msg: str):
        print(f"❌ [show_error] 显示错误：{msg}")
        self.btn_calc.setEnabled(True)
        self.progress_bar.setVisible(False)  # 出错时隐藏进度条
        QMessageBox.critical(self, "错误", msg)

    def on_search_changed(self):
        """当搜索条件变化时，自动重新计算"""
        print(f"🔍 [on_search_changed] 搜索条件变化，触发重新计算")
        # 如果当前有数据，则重新渲染（不重新计算，只过滤显示）
        if hasattr(self, '_current_data') and self._current_data:
            self.render_board(self._current_data)

    def on_reset_search(self):
        """重置搜索条件"""
        print(f"🔄 [on_reset_search] 重置搜索条件")
        self.search_edit.clear()
        # 如果当前有数据，则重新渲染显示所有数据
        if hasattr(self, '_current_data') and self._current_data:
            self.render_board(self._current_data)

    # ---- 渲染 ----
    def render_board(self, data: dict):
        print(f"🎨 [render_board] 开始渲染MRP看板")
        print(f"🎨 [render_board] 接收数据：{data}")
        
        # 隐藏进度条
        self.progress_bar.setVisible(False)
        
        self.btn_calc.setEnabled(True)
        self.btn_export.setEnabled(True)  # 启用导出按钮
        
        # 保存当前数据用于导出
        self._current_data = data
        
        if not data:
            print(f"❌ [render_board] 数据为空，清空表格")
            self.tbl.setRowCount(0); self.tbl.setColumnCount(0); return

        weeks = data.get("weeks", [])
        rows = data.get("rows", [])
        
        # 根据搜索条件过滤数据
        search_text = self.search_edit.text().strip().lower()
        if search_text:
            print(f"🔍 [render_board] 应用搜索过滤：{search_text}")
            filtered_rows = []
            for row in rows:
                item_name = row.get("ItemName", "").lower()
                item_spec = row.get("ItemSpec", "").lower()
                if search_text in item_name or search_text in item_spec:
                    filtered_rows.append(row)
            rows = filtered_rows
            print(f"🔍 [render_board] 过滤后数据行数：{len(rows)}")
        
        print(f"🎨 [render_board] 数据解析：weeks={weeks}, rows数量={len(rows)}")

        # 构建年份分组和合计列
        colspec = self._build_week_columns_with_totals(weeks)
        
        # 根据计算类型设置不同的列标题
        calc_type = self.calc_type_combo.currentText()
        if calc_type == "零部件MRP":
            # 零部件MRP：物料名称、规格、类型、行别、期初库存、各周、合计
            fixed_headers = ["物料名称", "物料规格", "物料类型", "行别", "期初库存"]
        elif calc_type == "成品MRP":
            # 成品MRP：物料名称、规格、类型、行别、期初库存、各周、合计
            fixed_headers = ["物料名称", "物料规格", "成品类型", "行别", "期初库存"]
        else:  # 综合MRP
            # 综合MRP：物料名称、规格、类型、行别、期初库存、各周、合计
            fixed_headers = ["物料名称", "物料规格", "物料类型", "行别", "期初库存"]
        
        # 设置列数和标题
        headers_count = len(fixed_headers) + len(colspec) + 1  # +1 for Total column
        self.tbl.setColumnCount(headers_count)
        
        # 设置固定列标题
        for i, title in enumerate(fixed_headers):
            item = QTableWidgetItem(title)
            self.tbl.setHorizontalHeaderItem(i, item)
        
        # 设置周列和年份合计列标题
        base_col = len(fixed_headers)
        for i, (kind, val) in enumerate(colspec):
            if kind == "week":
                it = QTableWidgetItem(val)  # val is already CW format
                # 设置日期作为用户数据
                date_str = self._convert_cw_to_date(val)
                it.setData(Qt.UserRole, date_str)
            else:
                it = QTableWidgetItem(f"{val}合计")
            self.tbl.setHorizontalHeaderItem(base_col + i, it)
        
        # 设置总计列标题
        self.tbl.setHorizontalHeaderItem(headers_count - 1, QTableWidgetItem("Total"))

        # 增加行用于显示日期和总计行
        calc_type = self.calc_type_combo.currentText()
        if calc_type == "成品MRP":
            # 成品MRP：日期行 + 数据行 + 生产计划总计行 + 即时库存总计行
            self.tbl.setRowCount(len(rows) + 3)  # +1 for date row, +2 for total rows
        else:
            # 零部件MRP：日期行 + 数据行 + 总计行
            self.tbl.setRowCount(len(rows) + 2)  # +1 for date row, +1 for total row
        
        # 设置颜色
        green_bg = QBrush(QColor(235, 252, 239))  # 生产计划绿色
        red_bg = QBrush(QColor(255, 235, 238))     # 库存不足红色
        blue_bg = QBrush(QColor(221, 235, 247))   # 合计列蓝色
        date_bg = QBrush(QColor(248, 249, 250))   # 日期行的背景色

        # 第一行：显示CW对应的日期
        date_row = 0
        for c in range(base_col):  # 基本信息列
            it = self._set_item(date_row, c, "")
            it.setBackground(date_bg)
        
        for i, (kind, val) in enumerate(colspec):
            if kind == "week":
                # 将CW转换为对应的日期
                date_str = self._convert_cw_to_date(val)
                it = self._set_item(date_row, base_col + i, date_str)
                it.setBackground(date_bg)
                # 设置日期行的字体样式
                font = it.font()
                font.setPointSize(9)
                it.setFont(font)
            else:
                # 年份合计列显示年份
                it = self._set_item(date_row, base_col + i, str(val))
                it.setBackground(date_bg)
                font = it.font()
                font.setPointSize(9)
                it.setFont(font)
        
        # 总计列
        it = self._set_item(date_row, headers_count - 1, "")
        it.setBackground(date_bg)

        # 数据行（从第二行开始）
        for r, row in enumerate(rows):
            actual_row = r + 1  # 实际行号要+1，因为第一行是日期行
            
            # 基本信息列
            self._set_item(actual_row, 0, row.get("ItemName", ""))
            self._set_item(actual_row, 1, row.get("ItemSpec", ""))
            self._set_item(actual_row, 2, row.get("ItemType", ""))
            self._set_item(actual_row, 3, row.get("RowType", ""))
            
            # 期初库存列：综合MRP显示"XXX+XXX"格式，其他显示数字
            start_onhand = row.get("StartOnHand", 0)
            if isinstance(start_onhand, str) and "+" in start_onhand:
                # 综合MRP的"XXX+XXX"格式，直接显示
                self._set_item(actual_row, 4, start_onhand)
            else:
                # 其他类型，格式化为数字
                self._set_item(actual_row, 4, self._fmt(start_onhand))

            # 基本信息列不设置背景色

            # 周数据列和年份合计列
            row_total = 0
            cursor_col = base_col
            for kind, val in colspec:
                if kind == "week":
                    val_float = float(row["cells"].get(val, 0.0))
                    row_total += val_float
                    it = self._set_item(actual_row, cursor_col, self._fmt(val_float))
                    
                    # 新的着色规则：
                    # 1. 生产计划行（非即时库存）且数值大于0时标绿色
                    # 2. 即时库存行且数值小于0时标红色
                    is_stock_row = (row.get("RowType") == "即时库存")
                    if not is_stock_row and val_float > 0:
                        it.setBackground(green_bg)  # 生产计划标绿
                    elif is_stock_row and val_float < 0:
                        it.setBackground(red_bg)    # 库存不足标红
                else:
                    # 年份合计列
                    year_total = sum(float(row["cells"].get(w, 0.0)) for w in self._get_weeks_in_year(val))
                    it = QTableWidgetItem(self._fmt(year_total))
                    it.setBackground(blue_bg)  # 合计列标蓝色
                    font = it.font()
                    font.setBold(True)
                    it.setFont(font)
                    self.tbl.setItem(actual_row, cursor_col, it)
                    row_total += year_total
                
                cursor_col += 1

            # 总计列
            total_item = QTableWidgetItem(self._fmt(row_total))
            total_item.setBackground(blue_bg)  # 总计列标蓝色
            font = total_item.font()
            font.setBold(True)
            total_item.setFont(font)
            self.tbl.setItem(actual_row, headers_count - 1, total_item)

        # 总计行
        calc_type = self.calc_type_combo.currentText()
        if calc_type == "成品MRP":
            # 成品MRP：两行总计行
            # 第一行：生产计划总计
            plan_total_row = len(rows) + 1
            self.tbl.setItem(plan_total_row, 0, QTableWidgetItem("生产计划TOTAL"))
            self.tbl.setItem(plan_total_row, 1, QTableWidgetItem(""))
            self.tbl.setItem(plan_total_row, 2, QTableWidgetItem(""))
            self.tbl.setItem(plan_total_row, 3, QTableWidgetItem("生产计划"))
            self.tbl.setItem(plan_total_row, 4, QTableWidgetItem(""))
            
            # 第二行：即时库存总计
            stock_total_row = len(rows) + 2
            self.tbl.setItem(stock_total_row, 0, QTableWidgetItem("即时库存TOTAL"))
            self.tbl.setItem(stock_total_row, 1, QTableWidgetItem(""))
            self.tbl.setItem(stock_total_row, 2, QTableWidgetItem(""))
            self.tbl.setItem(stock_total_row, 3, QTableWidgetItem("即时库存"))
            self.tbl.setItem(stock_total_row, 4, QTableWidgetItem(""))
            
            # 计算生产计划总计（只统计生产计划行）
            for col in range(base_col, headers_count):
                plan_sum = 0
                for r in range(1, plan_total_row):  # 从1开始，跳过日期行
                    it = self.tbl.item(r, col)
                    row_type_it = self.tbl.item(r, 3)  # 行别列
                    try:
                        if it and it.text().strip() and row_type_it and row_type_it.text() == "生产计划":
                            plan_sum += float(it.text().replace(',', ''))
                    except:
                        pass
                item = QTableWidgetItem(self._fmt(plan_sum))
                font = item.font()
                font.setBold(True)
                item.setFont(font)
                item.setBackground(green_bg)  # 生产计划总计标绿色
                self.tbl.setItem(plan_total_row, col, item)
            
            # 计算即时库存总计（只统计即时库存行）
            for col in range(base_col, headers_count):
                stock_sum = 0
                for r in range(1, stock_total_row):  # 从1开始，跳过日期行
                    it = self.tbl.item(r, col)
                    row_type_it = self.tbl.item(r, 3)  # 行别列
                    try:
                        if it and it.text().strip() and row_type_it and row_type_it.text() == "即时库存":
                            stock_sum += float(it.text().replace(',', ''))
                    except:
                        pass
                item = QTableWidgetItem(self._fmt(stock_sum))
                font = item.font()
                font.setBold(True)
                item.setFont(font)
                item.setBackground(red_bg)  # 即时库存总计标红色
                self.tbl.setItem(stock_total_row, col, item)
        else:
            # 零部件MRP：一行总计行
            total_row = len(rows) + 1
            self.tbl.setItem(total_row, 0, QTableWidgetItem("TOTAL"))
            
            # 只从周列开始统计（前5列不算）
            for col in range(base_col, headers_count):
                s = 0
                for r in range(1, total_row):  # 从1开始，跳过日期行
                    it = self.tbl.item(r, col)
                    try:
                        if it and it.text().strip():
                            s += float(it.text().replace(',', ''))
                    except:
                        pass
                item = QTableWidgetItem(self._fmt(s))
                font = item.font()
                font.setBold(True)
                item.setFont(font)
                item.setBackground(blue_bg)  # 总计行标蓝色
                self.tbl.setItem(total_row, col, item)

        # 小优化：把计划/库存两行当作一个分组阅读
        if calc_type == "零部件MRP":
            self.tbl.setAlternatingRowColors(False)
        else:
            self.tbl.setAlternatingRowColors(True)

    def _set_item(self, r: int, c: int, text: str):
        it = QTableWidgetItem(str(text))
        it.setTextAlignment(Qt.AlignCenter)
        self.tbl.setItem(r, c, it)
        return it

    @staticmethod
    def _fmt(v: float) -> str:
        # 千分位，不带多余小数
        if abs(v - int(v)) < 1e-6:
            return f"{int(v):,}"
        return f"{v:,.3f}"

    def _build_week_columns_with_totals(self, weeks: list) -> list:
        """构建周列和年份合计列，参考客户订单处理页面的逻辑"""
        from collections import defaultdict
        from datetime import datetime, timedelta
        
        # 如果有当前数据，使用实际的订单日期来分组
        if hasattr(self, '_current_data') and self._current_data:
            # 获取订单版本ID
            import_id = self.order_version_combo.currentData()
            
            if import_id is not None:
                # 获取该订单版本的所有唯一订单日期
                from app.db import query_all
                sql = """
                SELECT DISTINCT col.DeliveryDate
                FROM CustomerOrderLines col
                JOIN CustomerOrders co ON col.OrderId = co.OrderId
                WHERE co.ImportId = ? AND col.LineStatus = 'Active' AND col.DeliveryDate IS NOT NULL
                ORDER BY col.DeliveryDate
                """
                rows = query_all(sql, (import_id,))
                
                # 按年份分组实际的订单日期
                by_year = defaultdict(list)
                for row in rows:
                    try:
                        date_obj = datetime.strptime(row["DeliveryDate"], "%Y-%m-%d").date()
                        by_year[date_obj.isocalendar()[0]].append(date_obj)
                    except:
                        continue
                
                # 对每年的日期排序
                for y in by_year:
                    by_year[y].sort()
                
                # 构建列规范
                colspec = []
                years = sorted(by_year.keys())
                for year in years:
                    # 为每年的每个日期创建列
                    for d in by_year[year]:
                        colspec.append(("week", f"CW{d.isocalendar()[1]:02d}"))
                    # 添加年份合计列
                    colspec.append(("total", year))
                
                return colspec
        
        # 如果没有订单数据，使用默认逻辑
        # 按年份分组
        by_year = defaultdict(list)
        
        # 获取开始日期和结束日期来确定年份范围
        start_date = self.dt_start.date()
        end_date = self.dt_end.date()
        
        # 从日期范围推断年份
        start_year = start_date.year()
        end_year = end_date.year()
        
        # 如果跨年，则按年份分组
        if start_year == end_year:
            # 同一年，所有周都归到这一年
            for week in weeks:
                by_year[start_year].append(week)
        else:
            # 跨年，需要根据CW的实际日期来分组
            # 这里简化处理，按CW的顺序分组
            # 假设前半部分属于开始年份，后半部分属于结束年份
            mid_point = len(weeks) // 2
            for i, week in enumerate(weeks):
                if i < mid_point:
                    by_year[start_year].append(week)
                else:
                    by_year[end_year].append(week)
        
        # 构建列规范
        colspec = []
        years = sorted(by_year.keys())
        for year in years:
            # 添加该年的所有周
            for week in by_year[year]:
                colspec.append(("week", week))
            # 添加年份合计列
            colspec.append(("total", year))
        
        return colspec

    def _get_weeks_in_year(self, year: int) -> list:
        """获取指定年份的所有周，参考客户订单处理页面的逻辑"""
        from collections import defaultdict
        
        # 如果有当前数据，使用实际的订单日期来分组
        if hasattr(self, '_current_data') and self._current_data:
            # 获取订单版本ID
            import_id = self.order_version_combo.currentData()
            
            if import_id is not None:
                # 获取该订单版本的所有唯一订单日期
                from app.db import query_all
                sql = """
                SELECT DISTINCT col.DeliveryDate
                FROM CustomerOrderLines col
                JOIN CustomerOrders co ON col.OrderId = co.OrderId
                WHERE co.ImportId = ? AND col.LineStatus = 'Active' AND col.DeliveryDate IS NOT NULL
                ORDER BY col.DeliveryDate
                """
                rows = query_all(sql, (import_id,))
                
                # 按年份分组实际的订单日期
                by_year = defaultdict(list)
                for row in rows:
                    try:
                        from datetime import datetime
                        date_obj = datetime.strptime(row["DeliveryDate"], "%Y-%m-%d").date()
                        by_year[date_obj.isocalendar()[0]].append(f"CW{date_obj.isocalendar()[1]:02d}")
                    except:
                        continue
                
                return by_year.get(year, [])
        
        # 如果没有订单数据，使用默认逻辑
        # 从当前数据中获取该年份的所有周
        if hasattr(self, '_current_data'):
            weeks = self._current_data.get("weeks", [])
            by_year = defaultdict(list)
            
            # 获取开始日期和结束日期来确定年份范围
            start_date = self.dt_start.date()
            end_date = self.dt_end.date()
            
            # 从日期范围推断年份
            start_year = start_date.year()
            end_year = end_date.year()
            
            # 如果跨年，则按年份分组
            if start_year == end_year:
                # 同一年，所有周都归到这一年
                for week in weeks:
                    by_year[start_year].append(week)
            else:
                # 跨年，需要根据CW的顺序分组
                mid_point = len(weeks) // 2
                for i, week in enumerate(weeks):
                    if i < mid_point:
                        by_year[start_year].append(week)
                    else:
                        by_year[end_year].append(week)
            
            return by_year.get(year, [])
        
        return []

    def _convert_cw_to_date(self, cw: str) -> str:
        """将CW格式转换为对应的日期字符串，基于实际的订单日期"""
        try:
            # 从CW中提取周数
            if cw.startswith("CW"):
                week_num = int(cw[2:])
                
                # 如果有当前数据，尝试从实际的订单日期中找到对应的日期
                if hasattr(self, '_current_data') and self._current_data:
                    # 获取订单版本ID
                    import_id = self.order_version_combo.currentData()
                    
                    if import_id is not None:
                        # 获取该订单版本的所有唯一订单日期
                        from app.db import query_all
                        sql = """
                        SELECT DISTINCT col.DeliveryDate
                        FROM CustomerOrderLines col
                        JOIN CustomerOrders co ON col.OrderId = co.OrderId
                        WHERE co.ImportId = ? AND col.LineStatus = 'Active' AND col.DeliveryDate IS NOT NULL
                        ORDER BY col.DeliveryDate
                        """
                        rows = query_all(sql, (import_id,))
                        
                        # 找到对应周数的日期
                        for row in rows:
                            try:
                                from datetime import datetime
                                date_obj = datetime.strptime(row["DeliveryDate"], "%Y-%m-%d").date()
                                if date_obj.isocalendar()[1] == week_num:
                                    return date_obj.strftime("%Y/%m/%d")
                            except:
                                continue
                
                # 如果没有找到对应的订单日期，使用默认计算
                # 获取开始日期和结束日期来确定年份范围
                start_date = self.dt_start.date()
                end_date = self.dt_end.date()
                
                # 从日期范围推断年份
                start_year = start_date.year()
                end_year = end_date.year()
                
                # 如果跨年，需要根据CW的实际日期来分组
                if start_year == end_year:
                    # 同一年，使用开始年份
                    target_year = start_year
                else:
                    # 跨年，需要根据CW的实际日期来分组
                    # 这里简化处理，按CW的顺序分组
                    # 假设前半部分属于开始年份，后半部分属于结束年份
                    weeks = self._current_data.get("weeks", []) if hasattr(self, '_current_data') else []
                    mid_point = len(weeks) // 2
                    cw_index = weeks.index(cw) if cw in weeks else 0
                    if cw_index < mid_point:
                        target_year = start_year
                    else:
                        target_year = end_year
                
                # 计算该年的第week_num周的第一天
                jan1 = QDate(target_year, 1, 1)
                
                # 找到该年第一个周一
                days_to_monday = (8 - jan1.dayOfWeek()) % 7
                if days_to_monday == 0:
                    days_to_monday = 7
                
                first_monday = jan1.addDays(days_to_monday - 1)
                
                # 计算目标周的第一天
                target_week_start = first_monday.addDays((week_num - 1) * 7)
                
                # 返回格式化的日期字符串
                return target_week_start.toString("yyyy/MM/dd")
            else:
                return cw
        except:
            return cw

    def on_export(self):
        """导出Excel文件"""
        if not hasattr(self, '_current_data') or not self._current_data:
            QMessageBox.warning(self, "提示", "请先生成看板数据")
            return
        
        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "导出Excel文件", 
            f"MRP看板_{QDate.currentDate().toString('yyyyMMdd')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            self.export_to_excel(file_path, self._current_data)
            QMessageBox.information(self, "导出成功", f"文件已保存到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误：\n{str(e)}")

    def export_to_excel(self, file_path: str, data: dict):
        """导出数据到Excel文件"""
        weeks = data.get("weeks", [])
        rows = data.get("rows", [])
        calc_type = self.calc_type_combo.currentText()
        
        # 构建年份分组和合计列
        colspec = self._build_week_columns_with_totals(weeks)
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"MRP看板_{calc_type}"
        
        # 定义颜色样式
        green_fill = PatternFill(start_color="E7F5E7", end_color="E7F5E7", fill_type="solid")
        red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        date_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # 定义字体样式 - 统一使用Arial字体
        header_font = Font(name="Arial", bold=True, size=12)
        date_font = Font(name="Arial", size=9)
        normal_font = Font(name="Arial", size=10)
        total_font = Font(name="Arial", bold=True, size=10)
        
        # 定义对齐方式
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # 定义边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置列标题
        if calc_type == "零部件MRP":
            fixed_headers = ["物料名称", "物料规格", "物料类型", "行别", "期初库存"]
        elif calc_type == "成品MRP":
            fixed_headers = ["物料名称", "物料规格", "成品类型", "行别", "期初库存"]
        else:  # 综合MRP
            fixed_headers = ["物料名称", "物料规格", "物料类型", "行别", "期初库存"]
        
        headers_count = len(fixed_headers) + len(colspec) + 1  # +1 for Total column
        
        # 写入固定列标题
        for col, header in enumerate(fixed_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # 写入周列和年份合计列标题
        base_col = len(fixed_headers)
        for i, (kind, val) in enumerate(colspec):
            col = base_col + i + 1
            if kind == "week":
                cell = ws.cell(row=1, column=col, value=val)
            else:
                cell = ws.cell(row=1, column=col, value=f"{val}合计")
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # 写入总计列标题
        total_col = headers_count
        cell = ws.cell(row=1, column=total_col, value="Total")
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # 写入日期行（第二行）
        row_num = 2
        for col in range(1, base_col + 1):  # 基本信息列
            cell = ws.cell(row=row_num, column=col, value="")
            cell.fill = date_fill
            cell.border = thin_border
        
        for i, (kind, val) in enumerate(colspec):
            col = base_col + i + 1
            if kind == "week":
                date_str = self._convert_cw_to_date(val)
                cell = ws.cell(row=row_num, column=col, value=date_str)
            else:
                cell = ws.cell(row=row_num, column=col, value=str(val))
            cell.font = date_font
            cell.alignment = center_alignment
            cell.fill = date_fill
            cell.border = thin_border
        
        # 总计列
        cell = ws.cell(row=row_num, column=total_col, value="")
        cell.fill = date_fill
        cell.border = thin_border
        
        # 写入数据行
        for row_data in rows:
            row_num += 1
            
            # 基本信息列
            start_onhand = row_data.get("StartOnHand", 0)
            if isinstance(start_onhand, str) and "+" in start_onhand:
                # 综合MRP的"XXX+XXX"格式，直接显示
                start_onhand_display = start_onhand
            else:
                # 其他类型，格式化为数字
                start_onhand_display = self._fmt(start_onhand)
            
            basic_info = [
                row_data.get("ItemName", ""),
                row_data.get("ItemSpec", ""),
                row_data.get("ItemType", ""),
                row_data.get("RowType", ""),
                start_onhand_display
            ]
            
            for col, value in enumerate(basic_info, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border
                # 基本信息列不设置背景色
            
            # 周数据列和年份合计列
            row_total = 0
            for i, (kind, val) in enumerate(colspec):
                col = base_col + i + 1
                if kind == "week":
                    val_float = float(row_data["cells"].get(val, 0.0))
                    row_total += val_float
                    cell = ws.cell(row=row_num, column=col, value=val_float)
                    cell.font = normal_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    # 新的着色规则：
                    # 1. 生产计划行（非即时库存）且数值大于0时标绿色
                    # 2. 即时库存行且数值小于0时标红色
                    is_stock_row = (row_data.get("RowType") == "即时库存")
                    if not is_stock_row and val_float > 0:
                        cell.fill = green_fill  # 生产计划标绿
                    elif is_stock_row and val_float < 0:
                        cell.fill = red_fill    # 库存不足标红
                else:
                    # 年份合计列
                    year_total = sum(float(row_data["cells"].get(w, 0.0)) for w in self._get_weeks_in_year(val))
                    row_total += year_total
                    cell = ws.cell(row=row_num, column=col, value=year_total)
                    cell.font = total_font
                    cell.alignment = center_alignment
                    cell.fill = blue_fill  # 合计列标蓝色
                    cell.border = thin_border
            
            # 总计列
            cell = ws.cell(row=row_num, column=total_col, value=row_total)
            cell.font = total_font
            cell.alignment = center_alignment
            cell.fill = blue_fill  # 总计列标蓝色
            cell.border = thin_border
        
        # 总计行
        total_row = row_num + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        
        # 只从周列开始统计（前5列不算）
        for col in range(base_col + 1, total_col + 1):
            s = 0
            for r in range(3, total_row):  # 从第3行开始，跳过标题行和日期行
                cell = ws.cell(row=r, column=col)
                if cell.value is not None:
                    try:
                        s += float(cell.value)
                    except:
                        pass
            cell = ws.cell(row=total_row, column=col, value=s)
            cell.font = total_font
            cell.alignment = center_alignment
            cell.fill = blue_fill  # 总计行标蓝色
            cell.border = thin_border
        
        # 调整列宽
        for col in range(1, headers_count + 1):
            if col <= base_col:  # 基本信息列
                ws.column_dimensions[get_column_letter(col)].width = 15
            else:  # 周数据列和合计列
                ws.column_dimensions[get_column_letter(col)].width = 12
        
        # 保存文件
        wb.save(file_path)
        wb.close()
