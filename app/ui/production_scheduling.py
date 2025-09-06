#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from collections import defaultdict
from typing import Optional, List, Dict, Tuple
from datetime import datetime, timedelta, date as _date

from PySide6.QtCore import Qt, QDate, Signal, QThread, QRect
from PySide6.QtGui import QFont, QColor, QPainter
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QFrame, QLineEdit, QComboBox, QAbstractItemView,
    QMessageBox, QTabWidget, QFrame, QScrollArea, QSpinBox,
    QGroupBox, QGridLayout, QSplitter, QCheckBox, QProgressBar, QDialog,
    QHeaderView, QDateEdit, QFileDialog, QProgressBar, QTextBrowser, QAbstractScrollArea, QSizePolicy
)

from app.services.production_scheduling_service import ProductionSchedulingService
from app.services.customer_order_service import CustomerOrderService

# Excel导出相关导入
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -------------------- 小工具 --------------------
def _safe_parse_date(s):
    if not s:
        return None
    if isinstance(s, _date):
        return s
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%y", "%m/%d/%Y"):
        try:
            d = datetime.strptime(str(s), fmt).date()
            if d.year < 2000:
                d = d.replace(year=d.year + 2000)
            return d
        except Exception:
            pass
    return None

def _cw(d: _date) -> int:
    return d.isocalendar()[1]

def _norm_int(v, default=0):
    try:
        if v in (None, ""):
            return default
        return int(float(v))
    except Exception:
        return default

def _get(line: dict, *names, default=None):
    for n in names:
        if n in line and line[n] not in (None, ""):
            return line[n]
    return default

def _week_start(d: _date) -> _date:
    return d - timedelta(days=d.weekday())

def _build_week_cols(min_d: _date, max_d: _date, order_dates: List[_date] = None) -> Tuple[List[Tuple[str, _date|int]], Dict[int, List[_date]], List[int]]:
    """根据起止日期构造连续周列以及每年的周分组，并在每年后追加"合计"列。"""
    if not (min_d and max_d):
        return [], {}, []
    
    # 如果有订单日期，使用订单中的唯一日期；否则使用日期范围内的所有日期
    if order_dates:
        dates = sorted(set(order_dates))  # 去重并排序
    else:
        dates = []
        cur = min_d
        while cur <= max_d:
            dates.append(cur)
            cur += timedelta(days=1)
    
    # 按年分组
    by_year = defaultdict(list)
    for d in dates:
        by_year[d.isocalendar()[0]].append(d)
    
    # 对每年的日期排序
    for y in by_year:
        by_year[y].sort()
    
    years = sorted(by_year.keys())
    colspec: List[Tuple[str, _date|int]] = []
    
    for y in years:
        # 为每年的每个日期创建列
        for d in by_year[y]:
            colspec.append(("date", d))
        colspec.append(("sum", y))
    
    return colspec, by_year, years


# -------------------- 两行表头 --------------------
class TwoRowHeader(QHeaderView):
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setDefaultAlignment(Qt.AlignCenter)
        self._top_font = QFont(); self._top_font.setBold(True)
        self._bottom_font = QFont()
        self._bottom_font.setPointSize(self._bottom_font.pointSize() - 1.5)  # 日期字体更小
        self._bottom_font.setBold(True)  # 日期字体加粗
        h = self.fontMetrics().height()
        self.setFixedHeight(int(h * 3.0))  # 进一步增加高度

    def sizeHint(self):
        s = super().sizeHint()
        h = self.fontMetrics().height()
        s.setHeight(int(h * 3.0))  # 与setFixedHeight保持一致
        return s

    def paintSection(self, painter: QPainter, rect: QRect, logicalIndex: int):
        if not rect.isValid():
            return
        
        # 完全自定义绘制，不使用父类方法
        painter.save()
        
        # 绘制背景
        painter.fillRect(rect, QColor("#fafafa"))
        
        # 绘制边框
        painter.setPen(QColor("#d9d9d9"))
        painter.drawRect(rect)
        
        table = self.parent()
        item = table.horizontalHeaderItem(logicalIndex) if table else None
        top = item.text() if item else ""
        bottom = item.data(Qt.UserRole) if (item and item.data(Qt.UserRole) is not None) else ""

        # 调试信息
        if bottom and "CW" in top:
            print(f"DEBUG: 绘制表头 - 列{logicalIndex}: '{top}' / '{bottom}', 矩形: {rect}")

        # 计算两行的矩形区域
        top_height = rect.height() // 2
        bottom_height = rect.height() - top_height
        
        # 绘制第一行（CW编号）
        painter.setPen(QColor("#333333"))
        painter.setFont(self._top_font)
        topRect = QRect(rect.left() + 1, rect.top() + 1, rect.width() - 2, top_height - 2)
        painter.drawText(topRect, Qt.AlignCenter, str(top))
        
        # 绘制第二行（日期）
        if bottom:  # 只有当有日期数据时才绘制
            painter.setFont(self._bottom_font)
            painter.setPen(QColor("#666666"))  # 使用稍浅的颜色
            
            # 为日期预留更多边距，确保不超出边界
            margin = 1  # 增加边距
            bottomRect = QRect(
                rect.left() + margin, 
                rect.top() + top_height, 
                rect.width() - margin * 2, 
                bottom_height - margin
            )
            
            # 使用Qt.TextWrapAnywhere确保文本不会超出边界
            painter.drawText(bottomRect, Qt.AlignCenter | Qt.TextWrapAnywhere, str(bottom))
            print(f"DEBUG: 绘制第二行 - 矩形: {bottomRect}, 文本: '{bottom}'")
        
        painter.restore()


# -------------------- 主界面 --------------------
class ProductionSchedulingWidget(QWidget):
    """生产排产管理主界面 - 完全复刻客户订单看板"""
    
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_version_list()

    def init_ui(self):
        self.setWindowTitle("生产排产管理")
        
        # 设置大小策略，让页面适应父容器
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        main_layout = QVBoxLayout()

        # 页签
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet(
            "QTabWidget::pane{border:1px solid #dee2e6;background:white;}"
            "QTabBar::tab{background:#f8f9fa;border:1px solid #dee2e6;padding:8px 16px;margin-right:2px;}"
            "QTabBar::tab:selected{background:white;border-bottom:2px solid #007bff;}"
        )

        self.create_kanban_tab()

        main_layout.addWidget(self.tab_widget)
        self.setLayout(main_layout)

    # ---------- 看板页 ----------
    def create_kanban_tab(self):
        kanban_widget = QWidget()
        kanban_layout = QVBoxLayout()

        control_panel = QFrame()
        control_panel.setFrameStyle(QFrame.StyledPanel)
        control_panel.setStyleSheet("QFrame{background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;}")
        control_layout = QVBoxLayout()

        # 版本选择 + 日期范围
        version_layout = QHBoxLayout()
        version_layout.addWidget(QLabel("订单版本:"))
        self.version_combo = QComboBox()
        self.version_combo.addItem("全部版本汇总")
        self.version_combo.currentTextChanged.connect(self.on_version_changed)
        self.version_combo.setMinimumWidth(200)  # 设置最小宽度
        self.version_combo.setMinimumHeight(12)   # 设置最小高度
        self.version_combo.setStyleSheet("""
            QComboBox {
                padding: 6px 12px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background: white;
                min-width: 200px;
                max-width: 300px;
                min-height: 12px;
                max-height: 12px;
            }
            QComboBox:focus {
                border-color: #007bff;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #6c757d;
                margin-right: 5px;
            }
        """)
        version_layout.addWidget(self.version_combo)

        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("日期范围:"))
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDate(QDate.currentDate().addDays(-30))
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        date_layout.addWidget(self.start_date_edit)

        date_layout.addWidget(QLabel("至"))
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDate(QDate.currentDate().addDays(30))
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        date_layout.addWidget(self.end_date_edit)

        apply_btn = QPushButton("应用筛选")
        apply_btn.setStyleSheet("QPushButton{background:#007bff;color:#fff;border:none;padding:6px 12px;border-radius:4px;}"
                                "QPushButton:hover{background:#0069d9;}")
        apply_btn.clicked.connect(self.load_kanban_data)

        export_btn = QPushButton("导出Excel")
        export_btn.setStyleSheet("QPushButton{background:#28a745;color:#fff;border:none;padding:6px 12px;border-radius:4px;}"
                                "QPushButton:hover{background:#218838;}")
        export_btn.clicked.connect(self.export_kanban_to_excel)

        version_layout.addLayout(date_layout)
        version_layout.addWidget(apply_btn)
        version_layout.addWidget(export_btn)
        version_layout.addStretch()
        control_layout.addLayout(version_layout)

        # 订单类型过滤
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("订单类型:"))
        self.order_type_combo = QComboBox()
        self.order_type_combo.addItems(["全部", "F(正式)", "P(预测)"])
        self.order_type_combo.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.order_type_combo)
        filter_layout.addStretch()
        control_layout.addLayout(filter_layout)

        # 排产模式控制
        scheduling_layout = QHBoxLayout()
        scheduling_layout.addWidget(QLabel("排产模式:"))
        
        self.enter_scheduling_btn = QPushButton("进入排产模式")
        self.enter_scheduling_btn.setStyleSheet("QPushButton{background:#ffc107;color:#000;border:none;padding:6px 12px;border-radius:4px;}"
                                               "QPushButton:hover{background:#e0a800;}")
        self.enter_scheduling_btn.clicked.connect(self.enter_scheduling_mode)
        self.enter_scheduling_btn.setEnabled(False)
        
        self.confirm_scheduling_btn = QPushButton("确定排产")
        self.confirm_scheduling_btn.setStyleSheet("QPushButton{background:#28a745;color:#fff;border:none;padding:6px 12px;border-radius:4px;}"
                                                 "QPushButton:hover{background:#218838;}")
        self.confirm_scheduling_btn.clicked.connect(self.confirm_scheduling)
        self.confirm_scheduling_btn.setVisible(False)
        
        self.back_to_order_btn = QPushButton("返回订单视图")
        self.back_to_order_btn.setStyleSheet("QPushButton{background:#6c757d;color:#fff;border:none;padding:6px 12px;border-radius:4px;}"
                                             "QPushButton:hover{background:#5a6268;}")
        self.back_to_order_btn.clicked.connect(self.back_to_order_view)
        self.back_to_order_btn.setVisible(False)
        
        scheduling_layout.addWidget(self.enter_scheduling_btn)
        scheduling_layout.addWidget(self.confirm_scheduling_btn)
        scheduling_layout.addWidget(self.back_to_order_btn)
        scheduling_layout.addStretch()
        control_layout.addLayout(scheduling_layout)

        control_panel.setLayout(control_layout)
        kanban_layout.addWidget(control_panel)

        # 看板表格
        self.kanban_table = QTableWidget()
        self.kanban_table.setAlternatingRowColors(True)
        self.kanban_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kanban_table.setHorizontalHeader(TwoRowHeader(Qt.Horizontal, self.kanban_table))
        hdr = self.kanban_table.horizontalHeader()
        hdr.setFixedHeight(int(self.fontMetrics().height()*3.0))  # 与TwoRowHeader保持一致
        try:
            hdr.repaint()
            hdr.updateGeometry()
        except Exception:
            pass
        try:
            policy = QAbstractScrollArea.SizeAdjustPolicy.AdjustToContentsOnFirstShow
        except AttributeError:
            policy = getattr(QAbstractScrollArea, "AdjustToContentsOnFirstShow", QAbstractScrollArea.AdjustToContents)
        self.kanban_table.setSizeAdjustPolicy(policy)
        self.kanban_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # 设置表格的最小高度，确保在客户端范围内显示
        self.kanban_table.setMinimumHeight(300)

        kanban_layout.addWidget(self.kanban_table)
        kanban_widget.setLayout(kanban_layout)
        self.tab_widget.addTab(kanban_widget, "看板视图")

    # ------- 数据加载 -------
    def load_version_list(self):
        """加载版本列表"""
        try:
            versions = CustomerOrderService.get_import_history()
            
            self.version_combo.clear()
            self.version_combo.addItem("全部版本汇总")
            for version in versions:
                version_text = f"{version['ImportId']} - {version['FileName']} ({version['ImportDate']})"
                self.version_combo.addItem(version_text, version['ImportId'])
                
        except Exception as e:
            print(f"加载版本列表失败: {e}")

    def on_version_changed(self, version_text: str):
        try:
            if version_text == "全部版本汇总":
                self.load_kanban_data()
                return
            import re
            m = re.search(r'^(\d+) - ', version_text)
            if not m:
                self.load_kanban_data()
                return
            version_id = int(m.group(1))
            data = CustomerOrderService.get_order_lines_by_import_version(version_id)
            all_dates = []
            for ln in data or []:
                d = _safe_parse_date(_get(ln, "DeliveryDate", "delivery_date", "DueDate", "due_date", ""))
                if d:
                    all_dates.append(d)
            if all_dates:
                all_dates.sort()
                self.start_date_edit.setDate(QDate(all_dates[0].year, all_dates[0].month, all_dates[0].day))
                self.end_date_edit.setDate(QDate(all_dates[-1].year, all_dates[-1].month, all_dates[-1].day))
            self.load_kanban_data()
        except Exception as e:
            print(f"版本切换失败: {e}")

    def on_filter_changed(self):
        self.load_kanban_data()

    def load_kanban_data(self):
        try:
            version_text = self.version_combo.currentText()
            sd = self.start_date_edit.date().toString("yyyy-MM-dd")
            ed = self.end_date_edit.date().toString("yyyy-MM-dd")
            if version_text != "全部版本汇总":
                import re
                m = re.search(r'^(\d+) - ', version_text)
                version_id = int(m.group(1)) if m else None
            else:
                version_id = None

            if version_id:
                data = CustomerOrderService.get_order_lines_by_import_version(version_id)
                self.display_kanban_data_by_version(data, sd, ed, self.order_type_combo.currentText())
            else:
                # 汇总口径：把 Firm/Predict 拆成伪明细
                rows = CustomerOrderService.get_ndlutil_kanban_data(start_date=sd, end_date=ed)
                expanded = []
                for r in rows or []:
                    common = {
                        "SupplierCode": r.get("SupplierCode",""),
                        "SupplierName": r.get("SupplierName",""),
                        "ItemNumber":   r.get("ItemNumber",""),
                        "ItemDescription": r.get("ItemDescription",""),
                        "DeliveryDate": r.get("DeliveryDate",""),
                        "CalendarWeek": r.get("CalendarWeek",""),
                        "ReleaseDate":  r.get("ReleaseDate",""),
                        "ReleaseId":    r.get("ReleaseId",""),
                        "PurchaseOrder":r.get("PurchaseOrder",""),
                        "ReceiptQuantity": 0,
                        "CumReceived":  0,
                    }
                    f = _norm_int(r.get("FirmQty"), 0)
                    p = _norm_int(r.get("ForecastQty"), 0)
                    if f:
                        e = dict(common); e["OrderType"]="F"; e["RequiredQty"]=f; expanded.append(e)
                    if p:
                        e = dict(common); e["OrderType"]="P"; e["RequiredQty"]=p; expanded.append(e)
                self.display_kanban_data_by_version(expanded, sd, ed, self.order_type_combo.currentText())
            
            # 启用进入排产模式按钮
            self.enter_scheduling_btn.setEnabled(True)
            
        except Exception as e:
            print(f"加载看板数据失败: {e}")

    # ------- 渲染（行集合固定，列按范围变动）-------
    def display_kanban_data_by_version(self, data: list, start_date: str, end_date: str, order_type: str):
        """
        行集合 = 该版本中检索到的所有 (Supplier, PN)，固定展示；
        列（CW/年合计）随页面日期范围变化；
        数量来自【按日期范围 + 类型过滤】后的数据进行周聚合。
        """
        sd = _safe_parse_date(start_date)
        ed = _safe_parse_date(end_date)
        if sd and ed and sd > ed:
            sd, ed = ed, sd
        page_typ = (order_type or "").upper()

        # 1) 收集所有订单日期
        order_dates = []
        for ln in (data or []):
            d = _safe_parse_date(_get(ln, "DeliveryDate", "delivery_date", "DueDate", "due_date", ""))
            if d:
                order_dates.append(d)
        
        # 2) 构造列
        colspec, by_year, years = _build_week_cols(sd, ed, order_dates)

        # 3) 行集合：整版数据收集 (Supplier, PN)
        groups_all = defaultdict(lambda: {"release": {}})
        for ln in (data or []):
            sup = _get(ln, "SupplierCode", "supplier_code", "Supplier", "supplier", "") or ""
            pn = _get(ln, "ItemNumber", "item_number", "Item", "item", "") or ""
            key = (sup, pn)
            ri = groups_all[key]["release"]

            def _set_if_val(k, v, cast_int=False):
                if v in (None, ""): return
                if cast_int: v = _norm_int(v, 0)
                if not ri.get(k): ri[k] = v

            _set_if_val("release_date", _get(ln, "ReleaseDate", "release_date", "release_date_text"))
            _set_if_val("release_id", _get(ln, "ReleaseId", "release_id"))
            _set_if_val("purchase_order", _get(ln, "PurchaseOrder", "OrderNumber", "purchase_order", "order_number"))
            _set_if_val("receipt_qty", _get(ln, "ReceiptQuantity", "receipt_qty"), cast_int=True)
            _set_if_val("cum_received", _get(ln, "CumReceived", "cum_received"), cast_int=True)

        # 4) 过滤后的数据
        lines_filtered = []
        for ln in (data or []):
            d = _safe_parse_date(_get(ln, "DeliveryDate", "delivery_date", "DueDate", "due_date", ""))
            if not d: continue
            if sd and d < sd: continue
            if ed and d > ed: continue
            typ = (_get(ln, "OrderType", "order_type", "FP", default="") or "").upper()
            if page_typ in ("F(正式)", "F") and typ != "F": continue
            if page_typ in ("P(预测)", "P") and typ != "P": continue
            ln = dict(ln)
            ln["__week__"] = _week_start(d)
            ln["__fp__"] = "F" if typ == "F" else "P"
            lines_filtered.append(ln)

        date_qty = defaultdict(lambda: defaultdict(int))
        fp_map = {}
        for ln in lines_filtered:
            sup = _get(ln, "SupplierCode", "supplier_code", "Supplier", "supplier", "")
            pn = _get(ln, "ItemNumber", "item_number", "Item", "item", "")
            # 使用已经解析过的日期
            d = _safe_parse_date(_get(ln, "DeliveryDate", "delivery_date", "DueDate", "due_date", ""))
            if not d:
                continue
            q = _norm_int(_get(ln, "RequiredQty", "req_qty", default=0), 0)
            date_qty[(sup, pn)][d] += q
            cur = fp_map.get((sup, pn, d))
            if (cur is None) or (cur == "P" and ln["__fp__"] == "F"):
                fp_map[(sup, pn, d)] = ln["__fp__"]

        # 5) 表头
        fixed_headers = [
            "Release Date", "Release ID", "成品名称", "成品规格", "成品品牌型号", "成品Project", "Item",
            "Purchase Order", "Receipt Quantity", "Cum Received"
        ]
        headers_count = len(fixed_headers) + len(colspec) + 1
        self.kanban_table.clear()
        self.kanban_table.setColumnCount(headers_count)

        for i, title in enumerate(fixed_headers):
            item = QTableWidgetItem(title)
            self.kanban_table.setHorizontalHeaderItem(i, item)

        base_col = len(fixed_headers)
        self.cw_checkboxes = {}  # 存储CW列的复选框
        for i, (kind, val) in enumerate(colspec):
            if kind == "date":
                cw = f"CW{val.isocalendar()[1]:02d}"
                date_str = val.strftime("%Y/%m/%d")
                it = QTableWidgetItem(cw)
                it.setData(Qt.UserRole, date_str)
            else:
                it = QTableWidgetItem(f"{val}合计")
            self.kanban_table.setHorizontalHeaderItem(base_col + i, it)
        self.kanban_table.setHorizontalHeaderItem(headers_count - 1, QTableWidgetItem("Total"))
        
        # 强制刷新表头显示
        header = self.kanban_table.horizontalHeader()
        header.updateGeometry()
        header.repaint()

        # 6) 行集合（按项目映射表的DisplayOrder排序）
        def sort_key(item):
            sup, pn = item
            if not pn:
                return (999999, 999, sup)  # 空PN排最后
            
            try:
                from app.services.project_service import ProjectService
                
                # 先尝试完整型号匹配
                project_code = ProjectService.get_project_by_item_brand(pn)
                if project_code:
                    mappings = ProjectService.get_project_mappings_by_project_code(project_code)
                    if mappings:
                        display_order = mappings[0].get('DisplayOrder', 999999)
                        return (display_order, sup, pn)
                
                # 去掉最后一位字母后缀，获取基础产品型号
                if len(pn) > 1 and pn[-1].isalpha():
                    base_pn = pn[:-1]  # 去掉最后一位字母
                else:
                    base_pn = pn
                
                # 尝试基础型号匹配
                project_code = ProjectService.get_project_by_item_brand(base_pn)
                if project_code:
                    mappings = ProjectService.get_project_mappings_by_project_code(project_code)
                    if mappings:
                        display_order = mappings[0].get('DisplayOrder', 999999)
                        return (display_order, sup, pn)
                
                # 如果完整匹配失败，尝试基础项目匹配
                if len(base_pn) > 1 and base_pn[-1].isdigit():
                    base = base_pn[:-1]  # 去掉最后一位数字
                else:
                    base = base_pn
                
                project_code = ProjectService.get_project_by_item_brand(base)
                if project_code:
                    mappings = ProjectService.get_project_mappings_by_project_code(project_code)
                    if mappings:
                        display_order = mappings[0].get('DisplayOrder', 999999)
                        return (display_order, sup, pn)
                
                # 如果都没有匹配到，使用硬编码的备用排序
                DESIRED_PN_ORDER = [
                    "R001H368E","R001H369E","R001P320B","R001P313B",
                    "R001J139B","R001J140B","R001J141B","R001J142B"
                ]
                
                if pn in DESIRED_PN_ORDER:
                    priority = DESIRED_PN_ORDER.index(pn) + 1000  # 给硬编码的排序一个较高的优先级
                else:
                    priority = 999999  # 不在列表中的排最后
                
                return (priority, sup, pn)
                
            except Exception as e:
                print(f"❌ [sort_key] 排序失败: {str(e)}")
                return (999999, sup, pn)
        
        keys_all = sorted(groups_all.keys(), key=sort_key)
        
        data_rows = len(keys_all)
        self.kanban_table.setRowCount(data_rows + 1)  # +1 行留给 TOTAL

        def project_name(pn: str) -> str:
            """根据产品型号获取项目名称，使用项目映射表"""
            if not pn:
                return "UNKNOWN"
            
            try:
                from app.services.project_service import ProjectService
                
                # 先尝试完整型号匹配
                project_code = ProjectService.get_project_by_item_brand(pn)
                if project_code:
                    # 获取项目名称
                    mappings = ProjectService.get_project_mappings_by_project_code(project_code)
                    if mappings:
                        return mappings[0].get('ProjectName', project_code)
                
                # 去掉最后一位字母后缀，获取基础产品型号
                if len(pn) > 1 and pn[-1].isalpha():
                    base_pn = pn[:-1]  # 去掉最后一位字母
                else:
                    base_pn = pn
                
                # 尝试基础型号匹配
                project_code = ProjectService.get_project_by_item_brand(base_pn)
                if project_code:
                    # 获取项目名称
                    mappings = ProjectService.get_project_mappings_by_project_code(project_code)
                    if mappings:
                        return mappings[0].get('ProjectName', project_code)
                
                # 如果完整匹配失败，尝试基础项目匹配
                if len(base_pn) > 1 and base_pn[-1].isdigit():
                    base = base_pn[:-1]  # 去掉最后一位数字
                else:
                    base = base_pn
                
                project_code = ProjectService.get_project_by_item_brand(base)
                if project_code:
                    # 获取项目名称
                    mappings = ProjectService.get_project_mappings_by_project_code(project_code)
                    if mappings:
                        return mappings[0].get('ProjectName', project_code)
                
                # 如果都没有匹配到，使用硬编码的备用映射
                base_map = {
                    "R001H368": "Passat rear double",
                    "R001H369": "Passat rear single",
                    "R001P320": "Tiguan L rear double",
                    "R001P313": "Tiguan L rear single",
                    "R001J139": "A5L rear double",
                    "R001J140": "A5L rear single",
                    "R001J141": "Lavida rear double",
                    "R001J142": "Lavida rear single"
                }
                
                project = base_map.get(base, "UNKNOWN")
                if project == "UNKNOWN":
                    print(f"警告：产品型号 '{pn}' 没有匹配到项目，默认放到最后")
                return project
                
            except Exception as e:
                print(f"❌ [project_name] 获取项目名称失败: {str(e)}")
                return "UNKNOWN"

        # 7) 填充数据行
        for row_idx, (sup, pn) in enumerate(keys_all):
            ri = groups_all[(sup, pn)]["release"]
            rd_obj = _safe_parse_date(ri.get("release_date"))
            rd_txt = rd_obj.strftime("%Y/%m/%d") if rd_obj else (ri.get("release_date") or "")

            project_result = project_name(pn)
            print(f"🔍 [客户订单视图] 处理第{row_idx}行，PN值: '{pn}'")
            
            # 根据PN字段获取成品信息
            from app.services.production_scheduling_service import ProductionSchedulingService
            product_info = ProductionSchedulingService.get_product_info_by_pn(pn)
            
            if product_info:
                product_name = product_info.get('CnName', '')
                product_spec = product_info.get('ItemSpec', '')
                product_brand = product_info.get('Brand', '')
                product_project = product_info.get('ProjectName', '') or product_info.get('ProjectCode', '')
            else:
                product_name = pn
                product_spec = "PEMM ASSY"
                product_brand = ""
                product_project = project_result

            fixed_vals = [
                rd_txt, str(ri.get("release_id", "") or ""), product_name,
                product_spec, product_brand, product_project, "Gross Reqs",
                sup, str(_norm_int(ri.get("receipt_qty"), 0)),
                str(_norm_int(ri.get("cum_received"), 0)),
            ]
            for c, v in enumerate(fixed_vals):
                self.kanban_table.setItem(row_idx, c, QTableWidgetItem(v))

            row_total = 0
            cursor_col = base_col
            for kind, val in colspec:
                if kind == "date":
                    qty = _norm_int(date_qty[(sup, pn)].get(val, 0), 0)
                    row_total += qty
                    cell = QTableWidgetItem(str(qty))
                    fp = fp_map.get((sup, pn, val))
                    if qty > 0 and fp:
                        cell.setBackground(QColor("#C6E0B4") if fp == "F" else QColor("#FFF2CC"))
                    self.kanban_table.setItem(row_idx, cursor_col, cell)
                else:
                    s = sum(_norm_int(date_qty[(sup, pn)].get(d, 0), 0) for d in by_year[val])
                    cell = QTableWidgetItem(str(s))
                    f = cell.font();
                    f.setBold(True);
                    cell.setFont(f)
                    cell.setBackground(QColor("#DDEBF7"))
                    self.kanban_table.setItem(row_idx, cursor_col, cell)
                cursor_col += 1

            self.kanban_table.setItem(row_idx, headers_count - 1, QTableWidgetItem(str(row_total)))

        # 8) TOTAL 行
        total_row = data_rows
        self.kanban_table.setItem(total_row, 0, QTableWidgetItem("TOTAL"))
        # 只从 CW 开始统计（前9列不算）
        for col in range(base_col, headers_count):
            s = 0
            for r in range(0, data_rows):
                it = self.kanban_table.item(r, col)
                try:
                    s += int(float(it.text())) if it and it.text().strip() else 0
                except:
                    pass
            item = QTableWidgetItem(str(s))
            f = item.font();
            f.setBold(True);
            item.setFont(f)
            self.kanban_table.setItem(total_row, col, item)

        # 列宽
        hdr = self.kanban_table.horizontalHeader()
        for i in range(9):
            hdr.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        for c in range(9, headers_count):
            hdr.setSectionResizeMode(c, QHeaderView.Fixed)
            self.kanban_table.setColumnWidth(c, 64)

        self.kanban_table.resizeRowsToContents()

    def export_kanban_to_excel(self):
        """导出看板到Excel"""
        try:
            # 获取保存路径
            file_path, _ = QFileDialog.getSaveFileName(
                self, "导出Excel文件", "生产排产看板.xlsx", "Excel文件 (*.xlsx)"
            )
            if not file_path:
                return
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "生产排产看板"
            
            # 获取表格数据
            row_count = self.kanban_table.rowCount()
            col_count = self.kanban_table.columnCount()
            
            # 写入表头
            for col in range(col_count):
                header_item = self.kanban_table.horizontalHeaderItem(col)
                if header_item:
                    ws.cell(row=1, column=col+1, value=header_item.text())
            
            # 写入数据
            for row in range(row_count):
                for col in range(col_count):
                    item = self.kanban_table.item(row, col)
                    if item:
                        ws.cell(row=row+2, column=col+1, value=item.text())
            
            # 保存文件
            wb.save(file_path)
            QMessageBox.information(self, "成功", f"Excel文件已保存到: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出Excel失败: {str(e)}")

    def enter_scheduling_mode(self):
        """进入排产模式"""
        try:
            # 显示复选框
            self.show_cw_checkboxes()
            
            # 更新按钮状态
            self.enter_scheduling_btn.setVisible(False)
            self.confirm_scheduling_btn.setVisible(True)
            self.back_to_order_btn.setVisible(True)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"进入排产模式失败: {str(e)}")

    def show_cw_checkboxes(self):
        """在CW列上方显示复选框"""
        try:
            # 在表格上方添加一行复选框
            current_rows = self.kanban_table.rowCount()
            self.kanban_table.insertRow(0)  # 在第一行插入复选框行
            
            # 获取CW列的起始位置
            fixed_headers = [
                "Release Date", "Release ID", "PN", "Des", "Project", "Item",
                "Purchase Order", "Receipt Quantity", "Cum Received"
            ]
            base_col = len(fixed_headers)
            
            # 为CW列添加复选框
            col_count = self.kanban_table.columnCount()
            for col in range(base_col, col_count - 1):  # 排除Total列
                header_item = self.kanban_table.horizontalHeaderItem(col)
                if header_item and "CW" in header_item.text():
                    # 创建复选框
                    checkbox = QCheckBox()
                    checkbox.setText("选择")
                    checkbox.setStyleSheet("""
                        QCheckBox {
                            font-size: 10px;
                            padding: 2px;
                        }
                    """)
                    
                    # 将复选框添加到表格中
                    self.kanban_table.setCellWidget(0, col, checkbox)
                    
                    # 存储复选框引用
                    self.cw_checkboxes[col] = checkbox
            
            # 设置复选框行的其他列为空
            for col in range(base_col):
                self.kanban_table.setItem(0, col, QTableWidgetItem(""))
            
            # 调整行高
            self.kanban_table.resizeRowsToContents()
            
        except Exception as e:
            print(f"显示CW复选框失败: {e}")

    def confirm_scheduling(self):
        """确定排产"""
        try:
            # 获取选中的CW列
            selected_cw_cols = []
            for col, checkbox in self.cw_checkboxes.items():
                if checkbox.isChecked():
                    selected_cw_cols.append(col)
            
            if not selected_cw_cols:
                QMessageBox.warning(self, "警告", "请至少选择一个CW列进行排产")
                return
            
            # 切换到排产看板视图
            self.switch_to_scheduling_kanban(selected_cw_cols)
            
            # 更新按钮状态
            self.confirm_scheduling_btn.setVisible(False)
            self.back_to_order_btn.setVisible(True)
            
            QMessageBox.information(self, "成功", f"已选择 {len(selected_cw_cols)} 个CW列进行排产")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"确定排产失败: {str(e)}")

    def switch_to_scheduling_kanban(self, selected_cw_cols):
        """切换到排产看板视图"""
        try:
            # 获取当前数据
            current_data = self.get_current_table_data()
            
            # 重新构建表格，只显示选中的CW列
            self.build_scheduling_kanban(current_data, selected_cw_cols)
            
        except Exception as e:
            print(f"切换到排产看板失败: {e}")

    def get_current_table_data(self):
        """获取当前表格数据"""
        try:
            data = []
            row_count = self.kanban_table.rowCount()
            col_count = self.kanban_table.columnCount()
            
            # 跳过复选框行（第一行）
            for row in range(1, row_count):
                row_data = {}
                for col in range(col_count):
                    item = self.kanban_table.item(row, col)
                    if item:
                        row_data[f"col_{col}"] = item.text()
                data.append(row_data)
            return data
        except Exception as e:
            print(f"获取表格数据失败: {e}")
            return []

    def build_scheduling_kanban(self, data, selected_cw_cols):
        """构建排产看板 - 将选中的CW列展开为每天"""
        try:
            # 在清空表格之前先保存表头信息
            header_info = []
            for col in range(self.kanban_table.columnCount()):
                header_item = self.kanban_table.horizontalHeaderItem(col)
                if header_item:
                    header_info.append({
                        'text': header_item.text(),
                        'date_str': header_item.data(Qt.UserRole)
                    })
                else:
                    header_info.append({'text': '', 'date_str': ''})
            
            # 清空表格
            self.kanban_table.clear()
            
            # 生成每天的日期列（使用保存的表头信息）
            daily_cols = self.generate_daily_columns(selected_cw_cols, header_info)
            
            # 设置列数：成品信息列 + 每天日期列 + 总合计列
            fixed_headers = ["成品名称", "成品规格", "成品品牌型号", "成品Project"]
            headers_count = len(fixed_headers) + len(daily_cols) + 1
            
            self.kanban_table.setColumnCount(headers_count)
            
            # 设置表头
            for i, title in enumerate(fixed_headers):
                item = QTableWidgetItem(title)
                self.kanban_table.setHorizontalHeaderItem(i, item)
            
            # 设置每天日期列表头
            base_col = len(fixed_headers)
            for i, day_info in enumerate(daily_cols):
                # 创建双行表头：第一行显示日期，第二行显示星期几
                header_item = QTableWidgetItem(day_info['date_str'])
                header_item.setData(Qt.UserRole, day_info['weekday_str'])
                
                # 为周日列的表头设置黄色背景
                if day_info['weekday_str'] == '日':
                    from PySide6.QtGui import QColor
                    header_item.setBackground(QColor("#FFFF99"))  # 黄色
                
                self.kanban_table.setHorizontalHeaderItem(base_col + i, header_item)
            
            
            # 设置总合计列表头
            self.kanban_table.setHorizontalHeaderItem(headers_count - 1, QTableWidgetItem("总合计"))
            
            # 填充数据
            self.populate_scheduling_data(data, selected_cw_cols, daily_cols, base_col, headers_count)
            
            # 设置列宽
            hdr = self.kanban_table.horizontalHeader()
            for i in range(len(fixed_headers)):
                hdr.setSectionResizeMode(i, QHeaderView.ResizeToContents)
            for c in range(len(fixed_headers), headers_count):
                hdr.setSectionResizeMode(c, QHeaderView.Fixed)
                self.kanban_table.setColumnWidth(c, 64)
            
            self.kanban_table.resizeRowsToContents()
            
        except Exception as e:
            print(f"构建排产看板失败: {e}")

    def generate_daily_columns(self, selected_cw_cols, header_info=None):
        """生成每天的日期列 - 每个CW往前展示7天"""
        try:
            daily_cols = []
            
            for cw_col in selected_cw_cols:
                if header_info and cw_col < len(header_info):
                    # 使用保存的表头信息
                    header_data = header_info[cw_col]
                    cw_text = header_data['text']
                    date_str = header_data['date_str']
                else:
                    # 尝试从当前表格获取表头信息
                    original_header = self.kanban_table.horizontalHeaderItem(cw_col)
                    
                    if original_header:
                        date_str = original_header.data(Qt.UserRole)
                        cw_text = original_header.text()
                    else:
                        continue
                
                if date_str:
                    # 解析CW的基准日期
                    from datetime import datetime, timedelta
                    base_date = datetime.strptime(date_str, "%Y/%m/%d").date()
                    
                    # 生成前面7天的日期（从7天前到基准日期）
                    for i in range(7):
                        day_date = base_date - timedelta(days=6-i)  # 从7天前开始
                        day_str = day_date.strftime("%m/%d")
                        
                        # 获取星期几
                        weekdays = ['一', '二', '三', '四', '五', '六', '日']
                        weekday_str = weekdays[day_date.weekday()]
                        
                        daily_cols.append({
                            'date_str': day_str,
                            'weekday_str': weekday_str,
                            'full_date': day_date.strftime("%Y-%m-%d"),
                            'cw_text': cw_text,
                            'original_col': cw_col,
                            'day_index': i
                        })
            return daily_cols
            
        except Exception as e:
            print(f"生成每天日期列失败: {e}")
            return []

    def populate_scheduling_data(self, data, selected_cw_cols, daily_cols, base_col, headers_count):
        """填充排产数据"""
        try:
            # 设置行数
            self.kanban_table.setRowCount(len(data))
            
            # 填充数据行
            for row_idx, row_data in enumerate(data):
                
                # 成品信息列
                pn_value = row_data.get("col_2", "")  # PN字段
                print(f"🔍 [排产视图] 处理第{row_idx}行，PN值: '{pn_value}'")
                
                # 根据PN字段获取成品信息
                from app.services.production_scheduling_service import ProductionSchedulingService
                product_info = ProductionSchedulingService.get_product_info_by_pn(pn_value)
                
                if product_info:
                    # 成品名称
                    self.kanban_table.setItem(row_idx, 0, QTableWidgetItem(product_info.get('CnName', '')))
                    # 成品规格
                    self.kanban_table.setItem(row_idx, 1, QTableWidgetItem(product_info.get('ItemSpec', '')))
                    # 成品品牌型号
                    self.kanban_table.setItem(row_idx, 2, QTableWidgetItem(product_info.get('Brand', '')))
                    # 成品Project
                    project_name = product_info.get('ProjectName', '') or product_info.get('ProjectCode', '')
                    self.kanban_table.setItem(row_idx, 3, QTableWidgetItem(project_name))
                else:
                    # 如果找不到成品信息，显示原始数据
                    self.kanban_table.setItem(row_idx, 0, QTableWidgetItem(pn_value))  # 成品名称显示PN
                    self.kanban_table.setItem(row_idx, 1, QTableWidgetItem(""))  # 成品规格
                    self.kanban_table.setItem(row_idx, 2, QTableWidgetItem(""))  # 成品品牌型号
                    self.kanban_table.setItem(row_idx, 3, QTableWidgetItem(row_data.get("col_4", "")))  # 成品Project显示原始Project
                
                # 每天日期列数据（可编辑）
                daily_total = 0
                for i, day_info in enumerate(daily_cols):
                    # 获取原始CW列的数量
                    original_cw_col = day_info['original_col']
                    original_qty = row_data.get(f"col_{original_cw_col}", "0")
                    
                    # 如果是CW的基准日期（第7天），显示原始数量，其他天初始化为0
                    if day_info['day_index'] == 6:  # 第7天是CW的基准日期
                        initial_qty = original_qty
                    else:
                        initial_qty = "0"
                    
                    item = QTableWidgetItem(initial_qty)
                    item.setData(Qt.UserRole, {
                        'type': 'daily_scheduling',
                        'date': day_info['full_date'],
                        'cw_text': day_info['cw_text'],
                        'original_col': day_info['original_col'],
                        'day_index': day_info['day_index'],
                        'qty': float(initial_qty) if initial_qty.isdigit() else 0
                    })
                    
                    # 设置列背景色
                    initial_qty_num = float(initial_qty) if initial_qty.isdigit() else 0
                    self.set_column_background_color(item, day_info, initial_qty_num)
                    
                    self.kanban_table.setItem(row_idx, base_col + i, item)
                
                
                # 总合计列
                grand_total = 0
                for cw_col in selected_cw_cols:
                    original_qty = row_data.get(f"col_{cw_col}", "0")
                    grand_total += float(original_qty) if original_qty.isdigit() else 0
                
                total_item = QTableWidgetItem(str(int(grand_total)))
                total_item.setFlags(total_item.flags() & ~Qt.ItemIsEditable)  # 不可编辑
                # 去掉蓝色背景
                total_item.setData(Qt.UserRole, {
                    'type': 'grand_total',
                    'total': grand_total
                })
                self.kanban_table.setItem(row_idx, headers_count - 1, total_item)
            
            # 连接单元格变更事件
            self.kanban_table.itemChanged.connect(self.on_scheduling_item_changed)
            
        except Exception as e:
            print(f"填充排产数据失败: {e}")

    def set_column_background_color(self, item, day_info, qty=0):
        """设置列的背景色"""
        try:
            from PySide6.QtGui import QColor
            
            weekday_str = day_info['weekday_str']
            day_index = day_info['day_index']
            
            # 每周日的列标为黄色（优先级最高）
            if weekday_str == '日':
                item.setBackground(QColor("#FFFF99"))  # 黄色
            # 基准日期列（第7天，即CW的基准日期）标为蓝色（优先级最高）
            elif day_index == 6:
                item.setBackground(QColor("#99CCFF"))  # 蓝色
            # 如果有数据不为0，设置为绿色（但不能覆盖黄色和蓝色）
            elif qty > 0:
                item.setBackground(QColor("#99FF99"))  # 绿色
            # 其他列保持默认背景色
            else:
                item.setBackground(QColor("#FFFFFF"))  # 白色
            
        except Exception as e:
            print(f"设置列背景色失败: {e}")

    def update_cell_background_color(self, item, qty):
        """动态更新单元格背景色"""
        try:
            from PySide6.QtGui import QColor
            
            data = item.data(Qt.UserRole)
            if not data:
                return
                
            weekday_str = data.get('weekday_str', '')
            day_index = data.get('day_index', 0)
            
            # 每周日的列标为黄色（优先级最高）
            if weekday_str == '日':
                item.setBackground(QColor("#FFFF99"))  # 黄色
            # 基准日期列（第7天，即CW的基准日期）标为蓝色（优先级最高）
            elif day_index == 6:
                item.setBackground(QColor("#99CCFF"))  # 蓝色
            # 如果有数据不为0，设置为绿色（但不能覆盖黄色和蓝色）
            elif qty > 0:
                item.setBackground(QColor("#99FF99"))  # 绿色
            # 其他列保持默认背景色
            else:
                item.setBackground(QColor("#FFFFFF"))  # 白色
            
        except Exception as e:
            print(f"更新单元格背景色失败: {e}")

    def on_scheduling_item_changed(self, item):
        """排产单元格变更事件"""
        try:
            data = item.data(Qt.UserRole)
            if not data or data.get('type') != 'daily_scheduling':
                return
            
            # 更新数量
            try:
                new_qty = float(item.text())
                data['qty'] = new_qty
                item.setData(Qt.UserRole, data)
                
                # 动态设置单元格背景色
                self.update_cell_background_color(item, new_qty)
                
                # 重新计算CW合计和总合计
                self.recalculate_cw_totals(item.row())
                
            except ValueError:
                QMessageBox.warning(self, "警告", "请输入有效的数字")
                item.setText(str(data['qty']))
                
        except Exception as e:
            print(f"排产单元格变更处理失败: {e}")

    def recalculate_cw_totals(self, row):
        """重新计算指定行的总合计"""
        try:
            col_count = self.kanban_table.columnCount()
            fixed_headers_count = 4  # 成品信息列数（成品名称、成品规格、成品品牌型号、成品Project）
            
            # 计算所有每天日期列的总和
            grand_total = 0
            for col in range(fixed_headers_count, col_count - 1):  # 排除总合计列
                item = self.kanban_table.item(row, col)
                if item:
                    data = item.data(Qt.UserRole)
                    if data and data.get('type') == 'daily_scheduling':
                        try:
                            grand_total += float(item.text())
                        except:
                            pass
            
            # 更新总合计列
            total_item = self.kanban_table.item(row, col_count - 1)
            if total_item:
                total_item.setText(str(int(grand_total)))
                total_item.setData(Qt.UserRole, {
                    'type': 'grand_total',
                    'total': grand_total
                })
                    
        except Exception as e:
            print(f"重新计算合计失败: {e}")

    def back_to_order_view(self):
        """返回订单视图"""
        try:
            # 重新加载原始数据
            self.load_kanban_data()
            
            # 更新按钮状态
            self.enter_scheduling_btn.setVisible(True)
            self.enter_scheduling_btn.setEnabled(True)
            self.confirm_scheduling_btn.setVisible(False)
            self.back_to_order_btn.setVisible(False)
            
            # 清除复选框
            self.cw_checkboxes.clear()
            
        except Exception as e:
            print(f"返回订单视图失败: {e}")
