#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
新的排产订单管理界面
- 创建排产订单
- 选择需要排产的成品
- 管理排产订单列表
"""

from typing import Optional, List, Dict
from datetime import datetime, timedelta

from PySide6.QtCore import Qt, QDate, Signal, QTimer, QRect
from PySide6.QtGui import QFont, QColor, QPainter, QBrush
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QFrame, QLineEdit, QComboBox, QAbstractItemView,
    QMessageBox, QTabWidget, QGroupBox, QGridLayout, QCheckBox, QDialog,
    QHeaderView, QDateEdit, QListWidget, QListWidgetItem, QSplitter,
    QSizePolicy, QScrollArea, QFormLayout, QDialogButtonBox, QTextEdit,
    QSpacerItem, QAbstractScrollArea, QFileDialog
)

from app.services.scheduling_order_service import SchedulingOrderService
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class TwoRowHeader(QHeaderView):
    """自定义两行表头，支持周日列背景色"""
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setDefaultAlignment(Qt.AlignCenter)
        self._top_font = QFont()
        self._top_font.setBold(True)
        self._bottom_font = QFont()
        self._bottom_font.setPointSize(self._bottom_font.pointSize() - 1)
        self._bottom_font.setBold(True)
        h = self.fontMetrics().height()
        self.setFixedHeight(int(h * 3.5))  # 增加高度，让两行之间有更多间距

    def sizeHint(self):
        s = super().sizeHint()
        h = self.fontMetrics().height()
        s.setHeight(int(h * 3.5))  # 与setFixedHeight保持一致
        return s

    def paintSection(self, painter: QPainter, rect: QRect, logicalIndex: int):
        if not rect.isValid():
            return
        
        # 完全自定义绘制，不使用父类方法
        painter.save()
        
        table = self.parent()
        item = table.horizontalHeaderItem(logicalIndex) if table else None
        top = item.text() if item else ""
        bottom = item.data(Qt.UserRole) if (item and item.data(Qt.UserRole) is not None) else ""

        # 检查是否是周日，设置整列黄色背景
        # 对于MRP表格，bottom是日期，需要检查日期对应的周几
        is_sunday = False
        if bottom and len(bottom) == 10 and bottom.count('-') == 2:  # 日期格式 YYYY-MM-DD
            try:
                from datetime import datetime
                date_obj = datetime.strptime(bottom, "%Y-%m-%d").date()
                weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
                is_sunday = (weekday == 6)
            except:
                pass
        elif bottom == "日":  # 兼容旧的周几格式
            is_sunday = True
        
        if is_sunday:
            painter.fillRect(rect, QColor("#fff3cd"))  # 使用更柔和的黄色
        else:
            painter.fillRect(rect, QColor("#fafafa"))  # 默认背景
        
        # 绘制边框
        painter.setPen(QColor("#d9d9d9"))
        painter.drawRect(rect)

        # 计算两行的矩形区域，增加间距
        top_height = int(rect.height() * 0.6)  # 第一行占60%
        bottom_height = rect.height() - top_height  # 第二行占剩余部分
        
        # 对于MRP表格，第一行显示周几，第二行显示日期
        if bottom and len(bottom) == 10 and bottom.count('-') == 2:  # 日期格式 YYYY-MM-DD
            try:
                from datetime import datetime
                date_obj = datetime.strptime(bottom, "%Y-%m-%d").date()
                weekday_names = ['一', '二', '三', '四', '五', '六', '日']
                weekday_text = weekday_names[date_obj.weekday()]
                date_text = date_obj.strftime("%m-%d")
                
                # 绘制第一行（周几）
                painter.setPen(QColor("#333333"))
                painter.setFont(self._top_font)
                topRect = QRect(rect.left() + 1, rect.top() + 1, rect.width() - 2, top_height - 2)
                painter.drawText(topRect, Qt.AlignCenter, weekday_text)
                
                # 绘制第二行（日期）
                painter.setFont(self._bottom_font)
                painter.setPen(QColor("#666666"))
                
                margin = 2
                bottomRect = QRect(
                    rect.left() + margin, 
                    rect.top() + top_height, 
                    rect.width() - margin * 2, 
                    bottom_height - margin
                )
                
                painter.drawText(bottomRect, Qt.AlignCenter | Qt.TextWrapAnywhere, date_text)
            except:
                # 如果解析失败，使用原来的逻辑
                painter.setPen(QColor("#333333"))
                painter.setFont(self._top_font)
                topRect = QRect(rect.left() + 1, rect.top() + 1, rect.width() - 2, top_height - 2)
                painter.drawText(topRect, Qt.AlignCenter, str(top))
                
                if bottom:
                    painter.setFont(self._bottom_font)
                    painter.setPen(QColor("#666666"))
                    
                    margin = 2
                    bottomRect = QRect(
                        rect.left() + margin, 
                        rect.top() + top_height, 
                        rect.width() - margin * 2, 
                        bottom_height - margin
                    )
                    
                    painter.drawText(bottomRect, Qt.AlignCenter | Qt.TextWrapAnywhere, str(bottom))
        else:
            # 原来的逻辑（用于排产看板）
            # 绘制第一行（日期）
            painter.setPen(QColor("#333333"))
            painter.setFont(self._top_font)
            topRect = QRect(rect.left() + 1, rect.top() + 1, rect.width() - 2, top_height - 2)
            painter.drawText(topRect, Qt.AlignCenter, str(top))
            
            # 绘制第二行（周几）
            if bottom:  # 只有当有周几数据时才绘制
                painter.setFont(self._bottom_font)
                painter.setPen(QColor("#666666"))
                
                # 为周几预留更多边距
                margin = 2
                bottomRect = QRect(
                    rect.left() + margin, 
                    rect.top() + top_height, 
                    rect.width() - margin * 2, 
                    bottom_height - margin
                )
                
                painter.drawText(bottomRect, Qt.AlignCenter | Qt.TextWrapAnywhere, str(bottom))
        
        painter.restore()


class ProductSelectionDialog(QDialog):
    """产品选择对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择排产产品")
        self.setModal(True)
        self.resize(800, 600)
        
        self.selected_products = []
        self.init_ui()
        self.load_products()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # 搜索框
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索:"))
        self.search_edit = QLineEdit()
        self.search_edit.textChanged.connect(self.filter_products)
        search_layout.addWidget(self.search_edit)
        search_layout.addStretch()
        layout.addLayout(search_layout)
        
        # 产品列表
        self.product_list = QListWidget()
        self.product_list.setSelectionMode(QAbstractItemView.MultiSelection)
        layout.addWidget(self.product_list)
        
        # 按钮
        button_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.select_all_btn.clicked.connect(self.select_all)
        self.clear_selection_btn = QPushButton("清空选择")
        self.clear_selection_btn.clicked.connect(self.clear_selection)
        button_layout.addWidget(self.select_all_btn)
        button_layout.addWidget(self.clear_selection_btn)
        button_layout.addStretch()
        
        self.ok_btn = QPushButton("确定")
        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def load_products(self):
        """加载可排产的产品列表"""
        try:
            products = SchedulingOrderService.get_available_products()
            self.all_products = products
            
            for product in products:
                item_text = f"{product['ItemCode']} - {product['CnName']}"
                if product.get('ItemSpec'):
                    item_text += f" ({product['ItemSpec']})"
                if product.get('Brand'):
                    item_text += f" [品牌: {product['Brand']}]"
                if product.get('ProjectName'):
                    item_text += f" [项目: {product['ProjectName']}]"
                
                item = QListWidgetItem(item_text)
                item.setData(Qt.UserRole, product)
                self.product_list.addItem(item)
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载产品列表失败: {str(e)}")
    
    def filter_products(self, text):
        """过滤产品列表"""
        for i in range(self.product_list.count()):
            item = self.product_list.item(i)
            product = item.data(Qt.UserRole)
            if product:
                search_text = text.lower()
                match = (
                    search_text in product['ItemCode'].lower() or
                    search_text in product['CnName'].lower() or
                    (product.get('ItemSpec', '').lower() and search_text in product['ItemSpec'].lower()) or
                    (product.get('Brand', '').lower() and search_text in product['Brand'].lower()) or
                    (product.get('ProjectName', '').lower() and search_text in product['ProjectName'].lower())
                )
                item.setHidden(not match)
    
    def select_all(self):
        """全选"""
        for i in range(self.product_list.count()):
            item = self.product_list.item(i)
            if not item.isHidden():
                item.setSelected(True)
    
    def clear_selection(self):
        """清空选择"""
        self.product_list.clearSelection()
    
    def get_selected_products(self):
        """获取选中的产品"""
        selected_items = self.product_list.selectedItems()
        return [item.data(Qt.UserRole) for item in selected_items]


class SchedulingOrderManagementWidget(QWidget):
    """排产订单管理主界面"""
    
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_orders()
    
    def init_ui(self):
        self.setWindowTitle("排产订单管理")
        
        # 设置大小策略
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        main_layout = QVBoxLayout()
        
        # 页签
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet(
            "QTabWidget::pane{border:1px solid #dee2e6;background:white;}"
            "QTabBar::tab{background:#f8f9fa;border:1px solid #dee2e6;padding:8px 16px;margin-right:2px;}"
            "QTabBar::tab:selected{background:white;border-bottom:2px solid #007bff;}"
        )
        
        self.create_order_list_tab()
        self.create_order_detail_tab()
        self.create_production_mrp_tab()
        
        main_layout.addWidget(self.tab_widget)
        self.setLayout(main_layout)
    
    def create_order_list_tab(self):
        """创建订单列表页签"""
        order_widget = QWidget()
        order_layout = QVBoxLayout()
        
        # 控制面板
        control_panel = QFrame()
        control_panel.setFrameStyle(QFrame.StyledPanel)
        control_panel.setStyleSheet("QFrame{background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;}")
        control_layout = QVBoxLayout()
        
        # 创建订单按钮
        create_layout = QHBoxLayout()
        self.create_order_btn = QPushButton("新建排产订单")
        self.create_order_btn.setStyleSheet("QPushButton{background:#007bff;color:#fff;border:none;padding:8px 16px;border-radius:4px;}"
                                           "QPushButton:hover{background:#0069d9;}")
        self.create_order_btn.clicked.connect(self.create_new_order)
        create_layout.addWidget(self.create_order_btn)
        create_layout.addStretch()
        control_layout.addLayout(create_layout)
        
        control_panel.setLayout(control_layout)
        order_layout.addWidget(control_panel)
        
        # 订单列表表格
        self.order_table = QTableWidget()
        self.order_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: white;
                alternate-background-color: #f8f9fa;
                selection-background-color: #e3f2fd;
                border: 1px solid #dee2e6;
                border-radius: 4px;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #dee2e6;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: #1976d2;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                color: #495057;
                padding: 12px 8px;
                border: none;
                border-bottom: 2px solid #dee2e6;
                font-weight: 600;
                font-size: 14px;
            }
        """)
        self.order_table.setAlternatingRowColors(True)
        self.order_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.order_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.order_table.itemSelectionChanged.connect(self.on_order_selected)
        
        # 设置表头
        headers = ["订单ID", "订单名称", "开始日期", "结束日期", "状态", "创建人", "创建时间", "备注", "操作"]
        self.order_table.setColumnCount(len(headers))
        self.order_table.setHorizontalHeaderLabels(headers)
        
        # 设置列宽
        header = self.order_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 订单ID
        header.setSectionResizeMode(1, QHeaderView.Stretch)          # 订单名称
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents) # 开始日期
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents) # 结束日期
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents) # 状态
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents) # 创建人
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents) # 创建时间
        header.setSectionResizeMode(7, QHeaderView.Stretch)          # 备注
        header.setSectionResizeMode(8, QHeaderView.Fixed)             # 操作列固定宽度
        self.order_table.setColumnWidth(8, 280)  # 设置操作列宽度为280像素
        
        order_layout.addWidget(self.order_table)
        order_widget.setLayout(order_layout)
        self.tab_widget.addTab(order_widget, "排产列表")
    
    def create_order_detail_tab(self):
        """创建排产详情页签"""
        detail_widget = QWidget()
        detail_layout = QVBoxLayout()
        
        # 排产订单选择面板
        selection_group = QGroupBox("选择排产订单")
        selection_layout = QVBoxLayout()
        
        # 订单选择下拉框
        order_layout = QHBoxLayout()
        order_layout.addWidget(QLabel("排产订单:"))
        self.order_selection_combo = QComboBox()
        self.order_selection_combo.setMinimumWidth(300)
        self.order_selection_combo.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background: white;
                min-width: 300px;
            }
            QComboBox:focus {
                border-color: #007bff;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #6c757d;
                margin-right: 5px;
            }
        """)
        self.order_selection_combo.currentTextChanged.connect(self.on_order_selection_changed)
        order_layout.addWidget(self.order_selection_combo)
        order_layout.addStretch()
        selection_layout.addLayout(order_layout)
        
        selection_group.setLayout(selection_layout)
        detail_layout.addWidget(selection_group)
        
        # 排产看板面板
        kanban_group = QGroupBox("排产看板")
        kanban_layout = QVBoxLayout()
        
        # 看板控制按钮
        kanban_btn_layout = QHBoxLayout()
        self.refresh_kanban_btn = QPushButton("刷新看板")
        self.refresh_kanban_btn.setStyleSheet("QPushButton{background:#6c757d;color:#fff;border:none;padding:8px 16px;border-radius:4px;}"
                                             "QPushButton:hover{background:#5a6268;}")
        self.refresh_kanban_btn.clicked.connect(self.refresh_kanban_data)
        self.refresh_kanban_btn.setEnabled(False)
        
        self.save_kanban_btn = QPushButton("保存排产")
        self.save_kanban_btn.setStyleSheet("QPushButton{background:#007bff;color:#fff;border:none;padding:8px 16px;border-radius:4px;}"
                                          "QPushButton:hover{background:#0056b3;}")
        self.save_kanban_btn.clicked.connect(self.save_kanban_data)
        self.save_kanban_btn.setEnabled(False)
        
        kanban_btn_layout.addWidget(self.refresh_kanban_btn)
        kanban_btn_layout.addWidget(self.save_kanban_btn)
        
        # 导入导出按钮
        self.export_kanban_btn = QPushButton("导出看板")
        self.export_kanban_btn.setStyleSheet("QPushButton{background:#6f42c1;color:#fff;border:none;padding:8px 16px;border-radius:4px;}"
                                           "QPushButton:hover{background:#5a32a3;}")
        self.export_kanban_btn.clicked.connect(self.export_kanban_data)
        self.export_kanban_btn.setEnabled(False)
        
        self.import_kanban_btn = QPushButton("导入看板")
        self.import_kanban_btn.setStyleSheet("QPushButton{background:#fd7e14;color:#fff;border:none;padding:8px 16px;border-radius:4px;}"
                                           "QPushButton:hover{background:#e8650e;}")
        self.import_kanban_btn.clicked.connect(self.import_kanban_data)
        self.import_kanban_btn.setEnabled(False)
        
        kanban_btn_layout.addWidget(self.export_kanban_btn)
        kanban_btn_layout.addWidget(self.import_kanban_btn)
        kanban_btn_layout.addStretch()
        kanban_layout.addLayout(kanban_btn_layout)
        
        # 看板表格
        self.kanban_table = QTableWidget()
        # 完全移除CSS样式，让背景色设置生效
        self.kanban_table.setStyleSheet("")
        
        # 设置自定义表头
        self.kanban_table.setHorizontalHeader(TwoRowHeader(Qt.Horizontal, self.kanban_table))
        
        # 禁用交替行颜色，避免覆盖我们的背景色设置
        self.kanban_table.setAlternatingRowColors(False)
        self.kanban_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.kanban_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.kanban_table.horizontalHeader().setStretchLastSection(True)
        
        # 设置行高
        self.kanban_table.verticalHeader().setDefaultSectionSize(35)
        
        # 连接单元格编辑信号
        self.kanban_table.itemChanged.connect(self.on_kanban_item_changed)
        self.kanban_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        
        # 设置表格大小调整策略
        try:
            policy = QAbstractScrollArea.SizeAdjustPolicy.AdjustToContentsOnFirstShow
        except AttributeError:
            policy = getattr(QAbstractScrollArea, "AdjustToContentsOnFirstShow", QAbstractScrollArea.AdjustToContents)
        self.kanban_table.setSizeAdjustPolicy(policy)
        self.kanban_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        kanban_layout.addWidget(self.kanban_table)
        kanban_group.setLayout(kanban_layout)
        detail_layout.addWidget(kanban_group)
        
        detail_widget.setLayout(detail_layout)
        self.tab_widget.addTab(detail_widget, "排产详情")
    
    def create_production_mrp_tab(self):
        """创建生产MRP计算页签"""
        mrp_widget = ProductionMRPWidget(self)
        self.tab_widget.addTab(mrp_widget, "生产MRP计算")
        
        # 初始化订单选择下拉框
        self.load_order_selection()
    
    def load_order_selection(self):
        """加载订单选择下拉框"""
        try:
            orders = SchedulingOrderService.get_scheduling_orders()
            
            self.order_selection_combo.clear()
            self.order_selection_combo.addItem("请选择排产订单", None)
            
            for order in orders:
                order_text = f"{order['OrderName']} ({order['StartDate']} - {order['EndDate']})"
                self.order_selection_combo.addItem(order_text, order['OrderId'])
                
        except Exception as e:
            print(f"加载订单选择列表失败: {e}")
    
    def on_order_selection_changed(self, text):
        """订单选择改变时的处理"""
        try:
            order_id = self.order_selection_combo.currentData()
            if order_id:
                self.current_order_id = order_id
                self.load_kanban_data()
                self.refresh_kanban_btn.setEnabled(True)
                self.save_kanban_btn.setEnabled(True)
                self.export_kanban_btn.setEnabled(True)
                self.import_kanban_btn.setEnabled(True)
            else:
                self.current_order_id = None
                self.clear_kanban_table()
                self.refresh_kanban_btn.setEnabled(False)
                self.save_kanban_btn.setEnabled(False)
                self.export_kanban_btn.setEnabled(False)
                self.import_kanban_btn.setEnabled(False)
        except Exception as e:
            print(f"订单选择处理失败: {e}")
    
    def load_kanban_data(self):
        """加载看板数据"""
        if not hasattr(self, 'current_order_id') or not self.current_order_id:
            return
            
        try:
            # 获取看板数据
            data = SchedulingOrderService.get_scheduling_kanban_data(self.current_order_id)
            
            if "error" in data:
                QMessageBox.critical(self, "错误", data["error"])
                return
            
            order_info = data["order_info"]
            date_range = data["date_range"]
            products = data["products"]
            
            # 设置表格列数
            # 固定列：产品名称、规格、型号、项目名称
            # 动态列：每天的日期
            fixed_cols = 4
            total_cols = fixed_cols + len(date_range)
            self.kanban_table.setColumnCount(total_cols)
            
            # 设置表头
            headers = ["产品名称", "规格", "型号", "项目名称"]
            for date_str in date_range:
                # 将日期转换为显示格式
                try:
                    from datetime import datetime
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                    # 周几改为中文显示
                    weekday_map = {
                        0: "一", 1: "二", 2: "三", 3: "四", 
                        4: "五", 5: "六", 6: "日"
                    }
                    weekday = weekday_map[date_obj.weekday()]
                    headers.append(date_str)  # 只添加日期，周几通过UserRole设置
                except:
                    headers.append(date_str)
            
            self.kanban_table.setHorizontalHeaderLabels(headers)
            
            # 为日期列设置周几数据
            fixed_cols = 4
            for i, date_str in enumerate(date_range):
                col_index = fixed_cols + i
                header_item = self.kanban_table.horizontalHeaderItem(col_index)
                if header_item:
                    try:
                        from datetime import datetime
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                        weekday_map = {
                            0: "一", 1: "二", 2: "三", 3: "四", 
                            4: "五", 5: "六", 6: "日"
                        }
                        weekday = weekday_map[date_obj.weekday()]
                        header_item.setData(Qt.UserRole, weekday)
                    except:
                        header_item.setData(Qt.UserRole, "")
            
            # 清空表格并设置行数
            self.kanban_table.clearContents()
            self.kanban_table.setRowCount(len(products))
            
            # 填充数据
            # 临时禁用所有信号
            self.kanban_table.blockSignals(True)
            
            for row, product in enumerate(products):
                # 固定列数据
                item_name = product.get("ItemName", "") or ""
                item_spec = product.get("ItemSpec", "") or ""
                item_model = product.get("Brand", "") or ""  # 使用Brand字段作为型号
                project_name = product.get("ProjectName", "") or ""
                
                # 创建QTableWidgetItem并设置文本
                item0 = QTableWidgetItem()
                item0.setText(item_name)
                item1 = QTableWidgetItem()
                item1.setText(item_spec)
                item2 = QTableWidgetItem()
                item2.setText(item_model)
                item3 = QTableWidgetItem()
                item3.setText(project_name)
                
                self.kanban_table.setItem(row, 0, item0)
                self.kanban_table.setItem(row, 1, item1)
                self.kanban_table.setItem(row, 2, item2)
                self.kanban_table.setItem(row, 3, item3)
                
                # 设置固定列不可编辑
                for col in range(fixed_cols):
                    item = self.kanban_table.item(row, col)
                    if item:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setBackground(QColor("#f8f9fa"))
                
                # 动态列数据（排产数量）
                for col, date_str in enumerate(date_range):
                    col_index = fixed_cols + col
                    qty = product["cells"].get(date_str, 0.0)
                    
                    item = QTableWidgetItem(str(qty))
                    item.setTextAlignment(Qt.AlignCenter)
                    
                    # 背景色将在set_sunday_column_backgrounds方法中统一设置
                    self.kanban_table.setItem(row, col_index, item)
            
            # 重新启用信号
            self.kanban_table.blockSignals(False)
            
            # 设置列宽和固定列效果
            header = self.kanban_table.horizontalHeader()
            
            # 设置固定列的宽度和调整模式
            for i in range(fixed_cols):
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
                # 设置固定列的具体宽度
                if i == 0:  # 产品名称
                    self.kanban_table.setColumnWidth(i, 150)
                elif i == 1:  # 规格
                    self.kanban_table.setColumnWidth(i, 100)
                elif i == 2:  # 型号
                    self.kanban_table.setColumnWidth(i, 80)
                elif i == 3:  # 项目名称
                    self.kanban_table.setColumnWidth(i, 120)
            
            # 设置日期列为固定宽度
            for c in range(fixed_cols, total_cols):
                header.setSectionResizeMode(c, QHeaderView.Fixed)
                self.kanban_table.setColumnWidth(c, 80)  # 增加列宽从64到80
            
            # 设置水平滚动模式
            self.kanban_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
            
            # 设置周日整列的单元格背景色
            self.set_sunday_column_backgrounds()
            
            # 刷新表格显示
            self.kanban_table.viewport().update()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载看板数据失败: {str(e)}")
    
    def set_sunday_column_backgrounds(self):
        """设置所有日期列的单元格背景色"""
        try:
            fixed_cols = 4
            col_count = self.kanban_table.columnCount()
            row_count = self.kanban_table.rowCount()
            
            for col in range(fixed_cols, col_count):
                # 检查是否是周日列
                header_item = self.kanban_table.horizontalHeaderItem(col)
                is_sunday = False
                if header_item:
                    weekday = header_item.data(Qt.UserRole)
                    is_sunday = weekday == "日"
                
                # 设置整列的单元格背景色
                for row in range(row_count):
                    item = self.kanban_table.item(row, col)
                    if item:
                        try:
                            qty = float(item.text() or 0)
                            if qty != 0:
                                # 有数量时显示绿色背景
                                item.setBackground(QColor("#d4edda"))  # 使用更柔和的绿色
                            elif is_sunday:
                                # 周日列且数量为0时显示黄色背景
                                item.setBackground(QColor("#fff3cd"))  # 使用更柔和的黄色
                            else:
                                # 非周日列且数量为0时显示白色背景
                                item.setBackground(QColor("#FFFFFF"))
                        except Exception as e:
                            print(f"解析数量失败: {e}, 文本: {item.text()}")
                            # 如果解析失败，根据是否周日设置背景色
                            if is_sunday:
                                item.setBackground(QColor("#fff3cd"))  # 使用更柔和的黄色
                            else:
                                item.setBackground(QColor("#FFFFFF"))
        except Exception as e:
            print(f"设置列背景色失败: {e}")
        
        # 刷新表格显示
        self.kanban_table.viewport().update()
    
    def clear_kanban_table(self):
        """清空看板表格"""
        self.kanban_table.clear()
        self.kanban_table.setRowCount(0)
        self.kanban_table.setColumnCount(0)
    
    def refresh_kanban_data(self):
        """刷新看板数据"""
        self.load_kanban_data()
    
    def on_kanban_item_changed(self, item):
        """看板单元格内容改变时的处理"""
        if item is None:
            return
        
        # 临时断开信号连接，避免递归
        self.kanban_table.itemChanged.disconnect(self.on_kanban_item_changed)
        
        try:
            # 获取单元格的值
            text = item.text().strip()
            if text == "":
                qty = 0.0
            else:
                qty = float(text)
            
            # 获取列索引
            col = item.column()
            fixed_cols = 4  # 前4列是固定列
            
            # 只处理日期列（非固定列）
            if col >= fixed_cols:
                # 检查是否是周日列
                header_item = self.kanban_table.horizontalHeaderItem(col)
                is_sunday = False
                if header_item:
                    weekday = header_item.data(Qt.UserRole)
                    is_sunday = weekday == "日"
                
                # 根据数量和是否周日设置背景色
                if qty != 0:
                    # 数量不为0时显示绿色背景
                    item.setBackground(QColor("#d4edda"))  # 使用更柔和的绿色
                elif is_sunday:
                    # 周日列且数量为0时显示黄色背景
                    item.setBackground(QColor("#fff3cd"))  # 使用更柔和的黄色
                else:
                    # 非周日列且数量为0时显示白色背景
                    item.setBackground(QColor("#FFFFFF"))
                        
        except ValueError:
            # 如果输入的不是有效数字，恢复为0
            item.setText("0")
            # 检查是否是周日列来决定背景色
            col = item.column()
            fixed_cols = 4
            if col >= fixed_cols:
                header_item = self.kanban_table.horizontalHeaderItem(col)
                is_sunday = False
                if header_item:
                    weekday = header_item.data(Qt.UserRole)
                    is_sunday = weekday == "日"
                
                if is_sunday:
                    item.setBackground(QColor("#fff3cd"))  # 使用更柔和的黄色
                else:
                    item.setBackground(QColor("#FFFFFF"))
        
        finally:
            # 重新连接信号
            self.kanban_table.itemChanged.connect(self.on_kanban_item_changed)
    
    def save_kanban_data(self):
        """保存看板数据"""
        if not hasattr(self, 'current_order_id') or not self.current_order_id:
            return
            
        try:
            # 获取看板数据
            data = SchedulingOrderService.get_scheduling_kanban_data(self.current_order_id)
            if "error" in data:
                QMessageBox.critical(self, "错误", data["error"])
                return
            
            date_range = data["date_range"]
            products = data["products"]
            fixed_cols = 4  # 前4列是固定列：产品名称、规格、品牌、项目名称
            
            # 收集更新的数据
            updates = []
            for row, product in enumerate(products):
                item_id = product["ItemId"]
                
                for col, date_str in enumerate(date_range):
                    col_index = fixed_cols + col
                    item = self.kanban_table.item(row, col_index)
                    
                    if item:
                        try:
                            planned_qty = float(item.text() or 0)
                            updates.append({
                                "ItemId": item_id,
                                "ProductionDate": date_str,
                                "PlannedQty": planned_qty
                            })
                        except ValueError:
                            QMessageBox.warning(self, "警告", f"第{row+1}行第{col+1}列的数据格式不正确")
                            return
            
            # 批量更新
            success, message = SchedulingOrderService.batch_update_scheduling_lines(
                self.current_order_id, updates
            )
            
            if success:
                QMessageBox.information(self, "成功", message)
            else:
                QMessageBox.critical(self, "错误", message)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存看板数据失败: {str(e)}")
    
    def export_kanban_data(self):
        """导出看板数据到Excel"""
        if not hasattr(self, 'current_order_id') or not self.current_order_id:
            QMessageBox.warning(self, "警告", "请先选择一个排产订单")
            return
        
        try:
            # 获取文件保存路径
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "导出看板数据", 
                f"排产看板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel文件 (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # 获取看板数据
            data = SchedulingOrderService.get_scheduling_kanban_data(self.current_order_id)
            if "error" in data:
                QMessageBox.critical(self, "错误", data["error"])
                return
            
            order_info = data["order_info"]
            date_range = data["date_range"]
            products = data["products"]
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "排产看板"
            
            # 设置样式
            header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            sunday_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            product_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # 绿色背景，与看板一致
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 直接从第1行开始，不显示基础信息
            current_row = 1
            
            # 创建表头 - 第一行：产品信息列 + 日期
            headers_row1 = ["产品名称", "规格", "型号", "项目名称"]
            headers_row2 = ["", "", "", ""]  # 第二行：产品信息列为空，日期列显示周几
            
            for date_str in date_range:
                headers_row1.append(date_str)
                # 计算周几
                try:
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                    weekday_map = {
                        0: "一", 1: "二", 2: "三", 3: "四", 
                        4: "五", 5: "六", 6: "日"
                    }
                    weekday = weekday_map[date_obj.weekday()]
                    headers_row2.append(weekday)
                except:
                    headers_row2.append("")
            
            # 写入第一行表头（日期）
            for col, header in enumerate(headers_row1, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = center_alignment
                
                # 如果是日期列且是周日，设置黄色背景
                if col > 4:  # 日期列从第5列开始
                    try:
                        date_obj = datetime.strptime(header, "%Y-%m-%d").date()
                        if date_obj.weekday() == 6:  # 周日
                            cell.fill = sunday_fill
                    except:
                        pass
            
            current_row += 1
            
            # 写入第二行表头（周几）
            for col, header in enumerate(headers_row2, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = center_alignment
                
                # 如果是周日列，设置黄色背景
                if col > 4 and header == "日":  # 周日列
                    cell.fill = sunday_fill
            
            current_row += 1
            
            # 写入数据
            for product in products:
                # 固定列数据
                ws.cell(row=current_row, column=1, value=product.get("ItemName", ""))
                ws.cell(row=current_row, column=2, value=product.get("ItemSpec", ""))
                ws.cell(row=current_row, column=3, value=product.get("Brand", ""))  # Brand字段作为型号
                ws.cell(row=current_row, column=4, value=product.get("ProjectName", ""))
                
                # 设置固定列样式
                for col in range(1, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = product_fill
                    cell.border = border
                    cell.alignment = center_alignment
                
                # 动态列数据（排产数量）
                for col, date_str in enumerate(date_range, 5):
                    qty = product["cells"].get(date_str, 0.0)
                    cell = ws.cell(row=current_row, column=col, value=qty)
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # 设置背景色：优先检查数量，然后检查周日
                    try:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                        is_sunday = date_obj.weekday() == 6
                        
                        if qty != 0:
                            # 不为0的数据设置绿色背景
                            cell.fill = green_fill
                        elif is_sunday:
                            # 周日且为0的数据设置黄色背景
                            cell.fill = sunday_fill
                        # 其他情况保持默认背景（白色）
                    except:
                        # 如果日期解析失败，只检查数量
                        if qty != 0:
                            cell.fill = green_fill
                
                current_row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20  # 产品名称
            ws.column_dimensions['B'].width = 15  # 规格
            ws.column_dimensions['C'].width = 12  # 型号
            ws.column_dimensions['D'].width = 15  # 项目名称
            
            # 设置日期列宽度
            for col in range(5, len(headers_row1) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10
            
            # 保存文件
            wb.save(file_path)
            QMessageBox.information(self, "成功", f"看板数据已导出到:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出看板数据失败: {str(e)}")
    
    def import_kanban_data(self):
        """从Excel导入看板数据"""
        if not hasattr(self, 'current_order_id') or not self.current_order_id:
            QMessageBox.warning(self, "警告", "请先选择一个排产订单")
            return
        
        try:
            # 获取文件路径
            file_path, _ = QFileDialog.getOpenFileName(
                self, 
                "导入看板数据", 
                "", 
                "Excel文件 (*.xlsx *.xls)"
            )
            
            if not file_path:
                return
            
            # 读取Excel文件
            df = pd.read_excel(file_path, header=None)
            
            # 查找数据开始行（从第1行开始，跳过两行表头）
            data_start_row = None
            for i in range(len(df)):
                if df.iloc[i, 0] == "产品名称":
                    data_start_row = i
                    break
            
            if data_start_row is None:
                QMessageBox.critical(self, "错误", "Excel文件格式不正确，找不到表头")
                return
            
            # 获取表头信息（使用第一行表头，包含日期）
            headers = df.iloc[data_start_row].tolist()
            
            # 验证表头格式
            expected_headers = ["产品名称", "规格", "型号", "项目名称"]
            if not all(header in headers[:4] for header in expected_headers):
                QMessageBox.critical(self, "错误", "Excel文件表头格式不正确")
                return
            
            # 获取日期列
            date_columns = []
            for i, header in enumerate(headers[4:], 4):
                if pd.notna(header):
                    try:
                        # 尝试解析日期
                        if isinstance(header, str):
                            datetime.strptime(header, "%Y-%m-%d")
                        date_columns.append(i)
                    except:
                        continue
            
            if not date_columns:
                QMessageBox.critical(self, "错误", "Excel文件中没有找到有效的日期列")
                return
            
            # 获取当前看板数据以验证产品匹配
            current_data = SchedulingOrderService.get_scheduling_kanban_data(self.current_order_id)
            if "error" in current_data:
                QMessageBox.critical(self, "错误", current_data["error"])
                return
            
            current_products = current_data["products"]
            current_date_range = current_data["date_range"]
            
            # 创建产品名称到ItemId的映射
            product_map = {}
            for product in current_products:
                product_name = product.get("ItemName", "")
                if product_name:
                    product_map[product_name] = product["ItemId"]
            
            # 收集导入的数据
            updates = []
            errors = []
            
            for row_idx in range(data_start_row + 1, len(df)):
                row_data = df.iloc[row_idx]
                
                # 获取产品名称
                product_name = str(row_data.iloc[0]) if pd.notna(row_data.iloc[0]) else ""
                if not product_name or product_name == "nan":
                    continue
                
                # 检查产品是否存在
                if product_name not in product_map:
                    errors.append(f"第{row_idx + 1}行: 产品 '{product_name}' 在当前订单中不存在")
                    continue
                
                item_id = product_map[product_name]
                
                # 处理每个日期列
                for col_idx in date_columns:
                    if col_idx >= len(row_data):
                        continue
                    
                    # 获取日期
                    date_header = headers[col_idx]
                    try:
                        if isinstance(date_header, str):
                            date_str = datetime.strptime(date_header, "%Y-%m-%d").strftime("%Y-%m-%d")
                        else:
                            date_str = date_header.strftime("%Y-%m-%d")
                    except:
                        continue
                    
                    # 检查日期是否在当前日期范围内
                    if date_str not in current_date_range:
                        continue
                    
                    # 获取数量
                    qty_value = row_data.iloc[col_idx]
                    try:
                        if pd.isna(qty_value):
                            planned_qty = 0.0
                        else:
                            planned_qty = float(qty_value)
                    except:
                        errors.append(f"第{row_idx + 1}行第{col_idx + 1}列: 数量格式不正确")
                        continue
                    
                    updates.append({
                        "ItemId": item_id,
                        "ProductionDate": date_str,
                        "PlannedQty": planned_qty
                    })
            
            # 显示错误信息
            if errors:
                error_msg = "导入过程中发现以下错误:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    error_msg += f"\n... 还有{len(errors) - 10}个错误"
                QMessageBox.warning(self, "导入警告", error_msg)
            
            if not updates:
                QMessageBox.warning(self, "警告", "没有找到有效的数据进行导入")
                return
            
            # 确认导入
            reply = QMessageBox.question(
                self, 
                "确认导入", 
                f"将导入 {len(updates)} 条排产数据，是否继续？",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # 批量更新数据
            success, message = SchedulingOrderService.batch_update_scheduling_lines(
                self.current_order_id, updates
            )
            
            if success:
                QMessageBox.information(self, "成功", f"成功导入 {len(updates)} 条数据")
                # 刷新看板显示
                self.load_kanban_data()
            else:
                QMessageBox.critical(self, "错误", message)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入看板数据失败: {str(e)}")
    
    def load_orders(self):
        """加载排产订单列表"""
        try:
            orders = SchedulingOrderService.get_scheduling_orders()
            
            self.order_table.setRowCount(len(orders))
            for row, order in enumerate(orders):
                self.order_table.setItem(row, 0, QTableWidgetItem(str(order["OrderId"])))
                self.order_table.setItem(row, 1, QTableWidgetItem(order["OrderName"]))
                self.order_table.setItem(row, 2, QTableWidgetItem(order["StartDate"]))
                self.order_table.setItem(row, 3, QTableWidgetItem(order["EndDate"]))
                
                # 状态列设置颜色
                status_item = QTableWidgetItem(order["Status"])
                if order["Status"] == "Draft":
                    status_item.setBackground(QColor("#fff3cd"))
                elif order["Status"] == "Active":
                    status_item.setBackground(QColor("#d4edda"))
                elif order["Status"] == "Completed":
                    status_item.setBackground(QColor("#d1ecf1"))
                elif order["Status"] == "Cancelled":
                    status_item.setBackground(QColor("#f8d7da"))
                self.order_table.setItem(row, 4, status_item)
                
                self.order_table.setItem(row, 5, QTableWidgetItem(order["CreatedBy"] or ""))
                self.order_table.setItem(row, 6, QTableWidgetItem(order["CreatedDate"]))
                self.order_table.setItem(row, 7, QTableWidgetItem(order["Remark"] or ""))
                
                # 操作按钮
                button_widget = QWidget()
                button_layout = QHBoxLayout(button_widget)
                button_layout.setContentsMargins(8, 4, 8, 4)
                button_layout.setSpacing(6)
                
                # 查看按钮
                view_btn = QPushButton("查看")
                view_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #17a2b8;
                        color: white;
                        border: none;
                        padding: 6px 12px;
                        border-radius: 4px;
                        font-size: 12px;
                        font-weight: bold;
                        min-width: 50px;
                    }
                    QPushButton:hover {
                        background-color: #138496;
                    }
                    QPushButton:pressed {
                        background-color: #117a8b;
                    }
                """)
                view_btn.clicked.connect(lambda checked, order_id=order['OrderId']: self.view_order_details(order_id))
                button_layout.addWidget(view_btn)
                
                
                # 编辑按钮
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #ffc107;
                        color: black;
                        border: none;
                        padding: 6px 12px;
                        border-radius: 4px;
                        font-size: 12px;
                        font-weight: bold;
                        min-width: 50px;
                    }
                    QPushButton:hover {
                        background-color: #e0a800;
                    }
                    QPushButton:pressed {
                        background-color: #d39e00;
                    }
                """)
                edit_btn.clicked.connect(lambda checked, order_id=order['OrderId']: self.edit_order(order_id))
                button_layout.addWidget(edit_btn)
                
                # 删除按钮
                delete_btn = QPushButton("删除")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #dc3545;
                        color: white;
                        border: none;
                        padding: 6px 12px;
                        border-radius: 4px;
                        font-size: 12px;
                        font-weight: bold;
                        min-width: 50px;
                    }
                    QPushButton:hover {
                        background-color: #c82333;
                    }
                    QPushButton:pressed {
                        background-color: #bd2130;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, order_id=order['OrderId']: self.delete_order(order_id))
                button_layout.addWidget(delete_btn)
                
                self.order_table.setCellWidget(row, 8, button_widget)
            
            self.order_table.resizeRowsToContents()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载订单列表失败: {str(e)}")
    
    def create_new_order(self):
        """创建新的排产订单"""
        dialog = CreateOrderDialog(self)
        if dialog.exec() == QDialog.Accepted:
            order_data = dialog.get_order_data()
            
            if not order_data["order_name"]:
                QMessageBox.warning(self, "警告", "请输入订单名称")
                return
            
            if not order_data["selected_products"]:
                QMessageBox.warning(self, "警告", "请至少选择一个成品物料")
                return
            
            # 创建排产订单
            success, message, order_id = SchedulingOrderService.create_scheduling_order(
                order_data["order_name"],
                order_data["start_date"],
                order_data["end_date"],
                "System",
                order_data["remark"]
            )
            
            if success:
                # 添加选中的产品到订单
                product_ids = [p['ItemId'] for p in order_data["selected_products"]]
                add_success, add_message = SchedulingOrderService.add_products_to_order(
                    order_id, product_ids
                )
                
                if add_success:
                    QMessageBox.information(self, "成功", f"{message}\n{add_message}")
                    self.load_orders()
                else:
                    QMessageBox.warning(self, "部分成功", f"{message}\n但添加产品失败: {add_message}")
                    self.load_orders()
            else:
                QMessageBox.critical(self, "错误", message)
    
    def on_order_selected(self):
        """订单选择事件 - 取消自动跳转到详情页"""
        current_row = self.order_table.currentRow()
        if current_row >= 0:
            order_id = int(self.order_table.item(current_row, 0).text())
            self.current_order_id = order_id
            # 不再自动跳转到详情页，用户需要点击"查看"按钮
        else:
            self.current_order_id = None
            
    
    def calculate_mrp(self):
        """计算MRP"""
        if not hasattr(self, 'current_order_id') or not self.current_order_id:
            return
        
        try:
            from app.ui.scheduling_mrp_calculation import SchedulingMRPDialog
            dialog = SchedulingMRPDialog(self.current_order_id, self)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"打开MRP计算界面失败: {str(e)}")
    
    def view_order_details(self, order_id):
        """查看订单详情"""
        # 切换到详情页签
        self.tab_widget.setCurrentIndex(1)
        # 设置订单选择下拉框为当前订单
        for i in range(self.order_selection_combo.count()):
            if self.order_selection_combo.itemData(i) == order_id:
                self.order_selection_combo.setCurrentIndex(i)
                break
    
    
    def edit_order(self, order_id):
        """编辑订单基础信息"""
        try:
            # 获取订单信息
            orders = SchedulingOrderService.get_scheduling_orders()
            order_info = None
            for order in orders:
                if order['OrderId'] == order_id:
                    order_info = order
                    break
            
            if not order_info:
                QMessageBox.critical(self, "错误", "找不到指定的订单信息")
            return
        
            # 打开编辑对话框
            dialog = EditOrderDialog(order_info, self)
            if dialog.exec() == QDialog.Accepted:
                updated_data = dialog.get_order_data()
                
                # 更新订单基础信息
                success, message = SchedulingOrderService.update_scheduling_order(
                    order_id,
                    updated_data["order_name"],
                    updated_data["start_date"],
                    updated_data["end_date"],
                    None,  # status参数
                    "System",  # updated_by参数
                    updated_data["remark"]
                )
                
                if success:
                    # 检查是否需要更新产品列表
                    if updated_data["selected_products"]:
                        # 获取当前订单的产品ID列表
                        current_products = SchedulingOrderService.get_order_products(order_id)
                        current_product_ids = [p['ItemId'] for p in current_products]
                        new_product_ids = [p['ItemId'] for p in updated_data["selected_products"]]
                        
                        # 如果产品列表有变化，需要更新
                        if set(current_product_ids) != set(new_product_ids):
                            # 删除不再需要的产品
                            for product in current_products:
                                if product['ItemId'] not in new_product_ids:
                                    SchedulingOrderService.remove_product_from_order(order_id, product['ItemId'])
                            
                            # 添加新产品
                            products_to_add = []
                            for product in updated_data["selected_products"]:
                                if product['ItemId'] not in current_product_ids:
                                    products_to_add.append(product['ItemId'])
                            
                            if products_to_add:
                                SchedulingOrderService.add_products_to_order(order_id, products_to_add)
                    
                    QMessageBox.information(self, "成功", "订单信息更新成功")
                    self.load_orders()
                    self.load_order_selection()  # 刷新订单选择下拉框
                    
                    # 如果当前正在查看这个订单的详情，需要重新加载
                    if hasattr(self, 'current_order_id') and self.current_order_id == order_id:
                        self.load_kanban_data()
                else:
                    QMessageBox.critical(self, "错误", message)
                    
        except Exception as e:
            QMessageBox.critical(self, "错误", f"编辑订单失败: {str(e)}")
    
    def delete_order(self, order_id):
        """删除订单"""
        reply = QMessageBox.question(
            self, "确认删除", 
            "确定要删除这个排产订单吗？删除后将无法恢复。",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            success, message = SchedulingOrderService.delete_scheduling_order(order_id)
            
            if success:
                QMessageBox.information(self, "成功", message)
                self.load_orders()
            else:
                QMessageBox.critical(self, "错误", message)


class CreateOrderDialog(QDialog):
    """创建排产订单对话框 - 美化版本"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建排产订单")
        self.setModal(True)
        self.resize(600, 700)
        self.setMinimumSize(500, 600)
        self.init_ui()
        self.load_products()
    
    def init_ui(self):
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)
        
        # 标题
        title_label = QLabel("新建排产订单")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #333;
                padding: 5px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        # 表单区域
        form_layout = QVBoxLayout()
        
        # 订单名称
        name_group = QGroupBox("订单信息")
        name_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                color: #495057;
                border: 1px solid #ccc;
                margin-top: 5px;
                padding-top: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 5px;
                padding: 0 3px 0 3px;
            }
        """)
        name_layout = QFormLayout(name_group)
        
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("请输入排产订单名称")
        self.name_edit.setStyleSheet("""
            QLineEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #007bff;
            }
        """)
        name_layout.addRow("订单名称 *:", self.name_edit)
        
        # 日期选择
        date_layout = QHBoxLayout()
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate())
        self.start_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QDateEdit:focus {
                border-color: #007bff;
            }
        """)
        self.start_date_edit.dateChanged.connect(self.update_end_date)
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QDateEdit:focus {
                border-color: #007bff;
            }
        """)
        # 结束日期可以手动修改，默认是一个月后
        
        # 初始化结束日期
        self.update_end_date()
        
        date_layout.addWidget(QLabel("开始日期 *:"))
        date_layout.addWidget(self.start_date_edit)
        date_layout.addWidget(QLabel("结束日期:"))
        date_layout.addWidget(self.end_date_edit)
        name_layout.addRow(date_layout)
        
        # 备注
        self.remark_edit = QTextEdit()
        self.remark_edit.setPlaceholderText("请输入备注信息（可选）")
        self.remark_edit.setMaximumHeight(60)
        self.remark_edit.setStyleSheet("""
            QTextEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QTextEdit:focus {
                border-color: #007bff;
            }
        """)
        name_layout.addRow("备注:", self.remark_edit)
        
        form_layout.addWidget(name_group)
        
        # 产品选择区域
        product_group = QGroupBox("选择成品物料")
        product_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                color: #495057;
                border: 1px solid #ccc;
                margin-top: 5px;
                padding-top: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 5px;
                padding: 0 3px 0 3px;
            }
        """)
        product_layout = QVBoxLayout(product_group)
        
        # 搜索框
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入产品编码或名称进行搜索...")
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #007bff;
            }
        """)
        self.search_edit.textChanged.connect(self.filter_products)
        search_layout.addWidget(self.search_edit)
        product_layout.addLayout(search_layout)
        
        # 产品列表 - 使用QTableWidget替代QListWidget以支持复选框
        self.product_table = QTableWidget()
        self.product_table.setColumnCount(2)
        self.product_table.setHorizontalHeaderLabels(["选择", "产品信息"])
        self.product_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 4px;
                border-bottom: 1px solid #f0f0f0;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 4px;
                border: none;
                font-weight: bold;
            }
        """)
        self.product_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.product_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.product_table.setMaximumHeight(250)
        self.product_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        product_layout.addWidget(self.product_table)
        
        # 产品操作按钮
        product_btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.select_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 4px 8px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.select_all_btn.clicked.connect(self.select_all_products)
        product_btn_layout.addWidget(self.select_all_btn)
        
        self.clear_selection_btn = QPushButton("清空选择")
        self.clear_selection_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 4px 8px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.clear_selection_btn.clicked.connect(self.clear_product_selection)
        product_btn_layout.addWidget(self.clear_selection_btn)
        
        product_btn_layout.addStretch()
        product_layout.addLayout(product_btn_layout)
        
        form_layout.addWidget(product_group)
        main_layout.addLayout(form_layout)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.ok_btn = QPushButton("创建订单")
        self.ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 6px 12px;
                font-size: 12px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        self.ok_btn.clicked.connect(self.accept)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 6px 12px;
                font-size: 12px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        main_layout.addLayout(button_layout)
    
    def load_products(self):
        """加载可排产的产品列表"""
        try:
            products = SchedulingOrderService.get_available_products()
            self.all_products = products
            
            self.product_table.setRowCount(len(products))
            
            for row, product in enumerate(products):
                # 复选框
                checkbox = QCheckBox()
                self.product_table.setCellWidget(row, 0, checkbox)
                
                # 产品信息
                item_text = f"{product['ItemCode']} - {product['CnName']}"
                if product.get('ItemSpec'):
                    item_text += f" ({product['ItemSpec']})"
                if product.get('Brand'):
                    item_text += f" [品牌: {product['Brand']}]"
                if product.get('ProjectName'):
                    item_text += f" [项目: {product['ProjectName']}]"
                
                item = QTableWidgetItem(item_text)
                item.setData(Qt.UserRole, product)
                self.product_table.setItem(row, 1, item)
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载产品列表失败: {str(e)}")
    
    def filter_products(self, text):
        """过滤产品列表"""
        for row in range(self.product_table.rowCount()):
            item = self.product_table.item(row, 1)
            if item:
                product = item.data(Qt.UserRole)
                if product:
                    item_text = f"{product['ItemCode']} - {product['CnName']}"
                    if product.get('ItemSpec'):
                        item_text += f" ({product['ItemSpec']})"
                    if product.get('Brand'):
                        item_text += f" [品牌: {product['Brand']}]"
                    if product.get('ProjectName'):
                        item_text += f" [项目: {product['ProjectName']}]"
                    
                    self.product_table.setRowHidden(row, text.lower() not in item_text.lower())
    
    def select_all_products(self):
        """全选产品"""
        for row in range(self.product_table.rowCount()):
            if not self.product_table.isRowHidden(row):
                checkbox = self.product_table.cellWidget(row, 0)
                if checkbox:
                    checkbox.setChecked(True)
    
    def clear_product_selection(self):
        """清空产品选择"""
        for row in range(self.product_table.rowCount()):
            checkbox = self.product_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(False)
    
    def update_end_date(self):
        """更新结束日期（开始日期后1个月）"""
        start_date = self.start_date_edit.date()
        end_date = start_date.addDays(30)  # 添加30天
        self.end_date_edit.setDate(end_date)
    
    def get_order_data(self):
        """获取订单数据"""
        selected_products = []
        for row in range(self.product_table.rowCount()):
            checkbox = self.product_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                # 从表格项中获取产品数据
                item = self.product_table.item(row, 1)
                if item:
                    product = item.data(Qt.UserRole)
                    selected_products.append(product)
        
        return {
            "order_name": self.name_edit.text().strip(),
            "start_date": self.start_date_edit.date().toString("yyyy-MM-dd"),
            "end_date": self.end_date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.toPlainText().strip(),
            "selected_products": selected_products
        }
    
    def accept(self):
        """确认创建"""
        if not self.name_edit.text().strip():
            QMessageBox.warning(self, "警告", "请输入订单名称")
            return
        
        super().accept()


class EditOrderDialog(QDialog):
    """编辑排产订单对话框"""
    
    def __init__(self, order_info, parent=None):
        super().__init__(parent)
        self.order_info = order_info
        self.setWindowTitle("编辑排产订单")
        self.setModal(True)
        self.resize(600, 700)
        self.setMinimumSize(500, 600)
        self.init_ui()
        self.load_current_data()
    
    def init_ui(self):
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)
        
        # 标题
        title_label = QLabel("编辑排产订单")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #333;
                padding: 5px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        # 表单区域
        form_layout = QVBoxLayout()
        
        # 订单名称
        name_group = QGroupBox("订单信息")
        name_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                color: #495057;
                border: 1px solid #ccc;
                margin-top: 5px;
                padding-top: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 5px;
                padding: 0 3px 0 3px;
            }
        """)
        name_layout = QFormLayout(name_group)
        
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("请输入排产订单名称")
        self.name_edit.setStyleSheet("""
            QLineEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #007bff;
            }
        """)
        name_layout.addRow("订单名称 *:", self.name_edit)
        
        # 日期选择
        date_layout = QHBoxLayout()
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QDateEdit:focus {
                border-color: #007bff;
            }
        """)
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QDateEdit:focus {
                border-color: #007bff;
            }
        """)
        
        date_layout.addWidget(QLabel("开始日期 *:"))
        date_layout.addWidget(self.start_date_edit)
        date_layout.addWidget(QLabel("结束日期 *:"))
        date_layout.addWidget(self.end_date_edit)
        name_layout.addRow(date_layout)
        
        # 备注
        self.remark_edit = QTextEdit()
        self.remark_edit.setPlaceholderText("请输入备注信息（可选）")
        self.remark_edit.setMaximumHeight(60)
        self.remark_edit.setStyleSheet("""
            QTextEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QTextEdit:focus {
                border-color: #007bff;
            }
        """)
        name_layout.addRow("备注:", self.remark_edit)
        
        form_layout.addWidget(name_group)
        
        # 产品选择区域
        product_group = QGroupBox("选择成品物料")
        product_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                color: #495057;
                border: 1px solid #ccc;
                margin-top: 5px;
                padding-top: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 5px;
                padding: 0 3px 0 3px;
            }
        """)
        product_layout = QVBoxLayout(product_group)
        
        # 搜索框
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入产品编码或名称进行搜索...")
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #007bff;
            }
        """)
        self.search_edit.textChanged.connect(self.filter_products)
        search_layout.addWidget(self.search_edit)
        product_layout.addLayout(search_layout)
        
        # 产品列表
        self.product_table = QTableWidget()
        self.product_table.setColumnCount(2)
        self.product_table.setHorizontalHeaderLabels(["选择", "产品信息"])
        self.product_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ccc;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 4px;
                border-bottom: 1px solid #f0f0f0;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 4px;
                border: none;
                font-weight: bold;
            }
        """)
        self.product_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.product_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.product_table.setMaximumHeight(250)
        self.product_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        product_layout.addWidget(self.product_table)
        
        # 产品操作按钮
        product_btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.select_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 4px 8px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.select_all_btn.clicked.connect(self.select_all_products)
        product_btn_layout.addWidget(self.select_all_btn)
        
        self.clear_selection_btn = QPushButton("清空选择")
        self.clear_selection_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 4px 8px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.clear_selection_btn.clicked.connect(self.clear_product_selection)
        product_btn_layout.addWidget(self.clear_selection_btn)
        
        product_btn_layout.addStretch()
        product_layout.addLayout(product_btn_layout)
        
        form_layout.addWidget(product_group)
        main_layout.addLayout(form_layout)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.ok_btn = QPushButton("保存修改")
        self.ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 6px 12px;
                font-size: 12px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        self.ok_btn.clicked.connect(self.accept)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 6px 12px;
                font-size: 12px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        main_layout.addLayout(button_layout)
    
    def load_current_data(self):
        """加载当前订单数据"""
        try:
            # 设置基本信息
            self.name_edit.setText(self.order_info['OrderName'])
            
            # 设置日期
            from datetime import datetime
            start_date = datetime.strptime(self.order_info['StartDate'], "%Y-%m-%d").date()
            end_date = datetime.strptime(self.order_info['EndDate'], "%Y-%m-%d").date()
            self.start_date_edit.setDate(QDate(start_date.year, start_date.month, start_date.day))
            self.end_date_edit.setDate(QDate(end_date.year, end_date.month, end_date.day))
            
            # 设置备注
            self.remark_edit.setPlainText(self.order_info.get('Remark', ''))
            
            # 加载产品列表
            products = SchedulingOrderService.get_available_products()
            self.all_products = products
            
            # 获取当前订单的产品
            current_products = SchedulingOrderService.get_order_products(self.order_info['OrderId'])
            current_product_ids = {p['ItemId'] for p in current_products}
            
            self.product_table.setRowCount(len(products))
            
            for row, product in enumerate(products):
                # 复选框
                checkbox = QCheckBox()
                # 如果当前产品在订单中，则选中
                if product['ItemId'] in current_product_ids:
                    checkbox.setChecked(True)
                self.product_table.setCellWidget(row, 0, checkbox)
                
                # 产品信息
                item_text = f"{product['ItemCode']} - {product['CnName']}"
                if product.get('ItemSpec'):
                    item_text += f" ({product['ItemSpec']})"
                if product.get('Brand'):
                    item_text += f" [品牌: {product['Brand']}]"
                if product.get('ProjectName'):
                    item_text += f" [项目: {product['ProjectName']}]"
                
                item = QTableWidgetItem(item_text)
                item.setData(Qt.UserRole, product)
                self.product_table.setItem(row, 1, item)
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载订单数据失败: {str(e)}")
    
    def filter_products(self, text):
        """过滤产品列表"""
        for row in range(self.product_table.rowCount()):
            item = self.product_table.item(row, 1)
            if item:
                product = item.data(Qt.UserRole)
                if product:
                    item_text = f"{product['ItemCode']} - {product['CnName']}"
                    if product.get('ItemSpec'):
                        item_text += f" ({product['ItemSpec']})"
                    if product.get('Brand'):
                        item_text += f" [品牌: {product['Brand']}]"
                    if product.get('ProjectName'):
                        item_text += f" [项目: {product['ProjectName']}]"
                    
                    self.product_table.setRowHidden(row, text.lower() not in item_text.lower())
    
    def select_all_products(self):
        """全选产品"""
        for row in range(self.product_table.rowCount()):
            if not self.product_table.isRowHidden(row):
                checkbox = self.product_table.cellWidget(row, 0)
                if checkbox:
                    checkbox.setChecked(True)
    
    def clear_product_selection(self):
        """清空产品选择"""
        for row in range(self.product_table.rowCount()):
            checkbox = self.product_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(False)
    
    def get_order_data(self):
        """获取订单数据"""
        selected_products = []
        for row in range(self.product_table.rowCount()):
            checkbox = self.product_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                # 从表格项中获取产品数据
                item = self.product_table.item(row, 1)
                if item:
                    product = item.data(Qt.UserRole)
                    selected_products.append(product)
        
        return {
            "order_name": self.name_edit.text().strip(),
            "start_date": self.start_date_edit.date().toString("yyyy-MM-dd"),
            "end_date": self.end_date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.toPlainText().strip(),
            "selected_products": selected_products
        }
    
    def accept(self):
        """确认修改"""
        if not self.name_edit.text().strip():
            QMessageBox.warning(self, "警告", "请输入订单名称")
            return
        
        super().accept()


class ProductionMRPWidget(QWidget):
    """生产MRP计算页面"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_order_id = None
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)
        
        # 控制面板
        control_group = QGroupBox("MRP计算控制")
        control_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                color: #495057;
                border: 1px solid #ccc;
                margin-top: 5px;
                padding-top: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 5px;
                padding: 0 3px 0 3px;
            }
        """)
        control_layout = QVBoxLayout(control_group)
        
        # 订单选择
        order_layout = QHBoxLayout()
        order_layout.addWidget(QLabel("选择排产订单:"))
        self.order_combo = QComboBox()
        self.order_combo.setStyleSheet("""
            QComboBox {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
                min-width: 200px;
            }
            QComboBox:focus {
                border-color: #007bff;
            }
        """)
        self.order_combo.currentTextChanged.connect(self.on_order_changed)
        order_layout.addWidget(self.order_combo)
        order_layout.addStretch()
        control_layout.addLayout(order_layout)
        
        # 计算类型选择
        calc_layout = QHBoxLayout()
        calc_layout.addWidget(QLabel("计算类型:"))
        self.calc_type_combo = QComboBox()
        self.calc_type_combo.addItems(["综合MRP", "零部件MRP", "成品MRP"])
        self.calc_type_combo.setCurrentText("综合MRP")
        self.calc_type_combo.setStyleSheet("""
            QComboBox {
                padding: 4px 8px;
                border: 1px solid #ccc;
                font-size: 12px;
                min-width: 120px;
            }
            QComboBox:focus {
                border-color: #007bff;
            }
        """)
        calc_layout.addWidget(self.calc_type_combo)
        calc_layout.addStretch()
        control_layout.addLayout(calc_layout)
        
        # 计算类型说明
        type_desc_label = QLabel("综合MRP：结合成品库存和零部件库存计算；零部件MRP：展开BOM计算原材料需求；成品MRP：直接显示成品需求")
        type_desc_label.setStyleSheet("""
            QLabel {
                color: #6c757d;
                font-size: 11px;
                font-style: italic;
            }
        """)
        type_desc_label.setWordWrap(True)
        control_layout.addWidget(type_desc_label)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        self.calc_btn = QPushButton("计算MRP")
        self.calc_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
            QPushButton:disabled {
                background-color: #6c757d;
            }
        """)
        self.calc_btn.clicked.connect(self.calculate_mrp)
        self.calc_btn.setEnabled(False)
        
        self.refresh_btn = QPushButton("刷新数据")
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        self.refresh_btn.clicked.connect(self.refresh_data)
        
        button_layout.addWidget(self.calc_btn)
        button_layout.addWidget(self.refresh_btn)
        button_layout.addStretch()
        control_layout.addLayout(button_layout)
        
        main_layout.addWidget(control_group)
        
        # MRP结果表格
        result_group = QGroupBox("MRP计算结果")
        result_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                color: #495057;
                border: 1px solid #ccc;
                margin-top: 5px;
                padding-top: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 5px;
                padding: 0 3px 0 3px;
            }
        """)
        result_layout = QVBoxLayout(result_group)
        
        # 创建表格
        self.mrp_table = QTableWidget()
        self.mrp_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dee2e6;
                font-size: 12px;
                gridline-color: #dee2e6;
                background-color: white;
                selection-background-color: #e3f2fd;
            }
            QTableWidget::item {
                padding: 6px 8px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                border: 2px solid #007bff;
                border-radius: 2px;
            }
            QTableWidget::item:selected:focus {
                background-color: transparent !important;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                color: #495057;
                padding: 8px 6px;
                border: none;
                border-bottom: 2px solid #dee2e6;
                font-weight: 600;
                font-size: 12px;
            }
        """)
        
        # 设置自定义表头
        self.mrp_table.setHorizontalHeader(TwoRowHeader(Qt.Horizontal, self.mrp_table))
        
        # 禁用交替行颜色
        self.mrp_table.setAlternatingRowColors(False)
        
        # 设置表格属性
        self.mrp_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.mrp_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.mrp_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        
        result_layout.addWidget(self.mrp_table)
        main_layout.addWidget(result_group)
        
        # 加载订单列表
        self.load_orders()
    
    def load_orders(self):
        """加载排产订单列表"""
        try:
            orders = SchedulingOrderService.get_scheduling_orders()
            self.order_combo.clear()
            self.order_combo.addItem("请选择排产订单", None)
            
            for order in orders:
                display_text = f"{order['OrderName']} ({order['StartDate']} ~ {order['EndDate']})"
                self.order_combo.addItem(display_text, order['OrderId'])
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载订单列表失败: {str(e)}")
    
    def on_order_changed(self, text):
        """订单选择改变时的处理"""
        self.current_order_id = self.order_combo.currentData()
        self.calc_btn.setEnabled(self.current_order_id is not None)
        
        if self.current_order_id:
            self.load_mrp_data()
        else:
            self.clear_mrp_table()
    
    def calculate_mrp(self):
        """计算MRP"""
        if not self.current_order_id:
            QMessageBox.warning(self, "警告", "请先选择一个排产订单")
            return
        
        try:
            # 获取排产订单信息
            order_info = SchedulingOrderService.get_scheduling_order_by_id(self.current_order_id)
            if not order_info:
                QMessageBox.critical(self, "错误", "找不到指定的排产订单")
                return
            
            start_date = order_info["StartDate"]
            end_date = order_info["EndDate"]
            
            # 获取计算类型 - 与订单MRP管理保持一致
            calc_type_text = self.calc_type_combo.currentText()
            if calc_type_text == "零部件MRP":
                calc_type = "child"
            elif calc_type_text == "成品MRP":
                calc_type = "parent"
            else:  # 综合MRP
                calc_type = "comprehensive"
            
            print(f"🔘 [calculate_mrp] 计算类型：{calc_type}")
            
            # 根据计算类型调用不同的计算方法 - 与订单MRP管理保持一致
            if calc_type == "child":
                # 计算零部件MRP - 展开BOM计算原材料需求
                result = SchedulingOrderService.calculate_child_mrp_for_order(
                    self.current_order_id, start_date, end_date
                )
            elif calc_type == "parent":
                # 计算成品MRP - 直接显示成品需求
                result = SchedulingOrderService.calculate_parent_mrp_for_order(
                    self.current_order_id, start_date, end_date
                )
            else:  # comprehensive
                # 计算综合MRP - 结合成品库存和零部件库存计算
                result = SchedulingOrderService.calculate_comprehensive_mrp_for_order(
                    self.current_order_id, start_date, end_date
                )
            
            if "error" in result:
                QMessageBox.critical(self, "错误", result["error"])
                return
            
            # 显示结果
            self.display_mrp_results(result)
            QMessageBox.information(self, "成功", f"MRP计算完成，共计算 {len(result.get('rows', []))} 个物料")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"计算MRP失败: {str(e)}")
    
    def display_mrp_results(self, result):
        """显示MRP计算结果 - 与订单MRP管理保持一致"""
        try:
            # 检查数据格式 - 与订单MRP管理保持一致
            if not result or "weeks" not in result or "rows" not in result:
                self.clear_mrp_table()
                return
            
            weeks = result.get("weeks", [])
            rows = result.get("rows", [])
            
            if not rows:
                self.clear_mrp_table()
                return
            
            print(f"🎨 [display_mrp_results] 数据解析：weeks={weeks}, rows数量={len(rows)}")
            
            # 构建年份分组和合计列 - 与订单MRP管理保持一致
            colspec = self._build_week_columns_with_totals(weeks)
            
            # 设置固定列标题 - 显示产品信息和MRP信息
            fixed_headers = ["产品名称", "规格", "品牌", "项目名称", "行别", "期初库存"]
            
            # 设置列数和标题
            headers_count = len(fixed_headers) + len(colspec) + 1  # +1 for Total column
            self.mrp_table.setColumnCount(headers_count)
            
            # 设置固定列标题
            for i, title in enumerate(fixed_headers):
                item = QTableWidgetItem(title)
                self.mrp_table.setHorizontalHeaderItem(i, item)
            
            # 设置周列和年份合计列标题
            base_col = len(fixed_headers)
            for i, (kind, val) in enumerate(colspec):
                if kind == "week":
                    # val 现在是具体的订单日期 (YYYY-MM-DD)
                    # CW位置显示日期，日期位置显示周几
                    try:
                        from datetime import datetime
                        date_obj = datetime.strptime(val, "%Y-%m-%d").date()
                        # CW位置显示日期 (MM-DD格式)
                        date_text = date_obj.strftime("%m-%d")
                        it = QTableWidgetItem(date_text)
                        it.setData(Qt.UserRole, val)  # 存储具体的订单日期
                    except:
                        it = QTableWidgetItem(val)
                else:
                    it = QTableWidgetItem(f"{val}合计")
                self.mrp_table.setHorizontalHeaderItem(base_col + i, it)
            
            # 设置总计列标题
            self.mrp_table.setHorizontalHeaderItem(headers_count - 1, QTableWidgetItem("Total"))
            
            # 更新表头显示
            try:
                hdr = self.mrp_table.horizontalHeader()
                hdr.updateGeometry()
                hdr.repaint()
            except Exception as e:
                print(f"更新表头时出错: {e}")
            
            # 增加行用于显示总计行
            calc_type = self.calc_type_combo.currentText()
            if calc_type == "成品MRP":
                # 成品MRP：数据行 + 生产计划总计行 + 即时库存总计行
                self.mrp_table.setRowCount(len(rows) + 2)  # +2 for total rows
            else:
                # 零部件MRP：数据行 + 总计行
                self.mrp_table.setRowCount(len(rows) + 1)  # +1 for total row
            
            # 设置颜色 - 与订单MRP管理保持一致
            green_bg = QBrush(QColor(235, 252, 239))  # 生产计划绿色
            red_bg = QBrush(QColor(255, 235, 238))     # 库存不足红色
            blue_bg = QBrush(QColor(221, 235, 247))   # 合计列蓝色
            
            # 数据行（从第一行开始）
            for r, row in enumerate(rows):
                actual_row = r  # 数据行从第一行开始
                
                # 基本信息列 - 显示产品信息和MRP信息
                self._set_item(actual_row, 0, row.get("ItemName", ""))
                self._set_item(actual_row, 1, row.get("ItemSpec", ""))
                self._set_item(actual_row, 2, row.get("Brand", ""))  # 品牌字段
                self._set_item(actual_row, 3, row.get("ProjectName", ""))
                self._set_item(actual_row, 4, row.get("RowType", ""))  # 行别
                
                # 期初库存列：综合MRP显示"XXX+XXX"格式，其他显示数字
                start_onhand = row.get("StartOnHand", 0)
                if isinstance(start_onhand, str) and "+" in start_onhand:
                    # 综合MRP的"XXX+XXX"格式，直接显示
                    self._set_item(actual_row, 5, start_onhand)
                else:
                    # 其他类型，格式化为数字
                    self._set_item(actual_row, 5, self._fmt(start_onhand))
                
                # 基本信息列不设置背景色
                
                # 周数据列和年份合计列
                row_total = 0
                cursor_col = base_col
                for kind, val in colspec:
                    if kind == "week":
                        val_float = float(row["cells"].get(val, 0.0))
                        row_total += val_float
                        it = self._set_item(actual_row, cursor_col, self._fmt(val_float))
                        
                        # 着色规则 - 与订单MRP管理保持一致：
                        # 1. 生产计划行（非即时库存）且数值大于0时标绿色
                        # 2. 即时库存行且数值小于0时标红色
                        is_stock_row = (row.get("RowType") == "即时库存")
                        if not is_stock_row and val_float > 0:
                            it.setBackground(green_bg)  # 生产计划标绿
                        elif is_stock_row and val_float < 0:
                            it.setBackground(red_bg)    # 库存不足标红
                    else:
                        # 年份合计列 - val 现在是年份
                        # 需要计算该年份所有日期的总和
                        year_total = 0.0
                        for kind2, val2 in colspec:
                            if kind2 == "week":
                                # val2 是具体的订单日期
                                try:
                                    from datetime import datetime
                                    date_obj = datetime.strptime(val2, "%Y-%m-%d").date()
                                    if date_obj.isocalendar()[0] == val:  # 同一年
                                        year_total += float(row["cells"].get(val2, 0.0))
                                except:
                                    continue
                        
                        it = QTableWidgetItem(self._fmt(year_total))
                        it.setBackground(blue_bg)  # 合计列标蓝色
                        font = it.font()
                        font.setBold(True)
                        it.setFont(font)
                        self.mrp_table.setItem(actual_row, cursor_col, it)
                        row_total += year_total
                    
                    cursor_col += 1
                
                # 总计列
                total_item = QTableWidgetItem(self._fmt(row_total))
                total_item.setBackground(blue_bg)  # 总计列标蓝色
                font = total_item.font()
                font.setBold(True)
                total_item.setFont(font)
                self.mrp_table.setItem(actual_row, headers_count - 1, total_item)
            
            # 总计行 - 与订单MRP管理保持一致
            calc_type = self.calc_type_combo.currentText()
            if calc_type == "成品MRP":
                # 成品MRP：两行总计行
                # 第一行：生产计划总计
                plan_total_row = len(rows)
                self.mrp_table.setItem(plan_total_row, 0, QTableWidgetItem("生产计划TOTAL"))
                self.mrp_table.setItem(plan_total_row, 1, QTableWidgetItem(""))
                self.mrp_table.setItem(plan_total_row, 2, QTableWidgetItem(""))
                self.mrp_table.setItem(plan_total_row, 3, QTableWidgetItem(""))
                self.mrp_table.setItem(plan_total_row, 4, QTableWidgetItem("生产计划"))
                self.mrp_table.setItem(plan_total_row, 5, QTableWidgetItem(""))
                
                # 第二行：即时库存总计
                stock_total_row = len(rows) + 1
                self.mrp_table.setItem(stock_total_row, 0, QTableWidgetItem("即时库存TOTAL"))
                self.mrp_table.setItem(stock_total_row, 1, QTableWidgetItem(""))
                self.mrp_table.setItem(stock_total_row, 2, QTableWidgetItem(""))
                self.mrp_table.setItem(stock_total_row, 3, QTableWidgetItem(""))
                self.mrp_table.setItem(stock_total_row, 4, QTableWidgetItem("即时库存"))
                self.mrp_table.setItem(stock_total_row, 5, QTableWidgetItem(""))
                
                # 计算生产计划总计（只统计生产计划行）
                for col in range(base_col, headers_count):
                    plan_sum = 0
                    for r in range(0, plan_total_row):  # 从0开始，没有日期行
                        it = self.mrp_table.item(r, col)
                        row_type_it = self.mrp_table.item(r, 4)  # 行别列
                        try:
                            if it and it.text().strip() and row_type_it and row_type_it.text() == "生产计划":
                                plan_sum += float(it.text().replace(',', ''))
                        except:
                            pass
                    item = QTableWidgetItem(self._fmt(plan_sum))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    item.setBackground(green_bg)  # 生产计划总计标绿色
                    self.mrp_table.setItem(plan_total_row, col, item)
                
                # 计算即时库存总计（只统计即时库存行）
                for col in range(base_col, headers_count):
                    stock_sum = 0
                    for r in range(0, stock_total_row):  # 从0开始，没有日期行
                        it = self.mrp_table.item(r, col)
                        row_type_it = self.mrp_table.item(r, 4)  # 行别列
                        try:
                            if it and it.text().strip() and row_type_it and row_type_it.text() == "即时库存":
                                stock_sum += float(it.text().replace(',', ''))
                        except:
                            pass
                    item = QTableWidgetItem(self._fmt(stock_sum))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    item.setBackground(red_bg)  # 即时库存总计标红色
                    self.mrp_table.setItem(stock_total_row, col, item)
            else:
                # 零部件MRP：一行总计行
                total_row = len(rows)
                self.mrp_table.setItem(total_row, 0, QTableWidgetItem("TOTAL"))
                self.mrp_table.setItem(total_row, 1, QTableWidgetItem(""))
                self.mrp_table.setItem(total_row, 2, QTableWidgetItem(""))
                self.mrp_table.setItem(total_row, 3, QTableWidgetItem(""))
                self.mrp_table.setItem(total_row, 4, QTableWidgetItem(""))
                self.mrp_table.setItem(total_row, 5, QTableWidgetItem(""))
                
                # 计算总计行的数据
                for col in range(base_col, headers_count):
                    col_sum = 0
                    for r in range(len(rows)):
                        it = self.mrp_table.item(r, col)
                        try:
                            if it and it.text().strip():
                                col_sum += float(it.text().replace(',', ''))
                        except:
                            pass
                    
                    item = QTableWidgetItem(self._fmt(col_sum))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    item.setBackground(blue_bg)  # 总计行标蓝色
                    self.mrp_table.setItem(total_row, col, item)
            
            # 设置列宽
            header = self.mrp_table.horizontalHeader()
            
            # 设置固定列的宽度
            self.mrp_table.setColumnWidth(0, 150)  # 产品名称
            self.mrp_table.setColumnWidth(1, 100)  # 规格
            self.mrp_table.setColumnWidth(2, 100)  # 品牌
            self.mrp_table.setColumnWidth(3, 120)  # 项目名称
            self.mrp_table.setColumnWidth(4, 80)   # 行别
            self.mrp_table.setColumnWidth(5, 100)  # 期初库存
            
            # 设置日期列为固定宽度
            for c in range(len(fixed_headers), headers_count):
                header.setSectionResizeMode(c, QHeaderView.Fixed)
                self.mrp_table.setColumnWidth(c, 80)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"显示MRP结果失败: {str(e)}")
    
    def _set_item(self, row, col, text):
        """设置表格项 - 与订单MRP管理保持一致"""
        item = QTableWidgetItem(str(text))
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.mrp_table.setItem(row, col, item)
        return item
    
    def _build_week_columns_with_totals(self, weeks):
        """构建周列和年份合计列 - 与订单MRP管理保持一致"""
        colspec = []
        current_year = None
        
        for week in weeks:
            try:
                from datetime import datetime
                date_obj = datetime.strptime(week, "%Y-%m-%d").date()
                year = date_obj.isocalendar()[0]
                
                # 如果年份变化，添加年份合计列
                if current_year is not None and year != current_year:
                    colspec.append(("year_total", current_year))
                
                colspec.append(("week", week))
                current_year = year
            except:
                colspec.append(("week", week))
        
        # 添加最后一个年份的合计列
        if current_year is not None:
            colspec.append(("year_total", current_year))
        
        return colspec
    
    def _fmt(self, val):
        """格式化数字显示 - 与订单MRP保持一致"""
        try:
            if val is None:
                return "0"
            f = float(val)
            if f == 0:
                return "0"
            elif f == int(f):
                return str(int(f))
            else:
                return f"{f:.2f}".rstrip('0').rstrip('.')
        except:
            return str(val)
    
    def clear_mrp_table(self):
        """清空MRP表格"""
        self.mrp_table.clearContents()
        self.mrp_table.setRowCount(0)
        self.mrp_table.setColumnCount(0)
    
    def refresh_data(self):
        """刷新数据"""
        self.load_orders()
        if self.current_order_id:
            self.load_mrp_data()
    
    def load_mrp_data(self):
        """加载已保存的MRP数据"""
        if not self.current_order_id:
            return
        
        try:
            # 获取已保存的MRP结果
            mrp_results = SchedulingOrderService.get_mrp_results(self.current_order_id)
            
            if mrp_results:
                # 重新组织数据格式
                organized_data = self._organize_mrp_data(mrp_results)
                self.display_mrp_results(organized_data)
            else:
                self.clear_mrp_table()
                
        except Exception as e:
            print(f"加载MRP数据失败: {e}")
            self.clear_mrp_table()
    
    def _organize_mrp_data(self, mrp_results):
        """重新组织MRP数据格式"""
        try:
            # 按物料分组
            items_data = {}
            date_range = set()
            
            for result in mrp_results:
                item_id = result["ItemId"]
                if item_id not in items_data:
                    items_data[item_id] = {
                        "ItemId": item_id,
                        "ItemCode": result["ItemCode"],
                        "ItemName": result["ItemName"],
                        "ItemSpec": result["ItemSpec"],
                        "Brand": result["Brand"],  # 型号
                        "ItemType": result["ItemType"],
                        "cells": {}
                    }
                
                date_str = result["ProductionDate"]
                date_range.add(date_str)
                
                items_data[item_id]["cells"][date_str] = {
                    "RequiredQty": result["RequiredQty"],
                    "OnHandQty": result["OnHandQty"],
                    "NetQty": result["NetQty"]
                }
            
            # 转换为列表并排序
            mrp_list = []
            for item_id in sorted(items_data.keys(), 
                                key=lambda i: (items_data[i]["ItemType"], items_data[i]["ItemCode"])):
                mrp_list.append(items_data[item_id])
            
            return {
                "order_info": {"OrderName": "已保存的MRP数据"},
                "date_range": sorted(list(date_range)),
                "mrp_results": mrp_list
            }
            
        except Exception as e:
            print(f"组织MRP数据失败: {e}")
            return {"error": f"组织MRP数据失败: {str(e)}"}
