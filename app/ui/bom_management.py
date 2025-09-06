from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                               QPushButton, QTableWidget, QTableWidgetItem,
                               QFrame, QLineEdit, QComboBox, QSpinBox,
                               QDoubleSpinBox, QMessageBox, QTabWidget,
                               QHeaderView, QAbstractItemView, QGroupBox,
                               QFormLayout, QTextEdit, QDialog, QCheckBox,
                               QDialogButtonBox, QGridLayout, QSpacerItem,
                               QSizePolicy, QScrollArea, QSplitter, QDateEdit,
                               QTreeWidget, QTreeWidgetItem, QProgressBar)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QColor
from app.services.bom_service import BomService
from app.services.item_service import ItemService
from app.services.bom_matrix_import_service import BomMatrixImportService
from app.services.bom_history_service import BomHistoryService
from app.utils.resource_path import get_resource_path
import re
import os
import sys
import shutil
import openpyxl
from datetime import datetime
from PySide6.QtWidgets import QFileDialog


class BomEditorDialog(QDialog):
    """BOM编辑器对话框 - 支持无限层级嵌套的产品结构"""

    def __init__(self, parent=None, bom_data=None):
        super().__init__(parent)
        self.bom_data = bom_data
        # 修复SQLite Row对象的访问方式
        self.bom_id = bom_data['BomId'] if bom_data and 'BomId' in bom_data.keys() else None
        self.setWindowTitle("新增BOM" if not bom_data else "编辑BOM")
        self.resize(1000, 700)
        self.setMinimumSize(900, 600)
        self.setMaximumSize(1200, 800)
        self.setModal(True)
        self.setup_ui()
        if bom_data:
            self.load_bom_data()
        self.load_bom_lines()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # 标题
        self.title_label = QLabel("BOM编辑器")
        self.title_label.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #333;
                padding: 8px;
            }
        """)
        self.title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.title_label)

        # 创建分割器
        splitter = QSplitter(Qt.Vertical)
        splitter.setChildrenCollapsible(False)

        # 上半部分：BOM基础信息
        header_widget = self.create_bom_header_widget()
        splitter.addWidget(header_widget)

        # 下半部分：产品结构树形视图
        tree_widget = self.create_product_tree_widget()
        splitter.addWidget(tree_widget)

        # 设置分割器比例 - 调整结果区域更高
        splitter.setSizes([150, 550])
        layout.addWidget(splitter)

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 15, 0, 0)
        button_layout.setSpacing(20)
        button_layout.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: white;
                color: #666;
                border: 1px solid #ccc;
                padding: 10px 25px;
                border-radius: 5px;
                font-size: 14px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                border-color: #1890ff;
                color: #1890ff;
            }
        """)
        cancel_btn.clicked.connect(self.reject)

        save_btn = QPushButton("保存")
        save_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 10px 25px;
                border-radius: 5px;
                font-size: 14px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        save_btn.clicked.connect(self.save_bom)

        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(save_btn)

        layout.addLayout(button_layout)

        # 设置样式
        self.setStyleSheet("""
            QDialog {
                background: white;
            }
            QLineEdit, QComboBox, QTextEdit, QDoubleSpinBox {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 4px;
                font-size: 14px;
                min-width: 200px;
                min-height: 15px;
            }
            QLineEdit:focus, QComboBox:focus, QTextEdit:focus, QDoubleSpinBox:focus {
                border-color: #1890ff;
            }
        """)

    def create_bom_header_widget(self):
        """创建BOM基础信息组件"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(15)

        # BOM基础信息组
        header_group = QGroupBox("BOM基础信息")
        header_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 2px solid #e8e8e8;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-size: 14px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px 0 8px;
                color: #262626;
                font-weight: 600;
            }
        """)

        # 使用网格布局，两列显示
        header_layout = QGridLayout(header_group)
        header_layout.setSpacing(15)
        header_layout.setColumnStretch(1, 1)

        # 第一列
        # 父产品选择（支持搜索）
        self.parent_item_combo = QComboBox()
        self.parent_item_combo.setEditable(True)  # 设置为可编辑
        self.parent_item_combo.setInsertPolicy(QComboBox.NoInsert)  # 禁止插入新项
        self.parent_item_combo.setPlaceholderText("请输入或选择父产品（成品）")
        
        # 设置QCompleter实现包含式匹配
        from PySide6.QtWidgets import QCompleter
        from PySide6.QtCore import Qt
        completer = QCompleter()
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        self.parent_item_combo.setCompleter(completer)
        
        self.parent_item_combo.setStyleSheet("""
            QComboBox {
                padding: 4px 8px;
                min-height: 20px;
                max-height: 24px;
                border: 1px solid #d9d9d9;
                border-radius: 4px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #40a9ff;
            }
            QComboBox:focus {
                border-color: #1890ff;
                outline: none;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
                background-color: #f5f5f5;
                border-left: 1px solid #d9d9d9;
            }
            QComboBox::drop-down:hover {
                background-color: #e6f7ff;
            }
            QComboBox::down-arrow {
                width: 0;
                height: 0;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #666;
                margin-right: 6px;
            }
            QComboBox::down-arrow:hover {
                border-top-color: #1890ff;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #d9d9d9;
                border-radius: 4px;
                background-color: white;
                selection-background-color: #e6f7ff;
                selection-color: #262626;
                outline: none;
            }
            QComboBox QAbstractItemView::item {
                padding: 6px 8px;
                border-bottom: 1px solid #f0f0f0;
                min-height: 20px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #f0f0f0;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #e6f7ff;
                color: #1890ff;
            }
        """)
        
        # 连接信号
        self.parent_item_combo.currentIndexChanged.connect(self.on_parent_item_selected)
        self.parent_item_combo.lineEdit().textChanged.connect(self.on_parent_item_search_text_changed)
        self.parent_item_combo.activated.connect(self.on_parent_item_activated)
        
        header_layout.addWidget(QLabel("父产品 *:"), 0, 0)
        header_layout.addWidget(self.parent_item_combo, 0, 1)
        
        # 存储所有父产品数据
        self.all_parent_items = []
        
        # 加载父产品选项
        self.load_parent_items()

        # BOM名称
        self.bom_name_edit = QLineEdit()
        self.bom_name_edit.setPlaceholderText("请输入BOM名称，如：产品A的BOM结构")
        self.bom_name_edit.setMaxLength(100)
        self.bom_name_edit.setStyleSheet("QLineEdit { padding: 4px 8px; min-height: 20px; max-height: 24px; }")
        header_layout.addWidget(QLabel("BOM名称 *:"), 1, 0)
        header_layout.addWidget(self.bom_name_edit, 1, 1)

        # 版本号
        self.rev_edit = QLineEdit()
        self.rev_edit.setPlaceholderText("请输入版本号，如：A、B、1.0等")
        self.rev_edit.setMaxLength(20)
        self.rev_edit.setStyleSheet("QLineEdit { padding: 4px 8px; min-height: 20px; max-height: 24px; }")
        header_layout.addWidget(QLabel("版本号 *:"), 2, 0)
        header_layout.addWidget(self.rev_edit, 2, 1)

        # 生效日期
        self.effective_date_edit = QDateEdit()
        self.effective_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.effective_date_edit.setDate(QDate.currentDate())
        self.effective_date_edit.setCalendarPopup(True)
        self.effective_date_edit.setStyleSheet("QDateEdit { padding: 4px 8px; min-height: 20px; max-height: 24px; }")
        header_layout.addWidget(QLabel("生效日期 *:"), 3, 0)
        header_layout.addWidget(self.effective_date_edit, 3, 1)

        # 第二列
        # 失效日期
        self.expire_date_edit = QDateEdit()
        self.expire_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.expire_date_edit.setCalendarPopup(True)
        self.expire_date_edit.setDate(QDate.currentDate().addYears(10))
        self.expire_date_edit.setStyleSheet("QDateEdit { padding: 4px 8px; min-height: 20px; max-height: 24px; }")
        header_layout.addWidget(QLabel("失效日期:"), 0, 2)
        header_layout.addWidget(self.expire_date_edit, 0, 3)

        # 备注
        self.remark_edit = QTextEdit()
        self.remark_edit.setMaximumHeight(40)
        self.remark_edit.setPlaceholderText("请输入备注信息")
        self.remark_edit.setStyleSheet("QTextEdit { padding: 4px 8px; min-height: 20px; max-height: 32px; }")
        header_layout.addWidget(QLabel("备注:"), 1, 2)
        header_layout.addWidget(self.remark_edit, 1, 3, 3, 1)

        layout.addWidget(header_group)

        return widget

    def create_product_tree_widget(self):
        """创建产品结构树形视图组件"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(15)

        # 产品结构组
        tree_group = QGroupBox("产品结构")
        tree_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 2px solid #e8e8e8;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-size: 14px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px 0 8px;
                color: #262626;
                font-weight: 600;
            }
        """)

        tree_layout = QVBoxLayout(tree_group)

        # 按钮栏
        button_frame = QFrame()
        button_layout = QHBoxLayout(button_frame)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(8)

        # 添加子物料按钮
        self.add_child_btn = QPushButton("+ 添加子物料")
        self.add_child_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
                min-width: 100px;
                max-height: 28px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        self.add_child_btn.clicked.connect(self.add_child_material)

        # 展开所有按钮
        expand_all_btn = QPushButton("展开所有")
        expand_all_btn.setStyleSheet("""
            QPushButton {
                background: #52c41a;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
                min-width: 80px;
                max-height: 28px;
            }
            QPushButton:hover {
                background: #73d13d;
            }
        """)
        expand_all_btn.clicked.connect(self.expand_all_nodes)

        # 收起所有按钮
        collapse_all_btn = QPushButton("收起所有")
        collapse_all_btn.setStyleSheet("""
            QPushButton {
                background: #faad14;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
                min-width: 80px;
                max-height: 28px;
            }
            QPushButton:hover {
                background: #ffc53d;
            }
        """)
        collapse_all_btn.clicked.connect(self.collapse_all_nodes)

        button_layout.addWidget(self.add_child_btn)
        button_layout.addWidget(expand_all_btn)
        button_layout.addWidget(collapse_all_btn)
        button_layout.addStretch()

        tree_layout.addWidget(button_frame)

        # 产品结构树形视图
        self.product_tree = QTreeWidget()
        self.product_tree.setHeaderLabels(["编码", "名称", "规格", "用量", "损耗率", "备注", "操作"])
        self.product_tree.setIndentation(20)
        self.product_tree.setStyleSheet("""
            QTreeWidget {
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                background-color: white;
                font-size: 13px;
            }
            QTreeWidget::item {
                padding: 4px 0px;
                border-bottom: 1px solid #f0f0f0;
                min-height: 30px;
            }
            QTreeWidget::item:hover {
                background-color: #f0f0f0;
            }
            QTreeWidget::item:selected {
                background-color: #e6f7ff;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                color: #262626;
                padding: 6px 8px;
                border: none;
                border-bottom: 2px solid #e8e8e8;
                font-weight: 600;
                font-size: 13px;
            }
        """)

        # 设置列宽
        header = self.product_tree.header()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 编码
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # 名称
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 规格
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 用量
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 损耗率
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 备注
        header.setSectionResizeMode(6, QHeaderView.Fixed)  # 操作

        self.product_tree.setColumnWidth(6, 100)  # 操作列宽度
        
        # 设置树形视图为可编辑
        self.product_tree.setEditTriggers(QTreeWidget.DoubleClicked | QTreeWidget.EditKeyPressed)
        
        # 连接编辑完成信号
        self.product_tree.itemChanged.connect(self.on_tree_item_changed)

        tree_layout.addWidget(self.product_tree)
        layout.addWidget(tree_group)

        return widget

    def load_parent_items(self):
        """加载父产品选项（只显示启用的成品）"""
        try:
            # 设置更新标志
            self._updating_combo = True
            
            # 获取所有启用的成品和半成品
            items = ItemService.get_items_by_type_with_status('FG')  # 成品
            sfg_items = ItemService.get_items_by_type_with_status('SFG')  # 半成品
            
            # 合并列表
            all_items = items + sfg_items
            
            # 存储所有父产品数据
            self.all_parent_items = all_items
            
            
            # 清空现有选项
            self.parent_item_combo.clear()
            
            # 添加选项
            for item in all_items:
                self.add_parent_item_to_combo(item)
            
            # 重置更新标志
            self._updating_combo = False
                
        except Exception as e:
            self._updating_combo = False

    def on_parent_item_activated(self, index):
        """父产品被激活时的处理（点击下拉列表项）"""
        try:
            print(f"[DEBUG] on_parent_item_activated 被调用: index={index}, _updating_combo={getattr(self, '_updating_combo', False)}")
            
            # 如果正在更新组合框，不处理
            if getattr(self, '_updating_combo', False):
                print(f"[DEBUG] 跳过处理，因为 _updating_combo=True")
                return
            
            if index >= 0:
                # 使用itemData(index)而不是currentData()
                parent_item_id = self.parent_item_combo.itemData(index)
                print(f"[DEBUG] parent_item_id: {parent_item_id}")
                
                if parent_item_id:
                    print(f"[DEBUG] 调用 generate_bom_name_from_parent({parent_item_id})")
                    # 根据父产品自动生成BOM名称
                    self.generate_bom_name_from_parent(parent_item_id)
                else:
                    print(f"[DEBUG] parent_item_id 为空，跳过处理")
            else:
                print(f"[DEBUG] index < 0，跳过处理")
                    
        except Exception as e:
            print(f"[DEBUG] on_parent_item_activated 异常: {e}")
            pass
    
    def on_parent_item_search_text_changed(self, text):
        """父产品搜索文本变化时的处理"""
        try:
            # 如果正在更新组合框，不处理
            if hasattr(self, '_updating_combo') and self._updating_combo:
                return
                
            # 去抖处理：200ms后执行过滤
            if hasattr(self, '_filter_timer'):
                self._filter_timer.stop()
            
            from PySide6.QtCore import QTimer
            self._filter_timer = QTimer()
            self._filter_timer.setSingleShot(True)
            self._filter_timer.timeout.connect(lambda: self.debounced_filter_parent_items(text))
            self._filter_timer.start(200)
            
        except Exception as e:
            pass
    
    def debounced_filter_parent_items(self, search_text):
        """去抖后的父产品筛选，保护焦点和光标位置"""
        try:
            print(f"[DEBUG] debounced_filter_parent_items 被调用: search_text='{search_text}'")
            
            # 记录当前输入框的状态
            line_edit = self.parent_item_combo.lineEdit()
            current_text = line_edit.text()
            cursor_position = line_edit.cursorPosition()
            has_focus = line_edit.hasFocus()
            
            print(f"[DEBUG] 当前状态 - 文本: '{current_text}', 光标: {cursor_position}, 焦点: {has_focus}")
            
            # 设置更新标志
            self._updating_combo = True
            
            # 阻断lineEdit的信号，避免setText触发二次textChanged
            line_edit.blockSignals(True)
            
            # 清空现有选项
            self.parent_item_combo.clear()
            
            # 恢复输入框的文本
            line_edit.setText(current_text)
            
            # 恢复信号
            line_edit.blockSignals(False)
            
            matched_items = []
            
            if not search_text or not search_text.strip():
                # 显示所有选项
                print(f"[DEBUG] 显示所有选项，共 {len(self.all_parent_items)} 个")
                matched_items = self.all_parent_items
                for item in self.all_parent_items:
                    self.add_parent_item_to_combo(item)
            else:
                # 过滤匹配的选项
                search_lower = search_text.lower().strip()
                
                for item in self.all_parent_items:
                    item_code = item.get('ItemCode', '').lower()
                    item_name = item.get('CnName', '').lower()
                    item_spec = item.get('ItemSpec', '').lower()
                    item_brand = item.get('Brand', '').lower()
                    
                    # 检查是否在任意字段中匹配
                    if (search_lower in item_code or 
                        search_lower in item_name or 
                        search_lower in item_spec or 
                        search_lower in item_brand):
                        matched_items.append(item)
                
                print(f"[DEBUG] 找到 {len(matched_items)} 个匹配项")
                # 添加匹配的选项
                for item in matched_items:
                    self.add_parent_item_to_combo(item)
            
            # 重置更新标志
            self._updating_combo = False
            
            # 恢复光标位置
            line_edit.setCursorPosition(cursor_position)
            
            # 如果之前有焦点，重新设置焦点
            if has_focus:
                from PySide6.QtCore import QTimer
                QTimer.singleShot(1, lambda: line_edit.setFocus())
            
            # 如果有匹配项且用户正在输入，显示下拉列表
            if search_text and search_text.strip() and matched_items and has_focus:
                from PySide6.QtCore import QTimer
                QTimer.singleShot(50, lambda: self.parent_item_combo.showPopup())
            
            print(f"[DEBUG] debounced_filter_parent_items 完成")
            
        except Exception as e:
            print(f"[DEBUG] debounced_filter_parent_items 异常: {e}")
            self._updating_combo = False
            if hasattr(self, 'parent_item_combo'):
                self.parent_item_combo.lineEdit().blockSignals(False)
    
    def filter_parent_items(self, search_text):
        """根据搜索文本过滤父产品选项"""
        try:
            print(f"[DEBUG] filter_parent_items 被调用: search_text='{search_text}'")
            
            # 设置更新标志
            self._updating_combo = True
            
            # 保存当前输入框的文本
            current_text = self.parent_item_combo.lineEdit().text()
            print(f"[DEBUG] 当前输入框文本: '{current_text}'")
            
            # 使用blockSignals来避免触发信号
            self.parent_item_combo.blockSignals(True)
            
            # 清空现有选项
            self.parent_item_combo.clear()
            
            # 恢复输入框的文本
            self.parent_item_combo.lineEdit().setText(current_text)
            
            # 恢复信号
            self.parent_item_combo.blockSignals(False)
            
            if not search_text or not search_text.strip():
                # 显示所有选项
                print(f"[DEBUG] 显示所有选项，共 {len(self.all_parent_items)} 个")
                for item in self.all_parent_items:
                    self.add_parent_item_to_combo(item)
            else:
                # 过滤匹配的选项
                search_lower = search_text.lower().strip()
                matched_items = []
                
                for item in self.all_parent_items:
                    item_code = item.get('ItemCode', '').lower()
                    item_name = item.get('CnName', '').lower()
                    item_spec = item.get('ItemSpec', '').lower()
                    item_brand = item.get('Brand', '').lower()
                    
                    # 检查是否在任意字段中匹配
                    if (search_lower in item_code or 
                        search_lower in item_name or 
                        search_lower in item_spec or 
                        search_lower in item_brand):
                        matched_items.append(item)
                
                print(f"[DEBUG] 找到 {len(matched_items)} 个匹配项")
                # 添加匹配的选项
                for item in matched_items:
                    self.add_parent_item_to_combo(item)
            
            # 重置更新标志
            self._updating_combo = False
            print(f"[DEBUG] filter_parent_items 完成，_updating_combo=False")
            
        except Exception as e:
            print(f"[DEBUG] filter_parent_items 异常: {e}")
            self._updating_combo = False
            self.parent_item_combo.blockSignals(False)
    
    def add_parent_item_to_combo(self, item):
        """添加父产品到组合框"""
        try:
            item_code = item.get('ItemCode', '')
            item_name = item.get('CnName', '')
            item_spec = item.get('ItemSpec', '')
            item_brand = item.get('Brand', '')
            item_id = item.get('ItemId', 0)
            
            # 显示格式：品牌 - 名称 - 规格 (编码)
            display_text = f"{item_brand} - {item_name} - {item_spec} ({item_code})"
            
            self.parent_item_combo.addItem(display_text, item_id)
            
        except Exception as e:
            pass
    
    def on_parent_item_selected(self, index):
        """父产品选择时的处理"""
        try:
            print(f"[DEBUG] on_parent_item_selected 被调用: index={index}, _updating_combo={getattr(self, '_updating_combo', False)}")
            
            # 如果正在更新组合框，不处理
            if getattr(self, '_updating_combo', False):
                print(f"[DEBUG] 跳过处理，因为 _updating_combo=True")
                return
            
            # 如果索引无效，尝试从当前文本获取
            if index < 0:
                print(f"[DEBUG] index < 0，尝试从当前文本获取父产品")
                current_text = self.parent_item_combo.currentText()
                print(f"[DEBUG] 当前文本: '{current_text}'")
                
                # 从当前文本中提取父产品ID
                parent_item_id = self.get_parent_item_id_from_text(current_text)
                print(f"[DEBUG] 从文本提取的parent_item_id: {parent_item_id}")
                
                if parent_item_id:
                    print(f"[DEBUG] 调用 generate_bom_name_from_parent({parent_item_id})")
                    self.generate_bom_name_from_parent(parent_item_id)
                else:
                    print(f"[DEBUG] 无法从文本提取parent_item_id")
                return
            
            # 使用itemData(index)而不是currentData()
            parent_item_id = self.parent_item_combo.itemData(index)
            print(f"[DEBUG] parent_item_id: {parent_item_id}")
            
            if parent_item_id:
                print(f"[DEBUG] 调用 generate_bom_name_from_parent({parent_item_id})")
                # 根据父产品自动生成BOM名称
                self.generate_bom_name_from_parent(parent_item_id)
            else:
                print(f"[DEBUG] parent_item_id 为空，跳过处理")
                    
        except Exception as e:
            print(f"[DEBUG] on_parent_item_selected 异常: {e}")
            pass
    
    def on_parent_item_search_changed(self, text):
        """父产品搜索变化时的处理（兼容性方法）"""
        pass
    
    def get_parent_item_id_from_text(self, text):
        """从显示文本中提取父产品ID"""
        try:
            print(f"[DEBUG] get_parent_item_id_from_text 被调用: text='{text}'")
            
            # 遍历所有父产品，找到匹配的显示文本
            for item in self.all_parent_items:
                item_code = item.get('ItemCode', '')
                item_name = item.get('CnName', '')
                item_spec = item.get('ItemSpec', '')
                item_brand = item.get('Brand', '')
                item_id = item.get('ItemId', 0)
                
                # 生成显示格式
                display_text = f"{item_brand} - {item_name} - {item_spec} ({item_code})"
                
                if display_text == text:
                    print(f"[DEBUG] 找到匹配的父产品: ID={item_id}, 文本='{display_text}'")
                    return item_id
            
            print(f"[DEBUG] 未找到匹配的父产品")
            return None
            
        except Exception as e:
            print(f"[DEBUG] get_parent_item_id_from_text 异常: {e}")
            return None
    
    def generate_bom_name_from_parent(self, parent_item_id):
        """根据父产品ID生成BOM名称"""
        try:
            print(f"[DEBUG] generate_bom_name_from_parent 被调用: parent_item_id={parent_item_id}")
            
            # 从存储的数据中查找父产品信息
            parent_item = None
            for item in self.all_parent_items:
                if item.get('ItemId') == parent_item_id:
                    parent_item = item
                    break
            
            print(f"[DEBUG] 找到的父产品: {parent_item}")
            
            if parent_item:
                item_brand = parent_item.get('Brand', '')
                print(f"[DEBUG] 商品品牌: '{item_brand}'")
                
                # 使用商品品牌作为BOM名称
                if item_brand:
                    bom_name = item_brand
                else:
                    # 如果品牌为空，使用产品名称
                    item_name = parent_item.get('CnName', '')
                    bom_name = item_name
                
                print(f"[DEBUG] 生成的BOM名称: '{bom_name}'")
                self.bom_name_edit.setText(bom_name)
                
                # 自动生成版本号
                if not self.rev_edit.text():
                    self.rev_edit.setText("A")
                    print(f"[DEBUG] 设置版本号为: A")
                
                print(f"[DEBUG] BOM名称设置完成: {self.bom_name_edit.text()}")
            else:
                print(f"[DEBUG] 未找到ID为 {parent_item_id} 的父产品")
                    
        except Exception as e:
            print(f"[DEBUG] generate_bom_name_from_parent 异常: {e}")
            pass

    def load_bom_data(self):
        """加载BOM数据到表单"""
        if not self.bom_data:
            return

        try:
            # 将Row对象转换为字典
            if hasattr(self.bom_data, 'keys'):
                bom_data = dict(self.bom_data)
            else:
                bom_data = self.bom_data
            
            print(f"[DEBUG] 加载BOM数据: {bom_data}")
            
            # 设置父产品
            parent_item_id = bom_data.get('ParentItemId')
            if parent_item_id:
                print(f"[DEBUG] 设置父产品ID: {parent_item_id}")
                # 查找父产品在组合框中的位置
                for i in range(self.parent_item_combo.count()):
                    if self.parent_item_combo.itemData(i) == parent_item_id:
                        self.parent_item_combo.setCurrentIndex(i)
                        print(f"[DEBUG] 找到父产品，设置索引: {i}")
                        break
            
            # 设置BOM名称
            bom_name = bom_data.get('BomName', '')
            self.bom_name_edit.setText(str(bom_name))
            print(f"[DEBUG] 设置BOM名称: {bom_name}")

            # 设置版本号
            rev = bom_data.get('Rev', '')
            self.rev_edit.setText(str(rev))
            print(f"[DEBUG] 设置版本号: {rev}")
            
            # 设置生效日期
            effective_date = bom_data.get('EffectiveDate')
            if effective_date:
                try:
                    date_obj = QDate.fromString(str(effective_date), "yyyy-MM-dd")
                    if date_obj.isValid():
                        self.effective_date_edit.setDate(date_obj)
                        print(f"[DEBUG] 设置生效日期: {effective_date}")
                except Exception as e:
                    print(f"[DEBUG] 设置生效日期失败: {e}")
            
            # 设置失效日期
            expire_date = bom_data.get('ExpireDate')
            if expire_date:
                try:
                    date_obj = QDate.fromString(str(expire_date), "yyyy-MM-dd")
                    if date_obj.isValid():
                        self.expire_date_edit.setDate(date_obj)
                        print(f"[DEBUG] 设置失效日期: {expire_date}")
                except Exception as e:
                    print(f"[DEBUG] 设置失效日期失败: {e}")
            
            # 设置备注
            remark = bom_data.get('Remark', '')
            self.remark_edit.setPlainText(str(remark))
            print(f"[DEBUG] 设置备注: {remark}")

            # 更新标题显示父产品信息
            self.update_title_with_parent_info()

        except Exception as e:
            print(f"加载BOM数据失败: {e}")

    def update_title_with_parent_info(self):
        """更新标题显示父产品信息"""
        try:
            if not self.bom_data:
                return
                
            # 将Row对象转换为字典
            if hasattr(self.bom_data, 'keys'):
                bom_data = dict(self.bom_data)
            else:
                bom_data = self.bom_data
                
            parent_item_code = bom_data.get('ParentItemCode', '')
            parent_item_name = bom_data.get('ParentItemName', '')
            parent_item_spec = bom_data.get('ParentItemSpec', '')
            parent_item_brand = bom_data.get('ParentItemBrand', '')
            
            # 构建标题信息
            title_parts = ["BOM编辑器"]
            if parent_item_code or parent_item_name:
                product_info_parts = []
                if parent_item_code:
                    product_info_parts.append(parent_item_code)
                if parent_item_name:
                    product_info_parts.append(parent_item_name)
                if parent_item_spec:
                    product_info_parts.append(f"规格: {parent_item_spec}")
                
                title_parts.append(f"({' - '.join(product_info_parts)})")
            
            title_text = " - ".join(title_parts)
            self.title_label.setText(title_text)
            
        except Exception as e:
            print(f"更新标题失败: {e}")

    def load_bom_lines(self):
        """加载BOM明细列表到树形视图"""
        if not self.bom_id:
            return

        try:
            bom_lines = BomService.get_bom_lines(self.bom_id)
            self.populate_product_tree(bom_lines)
        except Exception as e:
            print(f"加载BOM明细失败: {e}")

    def populate_product_tree(self, bom_lines):
        """填充产品结构树形视图"""
        self.product_tree.clear()
        
        for line in bom_lines:
            # 修复SQLite Row对象的访问方式
            child_item_code = line['ChildItemCode'] if 'ChildItemCode' in line.keys() else ''
            child_item_name = line['ChildItemName'] if 'ChildItemName' in line.keys() else ''
            child_item_spec = line['ChildItemSpec'] if 'ChildItemSpec' in line.keys() else ''
            child_item_brand = line['ChildItemBrand'] if 'ChildItemBrand' in line.keys() else ''
            qty_per = line['QtyPer'] if 'QtyPer' in line.keys() else 0
            scrap_factor = line['ScrapFactor'] if 'ScrapFactor' in line.keys() else 0
            remark = line['Remark'] if 'Remark' in line.keys() else ''
            
            # 创建产品节点
            product_item = QTreeWidgetItem(self.product_tree)
            
            # 编码列
            product_item.setText(0, str(child_item_code))
            
            # 名称列
            product_item.setText(1, str(child_item_name))
            
            # 规格列
            product_item.setText(2, str(child_item_spec) if child_item_spec else "")
            
            # 用量列
            product_item.setText(3, str(qty_per))
            
            # 损耗率列
            scrap_factor_display = scrap_factor * 100 if scrap_factor else 0
            product_item.setText(4, f"{scrap_factor_display:.1f}%")
            
            # 备注列
            product_item.setText(5, str(remark) if remark else "")
            
            # 操作列
            self.create_tree_operation_widget(product_item, line)

    def create_tree_operation_widget(self, tree_item, line_data):
        """为树形视图项创建操作控件"""
        operation_widget = QWidget()
        operation_layout = QHBoxLayout(operation_widget)
        operation_layout.setContentsMargins(1, 1, 1, 1)
        operation_layout.setSpacing(2)

        # 添加子物料按钮
        add_child_btn = QPushButton("+ 新增")
        add_child_btn.setStyleSheet("""
            QPushButton {
                background: #52c41a;
                color: white;
                border: none;
                padding: 5px 8px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                min-width: 20px;
                min-height: 16px;
            }
            QPushButton:hover {
                background: #73d13d;
            }
        """)
        add_child_btn.clicked.connect(lambda: self.add_child_to_node(tree_item, line_data))

        # 编辑按钮
        edit_btn = QPushButton("编辑")
        edit_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 5px 8px;
                border-radius: 4px;
                font-size: 11px;
                min-width: 20px;
                min-height: 16px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        edit_btn.clicked.connect(lambda: self.edit_material_node(tree_item, line_data))

        # 删除按钮
        delete_btn = QPushButton("删除")
        delete_btn.setStyleSheet("""
            QPushButton {
                background: #ff4d4f;
                color: white;
                border: none;
                padding: 5px 8px;
                border-radius: 4px;
                font-size: 11px;
                min-width: 20px;
                min-height: 16px;
            }
            QPushButton:hover {
                background: #ff7875;
            }
        """)
        delete_btn.clicked.connect(lambda: self.delete_material_node(tree_item))

        operation_layout.addWidget(add_child_btn)
        operation_layout.addWidget(edit_btn)
        operation_layout.addWidget(delete_btn)

        self.product_tree.setItemWidget(tree_item, 6, operation_widget)

    def add_child_material(self):
        """添加子物料到根节点"""
        self.add_material_dialog = MaterialSelectionDialog(self)
        if self.add_material_dialog.exec() == QDialog.Accepted:
            materials_data = self.add_material_dialog.get_selected_materials()
            for material_data in materials_data:
                self.add_material_to_tree(None, material_data)

    def add_child_to_node(self, parent_item, parent_line_data):
        """在指定节点下添加子物料"""
        self.add_material_dialog = MaterialSelectionDialog(self)
        if self.add_material_dialog.exec() == QDialog.Accepted:
            materials_data = self.add_material_dialog.get_selected_materials()
            for material_data in materials_data:
                self.add_material_to_tree(parent_item, material_data)

    def add_material_to_tree(self, parent_item, material_data):
        """添加物料到树形视图"""
        # 创建新的物料节点
        if parent_item:
            new_item = QTreeWidgetItem(parent_item)
        else:
            new_item = QTreeWidgetItem(self.product_tree)

        # 设置物料信息
        new_item.setText(0, str(material_data.get('ItemCode', '')))  # 编码
        new_item.setText(1, str(material_data.get('CnName', '')))  # 名称
        new_item.setText(2, str(material_data.get('ItemSpec', '')))  # 规格
        new_item.setText(3, "1.0")  # 默认用量
        new_item.setText(4, "0.0%")  # 默认损耗率
        new_item.setText(5, "")  # 备注

        # 设置物料数据到节点，用于后续保存
        new_item.setData(0, Qt.UserRole, {
            'ItemId': material_data['ItemId'] if 'ItemId' in material_data.keys() else 0,
            'ItemCode': material_data['ItemCode'] if 'ItemCode' in material_data.keys() else '',
            'ItemName': material_data['CnName'] if 'CnName' in material_data.keys() else '',
            'ItemSpec': material_data['ItemSpec'] if 'ItemSpec' in material_data.keys() else '',
            'Brand': material_data['Brand'] if 'Brand' in material_data.keys() else '',
            'ItemType': material_data['ItemType'] if 'ItemType' in material_data.keys() else ''
        })

        # 设置节点为可编辑
        new_item.setFlags(new_item.flags() | Qt.ItemIsEditable)

        # 创建操作控件
        self.create_tree_operation_widget(new_item, material_data)

        # 展开父节点
        if parent_item:
            parent_item.setExpanded(True)

    def edit_material_node(self, tree_item, line_data):
        """编辑物料节点"""
        try:
            # 获取当前节点的数据
            current_qty = tree_item.text(3)  # 用量列
            current_scrap = tree_item.text(4)  # 损耗率列
            current_remark = tree_item.text(5)  # 备注列
            
            # 创建编辑对话框
            dialog = MaterialEditDialog(self, current_qty, current_scrap, current_remark)
            if dialog.exec() == QDialog.Accepted:
                # 获取编辑后的数据
                new_qty, new_scrap, new_remark = dialog.get_edited_data()
                
                # 更新节点数据
                tree_item.setText(3, str(new_qty))
                tree_item.setText(4, f"{new_scrap:.1f}%")
                tree_item.setText(5, str(new_remark))
                
                QMessageBox.information(self, "成功", "物料信息更新成功！")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"编辑物料失败: {str(e)}")

    def delete_material_node(self, tree_item):
        """删除物料节点"""
        reply = QMessageBox.question(
            self, "确认删除",
            "确定要删除该物料吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # 删除节点及其所有子节点
            parent = tree_item.parent()
            if parent:
                parent.removeChild(tree_item)
            else:
                self.product_tree.takeTopLevelItem(self.product_tree.indexOfTopLevelItem(tree_item))

    def expand_all_nodes(self):
        """展开所有节点"""
        self.product_tree.expandAll()

    def collapse_all_nodes(self):
        """收起所有节点"""
        self.product_tree.collapseAll()

    def on_tree_item_changed(self, item, column):
        """处理树形视图项编辑完成事件"""
        try:
            if column == 3:  # 用量列
                # 验证用量是否为有效数字
                qty_text = item.text(3)
                try:
                    qty = float(qty_text)
                    if qty <= 0:
                        item.setText(3, "1.0")
                        QMessageBox.warning(self, "警告", "用量必须大于0！")
                except ValueError:
                    item.setText(3, "1.0")
                    QMessageBox.warning(self, "警告", "用量必须是有效的数字！")
                    
            elif column == 4:  # 损耗率列
                # 验证损耗率是否为有效数字
                scrap_text = item.text(4)
                try:
                    # 移除百分号并转换为小数
                    if "%" in scrap_text:
                        scrap_text = scrap_text.replace("%", "")
                    scrap = float(scrap_text)
                    if scrap < 0:
                        item.setText(4, "0.0%")
                        QMessageBox.warning(self, "警告", "损耗率不能为负数！")
                    elif scrap > 100:
                        item.setText(4, "100.0%")
                        QMessageBox.warning(self, "警告", "损耗率不能超过100%！")
                    else:
                        # 重新格式化显示
                        item.setText(4, f"{scrap:.1f}%")
                except ValueError:
                    item.setText(4, "0.0%")
                    QMessageBox.warning(self, "警告", "损耗率必须是有效的数字！")
                    
        except Exception as e:
            print(f"处理树形视图项编辑失败: {e}")

    def get_bom_data(self):
        """获取BOM主表数据"""
        # 获取日期
        effective_date = self.effective_date_edit.date().toString("yyyy-MM-dd")
        expire_date = self.expire_date_edit.date().toString("yyyy-MM-dd") if self.expire_date_edit.date().isValid() else None
        
        # 获取父产品ID
        parent_item_id = self.parent_item_combo.currentData()
        
        # 如果currentData()返回None，尝试从当前文本获取
        if not parent_item_id:
            current_text = self.parent_item_combo.currentText()
            if current_text:
                parent_item_id = self.get_parent_item_id_from_text(current_text)
                print(f"[DEBUG] get_bom_data: 从文本提取的parent_item_id: {parent_item_id}")

        return {
            'BomName': self.bom_name_edit.text().strip(),
            'ParentItemId': parent_item_id,
            'Rev': self.rev_edit.text().strip(),
            'EffectiveDate': effective_date,
            'ExpireDate': expire_date,
            'Remark': self.remark_edit.toPlainText().strip()
        }

    def get_materials_data(self):
        """从树形视图获取所有物料数据"""
        materials_data = []
        self.collect_materials_from_tree(self.product_tree.invisibleRootItem(), materials_data)
        return materials_data

    def collect_materials_from_tree(self, parent_item, materials_data):
        """递归收集树形视图中的物料数据"""
        for i in range(parent_item.childCount()):
            child_item = parent_item.child(i)
            
            # 从节点的UserRole数据中获取物料信息
            item_data = child_item.data(0, Qt.UserRole)
            
            if item_data and isinstance(item_data, dict) and item_data.get('ItemId', 0):
                try:
                    # 获取物料ID
                    child_item_id = item_data['ItemId']
                    
                    # 获取用量
                    qty_text = child_item.text(3)
                    qty_per = float(qty_text) if qty_text else 1.0
                    
                    # 获取损耗率
                    scrap_text = child_item.text(4)
                    scrap_factor = float(scrap_text.replace("%", "")) / 100 if scrap_text and "%" in scrap_text else 0.0
                    
                    # 获取备注
                    remark = child_item.text(5)
                    
                    materials_data.append({
                        'ChildItemId': child_item_id,
                        'QtyPer': qty_per,
                        'ScrapFactor': scrap_factor,
                        'Remark': remark
                    })
                    
                    # 递归处理子节点
                    self.collect_materials_from_tree(child_item, materials_data)
                    
                except Exception as e:
                    print(f"处理物料数据失败: {e}")
            else:
                # 如果没有物料ID，尝试从文本中解析物料编码
                item_code = child_item.text(0)  # 编码列
                
                if item_code:
                    try:
                        # 根据物料编码查找物料ID
                        items = ItemService.search_items(item_code)
                        if items:
                            child_item_id = items[0]['ItemId'] if 'ItemId' in items[0].keys() else 0
                            
                            if child_item_id:
                                # 获取用量
                                qty_text = child_item.text(3)
                                qty_per = float(qty_text) if qty_text else 1.0

                                # 获取损耗率
                                scrap_text = child_item.text(4)
                                scrap_factor = float(scrap_text.replace("%", "")) / 100 if scrap_text and "%" in scrap_text else 0.0

                                # 获取备注
                                remark = child_item.text(5)

                                materials_data.append({
                                    'ChildItemId': child_item_id,
                                    'QtyPer': qty_per,
                                    'ScrapFactor': scrap_factor,
                                    'Remark': remark
                                })

                                # 递归处理子节点
                                self.collect_materials_from_tree(child_item, materials_data)
                                
                    except Exception as e:
                        print(f"处理物料数据失败: {e}")

    def save_bom(self):
        """保存BOM数据"""
        try:
            # 验证必填字段
            bom_data = self.get_bom_data()
            print(f"调试 - BOM数据: {bom_data}")  # 调试信息
            
            # 检查bom_data是否为None
            if bom_data is None:
                QMessageBox.critical(self, "错误", "获取BOM数据失败！")
                return
            
            # 验证BOM名称
            bom_name = bom_data.get('BomName', '').strip()
            if not bom_name:
                QMessageBox.warning(self, "警告", "请输入BOM名称！")
                return

            # 验证版本号
            rev = bom_data.get('Rev', '').strip()
            if not rev:
                QMessageBox.warning(self, "警告", "请输入版本号！")
                return

            # 验证版本号格式
            if not re.match(r'^[A-Za-z0-9._-]+$', rev):
                QMessageBox.warning(self, "警告", "版本号只能包含字母、数字、点、下划线和连字符！")
                return
                
            # 验证父产品
            if not bom_data.get('ParentItemId'):
                QMessageBox.warning(self, "警告", "请选择父产品！")
                return
                
            # 验证生效日期
            effective_date = bom_data.get('EffectiveDate')
            if not effective_date:
                QMessageBox.warning(self, "警告", "请选择生效日期！")
                return
                
            # 验证失效日期
            expire_date = bom_data.get('ExpireDate')
            if expire_date:
                effective_date_obj = QDate.fromString(effective_date, "yyyy-MM-dd")
                expire_date_obj = QDate.fromString(expire_date, "yyyy-MM-dd")
                if expire_date_obj <= effective_date_obj:
                    QMessageBox.warning(self, "警告", "失效日期必须晚于生效日期！")
                    return

            # 验证物料明细
            materials_data = self.get_materials_data()
            print(f"调试 - 物料数据: {materials_data}")  # 调试信息
            
            if not materials_data:
                QMessageBox.warning(self, "警告", "请至少添加一个物料明细！")
                return
                
            # 验证物料数据
            for i, material in enumerate(materials_data):
                print(f"调试 - 验证第{i+1}个物料: {material}")  # 调试信息
                qty_per = material.get('QtyPer', 0)
                scrap_factor = material.get('ScrapFactor', 0)
                print(f"调试 - 第{i+1}个物料用量: {qty_per}, 损耗率: {scrap_factor}")  # 调试信息
                
                if qty_per <= 0:
                    QMessageBox.warning(self, "警告", f"第{i+1}个物料的用量必须大于0！")
                    return
                if scrap_factor < 0:
                    QMessageBox.warning(self, "警告", f"第{i+1}个物料的损耗率不能为负数！")
                    return

            # 保存数据
            if self.bom_id:
                # 更新现有BOM
                print(f"调试 - 更新BOM，ID: {self.bom_id}")  # 调试信息
                BomService.update_bom_header(self.bom_id, bom_data)
                
                # 智能更新BOM明细
                self.smart_update_bom_lines(self.bom_id, materials_data)
                
                QMessageBox.information(self, "成功", "BOM更新成功！")
            else:
                # 创建新BOM
                print(f"调试 - 创建新BOM")  # 调试信息
                bom_id = BomService.create_bom_header(bom_data)
                print(f"调试 - 新BOM ID: {bom_id}")  # 调试信息
                
                # 添加物料明细
                for material_data in materials_data:
                    print(f"调试 - 创建BOM明细: {material_data}")  # 调试信息
                    BomService.create_bom_line(bom_id, material_data)
                QMessageBox.information(self, "成功", "BOM创建成功！")

            self.accept()

        except Exception as e:
            print(f"调试 - 保存BOM异常: {e}")  # 调试信息
            QMessageBox.critical(self, "错误", f"保存BOM失败: {str(e)}")

    def smart_update_bom_lines(self, bom_id, new_materials_data):
        """智能更新BOM明细 - 只更新有变化的明细"""
        try:
            # 获取现有明细
            existing_lines = BomService.get_bom_lines(bom_id)
            existing_dict = {}
            for line in existing_lines:
                if hasattr(line, 'keys'):
                    line = dict(line)
                child_item_id = line.get('ChildItemId')
                if child_item_id:
                    existing_dict[child_item_id] = line
            
            # 创建新明细字典
            new_dict = {}
            for material in new_materials_data:
                child_item_id = material.get('ChildItemId')
                if child_item_id:
                    new_dict[child_item_id] = material
            
            # 找出需要删除的明细（存在于现有但不在新数据中）
            to_delete = []
            for child_item_id, existing_line in existing_dict.items():
                if child_item_id not in new_dict:
                    to_delete.append(existing_line)
            
            # 找出需要添加的明细（存在于新数据但不在现有中）
            to_add = []
            for child_item_id, new_material in new_dict.items():
                if child_item_id not in existing_dict:
                    to_add.append(new_material)
            
            # 找出需要更新的明细（都存在但数据有变化）
            to_update = []
            for child_item_id, new_material in new_dict.items():
                if child_item_id in existing_dict:
                    existing_line = existing_dict[child_item_id]
                    # 比较关键字段是否有变化
                    if (existing_line.get('QtyPer', 0) != new_material.get('QtyPer', 0) or
                        existing_line.get('ScrapFactor', 0) != new_material.get('ScrapFactor', 0) or
                        existing_line.get('Remark', '') != new_material.get('Remark', '')):
                        to_update.append({
                            'line_id': existing_line.get('LineId'),
                            'new_data': new_material
                        })
            
            print(f"调试 - 智能更新明细: 删除{len(to_delete)}个, 添加{len(to_add)}个, 更新{len(to_update)}个")
            
            # 执行删除
            for line in to_delete:
                line_id = line.get('LineId')
                if line_id:
                    BomService.delete_bom_line(line_id)
                    print(f"调试 - 删除明细: {line.get('ChildItemName', '未知')}")
            
            # 执行添加
            for material in to_add:
                BomService.create_bom_line(bom_id, material)
                print(f"调试 - 添加明细: {material.get('ChildItemName', '未知')}")
            
            # 执行更新
            for update_item in to_update:
                BomService.update_bom_line(update_item['line_id'], update_item['new_data'])
                print(f"调试 - 更新明细: {update_item['new_data'].get('ChildItemName', '未知')}")
                
        except Exception as e:
            print(f"调试 - 智能更新明细失败: {e}")
            raise e


class MaterialSelectionDialog(QDialog):
    """物料选择对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_materials = []  # 改为支持多选
        self.setWindowTitle("选择物料")
        self.resize(750, 500)
        self.setMinimumSize(700, 450)
        self.setModal(True)
        # 设置全局复选框样式
        self.setStyleSheet("""
            QCheckBox {
                spacing: 8px;
                font-size: 14px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #d9d9d9;
                border-radius: 3px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                background-color: #1890ff;
                border-color: #1890ff;
            }
            QCheckBox::indicator:hover {
                border-color: #1890ff;
            }
            QCheckBox::indicator:checked:hover {
                background-color: #40a9ff;
                border-color: #40a9ff;
            }
        """)
        self.setup_ui()
        self.load_materials()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 搜索区域
        search_frame = QFrame()
        search_layout = QHBoxLayout(search_frame)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(10)

        search_label = QLabel("搜索物料:")
        search_label.setStyleSheet("font-size: 14px; font-weight: 500;")
        
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入物料编码、名称、规格或品牌搜索")
        self.search_edit.textChanged.connect(self.filter_materials)
        
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_edit)
        search_layout.addStretch()

        layout.addWidget(search_frame)

        # 物料列表表格
        self.materials_table = QTableWidget()
        self.materials_table.setColumnCount(6)
        self.materials_table.setHorizontalHeaderLabels([
            "选择", "物料编码", "物料名称", "规格型号", "商品品牌", "物料类型"
        ])

        # 设置表格样式
        self.materials_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e8e8e8;
                background-color: white;
                alternate-background-color: #fafafa;
                selection-background-color: #e6f7ff;
                selection-color: #262626;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                font-size: 13px;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                color: #262626;
                padding: 10px 8px;
                border: none;
                border-bottom: 2px solid #e8e8e8;
                font-weight: 600;
                font-size: 13px;
            }
        """)

        # 设置表格属性
        self.materials_table.setAlternatingRowColors(True)
        self.materials_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.materials_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.materials_table.itemDoubleClicked.connect(self.on_item_double_clicked)

        # 调整列宽
        header = self.materials_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 物料编码
        header.setSectionResizeMode(2, QHeaderView.Stretch)  # 物料名称
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 规格型号
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 商品品牌
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 物料类型

        layout.addWidget(self.materials_table)

        # 按钮区域
        button_layout = QHBoxLayout()
        
        # 多选操作按钮
        select_all_btn = QPushButton("全选")
        select_all_btn.setStyleSheet("""
            QPushButton {
                background: #52c41a;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
                min-width: 60px;
            }
            QPushButton:hover {
                background: #73d13d;
            }
        """)
        select_all_btn.clicked.connect(self.select_all_materials)
        
        clear_all_btn = QPushButton("清空")
        clear_all_btn.setStyleSheet("""
            QPushButton {
                background: #faad14;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
                min-width: 60px;
            }
            QPushButton:hover {
                background: #ffc53d;
            }
        """)
        clear_all_btn.clicked.connect(self.clear_all_selections)
        
        button_layout.addWidget(select_all_btn)
        button_layout.addWidget(clear_all_btn)
        button_layout.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)

        select_btn = QPushButton("选择")
        select_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-size: 14px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        select_btn.clicked.connect(self.select_material)

        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(select_btn)

        layout.addLayout(button_layout)

    def load_materials(self):
        """加载物料列表（只显示启用的物料）"""
        try:
            materials = ItemService.get_all_items_with_status()
            # 过滤掉禁用的物料
            active_materials = [m for m in materials if m.get('IsActive', 1) == 1]
            self.populate_materials_table(active_materials)
        except Exception as e:
            print(f"加载物料失败: {e}")

    def populate_materials_table(self, materials):
        """填充物料表格"""
        self.materials_table.setRowCount(len(materials))

        for row, material in enumerate(materials):
            # 添加复选框
            checkbox = QCheckBox()
            # 设置复选框样式，确保在不同系统上都能正确显示
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 8px;
                    font-size: 14px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                    border: 2px solid #d9d9d9;
                    border-radius: 3px;
                    background-color: white;
                }
                QCheckBox::indicator:checked {
                    background-color: #1890ff;
                    border-color: #1890ff;
                }
                QCheckBox::indicator:hover {
                    border-color: #1890ff;
                }
                QCheckBox::indicator:checked:hover {
                    background-color: #40a9ff;
                    border-color: #40a9ff;
                }
            """)
            # 强制设置复选框的文本为空，避免显示默认文本
            checkbox.setText("")
            # 强制更新样式
            checkbox.update()
            checkbox.stateChanged.connect(lambda state, r=row: self.on_checkbox_changed(state, r))
            self.materials_table.setCellWidget(row, 0, checkbox)
            
            # 修复SQLite Row对象的访问方式
            item_code = material['ItemCode'] if 'ItemCode' in material.keys() else ''
            item_name = material['CnName'] if 'CnName' in material.keys() else ''
            item_spec = material['ItemSpec'] if 'ItemSpec' in material.keys() else ''
            item_brand = material['Brand'] if 'Brand' in material.keys() else ''
            item_type = material['ItemType'] if 'ItemType' in material.keys() else ''
            
            self.materials_table.setItem(row, 1, QTableWidgetItem(str(item_code)))
            self.materials_table.setItem(row, 2, QTableWidgetItem(str(item_name)))
            self.materials_table.setItem(row, 3, QTableWidgetItem(str(item_spec)))
            self.materials_table.setItem(row, 4, QTableWidgetItem(str(item_brand)))
            self.materials_table.setItem(row, 5, QTableWidgetItem(str(item_type)))

    def filter_materials(self, search_text):
        """过滤物料 - 支持物料编码、名称、规格、品牌的模糊搜索"""
        for row in range(self.materials_table.rowCount()):
            # 获取所有搜索字段
            item_code_item = self.materials_table.item(row, 1)
            item_name_item = self.materials_table.item(row, 2)
            item_spec_item = self.materials_table.item(row, 3)
            item_brand_item = self.materials_table.item(row, 4)
            
            # 检查所有字段是否为空
            if (item_code_item and item_name_item and 
                item_spec_item and item_brand_item):
                
                item_code = item_code_item.text()
                item_name = item_name_item.text()
                item_spec = item_spec_item.text()
                item_brand = item_brand_item.text()
                
                search_lower = search_text.lower()
                
                # 如果搜索文本为空，显示所有行
                if not search_text:
                    self.materials_table.setRowHidden(row, False)
                else:
                    # 检查是否在任意字段中匹配
                    if (search_lower in item_code.lower() or 
                        search_lower in item_name.lower() or 
                        search_lower in item_spec.lower() or 
                        search_lower in item_brand.lower()):
                        self.materials_table.setRowHidden(row, False)
                    else:
                        self.materials_table.setRowHidden(row, True)
            else:
                # 如果表格项为空，隐藏该行
                self.materials_table.setRowHidden(row, True)
    
    def on_checkbox_changed(self, state, row):
        """复选框状态改变"""
        # 这里可以添加选中状态变化的处理逻辑
        pass
    
    def select_all_materials(self):
        """全选所有物料"""
        for row in range(self.materials_table.rowCount()):
            if not self.materials_table.isRowHidden(row):
                checkbox = self.materials_table.cellWidget(row, 0)
                if checkbox:
                    checkbox.setChecked(True)
    
    def clear_all_selections(self):
        """清空所有选择"""
        for row in range(self.materials_table.rowCount()):
            checkbox = self.materials_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(False)

    def on_item_double_clicked(self, item, column):
        """双击选择物料"""
        self.select_material()

    def select_material(self):
        """选择物料"""
        self.selected_materials = []
        
        for row in range(self.materials_table.rowCount()):
            checkbox = self.materials_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                item_code = self.materials_table.item(row, 1).text()
                item_name = self.materials_table.item(row, 2).text()
                item_spec = self.materials_table.item(row, 3).text()
                item_brand = self.materials_table.item(row, 4).text()
                item_type = self.materials_table.item(row, 5).text()
                
                # 根据物料编码查找完整的物料信息
                try:
                    items = ItemService.search_items(item_code)
                    if items and len(items) > 0:
                        item = items[0]
                        material_data = {
                            'ItemId': item['ItemId'] if 'ItemId' in item.keys() else 0,
                            'ItemCode': item_code,
                            'CnName': item_name,
                            'ItemSpec': item_spec,
                            'Brand': item_brand,
                            'ItemType': item_type
                        }
                        self.selected_materials.append(material_data)
                except Exception as e:
                    print(f"获取物料信息失败: {e}")
        
        if self.selected_materials:
            self.accept()
        else:
            QMessageBox.warning(self, "警告", "请至少选择一个物料！")

    def get_selected_materials(self):
        """获取选中的物料列表"""
        return self.selected_materials


class BomManagementWidget(QWidget):
    """BOM管理主界面"""

    def __init__(self, parent=None):
        super().__init__(parent)
        # 初始化服务
        self.bom_service = BomService()
        self.item_service = ItemService()
        self.setup_ui()
        self.load_boms()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(16)

        # 设置主窗口尺寸策略
        self.setMinimumSize(800, 600)
        self.resize(1000, 700)

        # 标题
        title_label = QLabel("BOM 管理")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #333;
                padding: 15px;
                text-align: center;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # 搜索栏
        search_frame = QFrame()
        search_layout = QHBoxLayout(search_frame)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(8)

        # 搜索标签
        search_label = QLabel("搜索零部件:")
        search_label.setStyleSheet("""
            QLabel {
                color: #262626;
                font-size: 13px;
                font-weight: 500;
            }
        """)
        search_layout.addWidget(search_label)

        # 搜索输入框
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入物料编码、名称或规格，查找使用了该零部件的BOM")
        self.search_edit.setMinimumWidth(400)
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #d9d9d9;
                border-radius: 4px;
                font-size: 13px;
                background-color: white;
            }
            QLineEdit:focus {
                border-color: #40a9ff;
                outline: none;
            }
        """)
        self.search_edit.textChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.search_edit)

        # 重置按钮
        reset_btn = QPushButton("重置")
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 60px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        reset_btn.clicked.connect(self.on_reset_search)
        search_layout.addWidget(reset_btn)

        search_layout.addStretch()
        layout.addWidget(search_frame)

        # 按钮栏
        button_frame = QFrame()
        button_layout = QHBoxLayout(button_frame)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(8)

        # 新增BOM按钮
        self.add_bom_btn = QPushButton("新增BOM")
        self.add_bom_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 90px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        self.add_bom_btn.clicked.connect(self.add_bom)

        # 刷新按钮
        self.refresh_btn = QPushButton("刷新")
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background: white;
                color: #595959;
                border: 1px solid #d9d9d9;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 70px;
            }
            QPushButton:hover {
                background: #f5f5f5;
                border-color: #40a9ff;
                color: #40a9ff;
            }
        """)
        self.refresh_btn.clicked.connect(lambda: self.load_boms())

        # 导入BOM按钮
        self.import_bom_btn = QPushButton("导入BOM")
        self.import_bom_btn.setStyleSheet("""
            QPushButton {
                background: #722ed1;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 90px;
            }
            QPushButton:hover {
                background: #9254de;
            }
        """)
        self.import_bom_btn.clicked.connect(self.show_bom_import_dialog)

        # 导出BOM按钮
        self.export_bom_btn = QPushButton("导出BOM")
        self.export_bom_btn.setStyleSheet("""
            QPushButton {
                background: #13c2c2;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 90px;
            }
            QPushButton:hover {
                background: #36cfc9;
            }
        """)
        self.export_bom_btn.clicked.connect(self.export_bom_to_excel)

        # 查看历史按钮
        self.view_history_btn = QPushButton("查看历史")
        self.view_history_btn.setStyleSheet("""
            QPushButton {
                background: #fa8c16;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 90px;
            }
            QPushButton:hover {
                background: #ffa940;
            }
        """)
        self.view_history_btn.clicked.connect(self.show_bom_history_dialog)

        # BOM展开按钮
        self.expand_bom_btn = QPushButton("BOM展开")
        self.expand_bom_btn.setStyleSheet("""
            QPushButton {
                background: #52c41a;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #73d13d;
            }
        """)
        self.expand_bom_btn.clicked.connect(self.show_bom_expand_dialog)

        button_layout.addWidget(self.add_bom_btn)
        button_layout.addWidget(self.import_bom_btn)
        button_layout.addWidget(self.export_bom_btn)
        button_layout.addWidget(self.view_history_btn)
        button_layout.addWidget(self.refresh_btn)
        button_layout.addWidget(self.expand_bom_btn)
        button_layout.addStretch()

        layout.addWidget(button_frame)

        # BOM表格
        self.bom_table = QTableWidget()
        self.bom_table.setColumnCount(11)
        self.bom_table.setHorizontalHeaderLabels([
            "BOM ID", "BOM名称", "父产品编码", "父产品名称", "父产品规格", "版本", "生效日期", "失效日期", "备注", "状态", "操作"
        ])

        # 设置表格样式
        self.bom_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e8e8e8;
                background-color: white;
                alternate-background-color: #fafafa;
                selection-background-color: #e6f7ff;
                selection-color: #262626;
                border: 1px solid #e8e8e8;
                border-radius: 4px;
                font-size: 13px;
            }
            QHeaderView::section {
                background-color: #fafafa;
                color: #262626;
                padding: 10px 6px;
                border: none;
                border-bottom: 1px solid #e8e8e8;
                font-weight: 500;
                font-size: 12px;
            }
        """)

        # 设置表格属性
        self.bom_table.setAlternatingRowColors(True)
        self.bom_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.bom_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # 连接状态列双击事件
        self.bom_table.cellDoubleClicked.connect(self.on_status_cell_double_clicked)

        # 调整列宽
        header = self.bom_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # BOM ID
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # BOM名称
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 父产品编码
        header.setSectionResizeMode(3, QHeaderView.Stretch)  # 父产品名称
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 父产品规格
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 版本
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 生效日期
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 失效日期
        header.setSectionResizeMode(8, QHeaderView.Stretch)  # 备注
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 状态
        header.setSectionResizeMode(10, QHeaderView.Fixed)  # 操作
        header.setDefaultSectionSize(120)  # 设置默认列宽
        self.bom_table.setColumnWidth(10, 150)  # 设置操作列宽度

        layout.addWidget(self.bom_table)

    def load_boms(self, search_filter: str = None):
        """加载BOM列表"""
        try:
            boms = BomService.get_bom_headers(search_filter)
            self.populate_bom_table(boms)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载BOM列表失败: {str(e)}")

    def on_search_changed(self):
        """当搜索条件变化时，重新加载BOM列表"""
        search_text = self.search_edit.text().strip()
        self.load_boms(search_text if search_text else None)

    def on_reset_search(self):
        """重置搜索条件"""
        self.search_edit.clear()
        self.load_boms()

    def show_bom_history_dialog(self):
        """显示BOM历史对话框"""
        dialog = BomHistoryDialog(self)
        dialog.exec()

    def on_status_cell_double_clicked(self, row, column):
        """状态列双击事件 - 显示BOM状态详情"""
        if column == 9:  # 状态列
            bom_id = self.bom_table.item(row, 0).text()
            if bom_id:
                self.show_bom_status_details(int(bom_id))

    def show_bom_status_details(self, bom_id):
        """显示BOM状态详情对话框"""
        try:
            # 获取BOM状态详情
            status_details = BomService.get_bom_status_details(bom_id)
            
            # 创建详情对话框
            dialog = QDialog(self)
            dialog.setWindowTitle("BOM状态详情")
            dialog.setModal(True)
            dialog.resize(500, 400)
            
            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(20, 20, 20, 20)
            layout.setSpacing(15)
            
            # 标题
            title_label = QLabel(f"BOM状态详情 (ID: {bom_id})")
            title_label.setStyleSheet("""
                QLabel {
                    font-size: 16px;
                    font-weight: bold;
                    color: #333;
                    padding: 10px;
                }
            """)
            layout.addWidget(title_label)
            
            # 状态信息
            status_frame = QFrame()
            status_frame.setStyleSheet("""
                QFrame {
                    background: #f5f5f5;
                    border: 1px solid #d9d9d9;
                    border-radius: 6px;
                    padding: 15px;
                }
            """)
            status_layout = QVBoxLayout(status_frame)
            
            # 整体状态
            overall_status = status_details['status']
            status_color = "#52c41a" if overall_status == "有效" else "#ff4d4f"
            status_text = QLabel(f"整体状态: {overall_status}")
            status_text.setStyleSheet(f"""
                QLabel {{
                    font-size: 14px;
                    font-weight: bold;
                    color: {status_color};
                    padding: 5px;
                }}
            """)
            status_layout.addWidget(status_text)
            
            # 父产品状态
            parent_status = status_details['parent_status']
            parent_text = QLabel(f"父产品状态: {parent_status}")
            parent_text.setStyleSheet("""
                QLabel {
                    font-size: 13px;
                    color: #666;
                    padding: 3px;
                }
            """)
            status_layout.addWidget(parent_text)
            
            layout.addWidget(status_frame)
            
            # 失效原因（如果有）
            disabled_components = status_details['disabled_components']
            if disabled_components:
                reason_frame = QFrame()
                reason_frame.setStyleSheet("""
                    QFrame {
                        background: #fff2f0;
                        border: 1px solid #ffccc7;
                        border-radius: 6px;
                        padding: 15px;
                    }
                """)
                reason_layout = QVBoxLayout(reason_frame)
                
                reason_title = QLabel("失效原因:")
                reason_title.setStyleSheet("""
                    QLabel {
                        font-size: 13px;
                        font-weight: bold;
                        color: #ff4d4f;
                        padding: 5px;
                    }
                """)
                reason_layout.addWidget(reason_title)
                
                for component in disabled_components:
                    component_text = QLabel(f"• {component['type']}: {component['name']} ({component['code']})")
                    component_text.setStyleSheet("""
                        QLabel {
                            font-size: 12px;
                            color: #666;
                            padding: 2px 0px 2px 15px;
                        }
                    """)
                    reason_layout.addWidget(component_text)
                
                layout.addWidget(reason_frame)
            else:
                # 有效状态说明
                valid_frame = QFrame()
                valid_frame.setStyleSheet("""
                    QFrame {
                        background: #f6ffed;
                        border: 1px solid #b7eb8f;
                        border-radius: 6px;
                        padding: 15px;
                    }
                """)
                valid_layout = QVBoxLayout(valid_frame)
                
                valid_text = QLabel("✓ 父产品和所有零部件都处于启用状态")
                valid_text.setStyleSheet("""
                    QLabel {
                        font-size: 13px;
                        color: #52c41a;
                        padding: 5px;
                    }
                """)
                valid_layout.addWidget(valid_text)
                
                layout.addWidget(valid_frame)
            
            # 按钮
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            
            close_btn = QPushButton("关闭")
            close_btn.setStyleSheet("""
                QPushButton {
                    background: #1890ff;
                    color: white;
                    border: none;
                    padding: 8px 20px;
                    border-radius: 4px;
                    font-size: 14px;
                    font-weight: 500;
                    min-width: 100px;
                }
                QPushButton:hover {
                    background: #40a9ff;
                }
            """)
            close_btn.clicked.connect(dialog.accept)
            button_layout.addWidget(close_btn)
            
            layout.addLayout(button_layout)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"显示BOM状态详情失败: {str(e)}")

    def show_bom_expand_dialog(self):
        """显示BOM展开对话框"""
        try:
            # 获取有BOM结构的成品物料（只显示启用的）
            items_with_bom = []
            all_items = ItemService.get_all_items_with_status()
            
            for item in all_items:
                # 修复SQLite Row对象的访问方式
                item_type = item['ItemType'] if 'ItemType' in item.keys() else ''
                is_active = item.get('IsActive', 1)
                
                # 只处理启用的成品和半成品
                if item_type in ['FG', 'SFG'] and is_active == 1:
                    # 检查这个物料是否有BOM结构
                    try:
                        bom = BomService.get_bom_by_parent_item(item['ItemId'])
                        if bom:
                            items_with_bom.append(item)
                    except:
                        continue
            
            if not items_with_bom:
                QMessageBox.information(self, "提示", "没有找到有BOM结构的成品或半成品物料")
                return
            
            # 创建BOM展开对话框
            dialog = BomExpandDialog(self, items_with_bom)
            if dialog.exec() == QDialog.Accepted:
                pass  # 可以在这里添加后续处理逻辑
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"显示BOM展开对话框失败: {str(e)}")

    def show_bom_import_dialog(self):
        """显示BOM导入对话框"""
        try:
            dialog = BomImportDialog(self)
            if dialog.exec() == QDialog.Accepted:
                # 导入成功后刷新BOM列表
                self.load_boms()
                QMessageBox.information(self, "成功", "BOM导入完成！")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"显示BOM导入对话框失败: {str(e)}")

    def export_bom_to_excel(self):
        """导出所有BOM到Excel文件"""
        try:
            # 获取所有BOM数据
            all_boms = self.bom_service.get_bom_headers()
            if not all_boms:
                QMessageBox.warning(self, "警告", "没有找到任何BOM数据！")
                return
            
            # 选择保存位置
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出所有BOM",
                "所有BOM清单.xlsx",
                "Excel文件 (*.xlsx);;所有文件 (*)"
            )
            
            if save_path:
                # 生成Excel文件
                self.generate_all_boms_excel(all_boms, save_path)
                QMessageBox.information(self, "成功", f"所有BOM已导出到：\n{save_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出BOM失败: {str(e)}")
    
    def generate_all_boms_excel(self, all_boms, file_path):
        """生成包含所有BOM的Excel文件（矩阵格式）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM矩阵"
            
            # 设置样式
            normal_font = Font(name='宋体', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 收集所有零部件和成品信息
            all_components = {}  # {component_id: component_info}
            all_products = []    # [product_info]
            bom_matrix = {}      # {(component_id, product_id): quantity}
            
            for bom_data in all_boms:
                # 将Row对象转换为字典
                if hasattr(bom_data, 'keys'):
                    bom_data = dict(bom_data)
                
                # 获取父产品信息
                parent_item_id = bom_data.get('ParentItemId')
                parent_item = None
                if parent_item_id:
                    parent_item = self.item_service.get_item_by_id(parent_item_id)
                
                if not parent_item:
                    continue
                
                # 添加到成品列表
                product_info = {
                    'ItemId': parent_item_id,
                    'ItemCode': parent_item.get('ItemCode', ''),
                    'CnName': parent_item.get('CnName', ''),
                    'ItemSpec': parent_item.get('ItemSpec', ''),
                    'Brand': parent_item.get('Brand', ''),
                    'BomId': bom_data.get('BomId')
                }
                all_products.append(product_info)
                
                # 获取BOM的物料清单
                bom_lines = self.bom_service.get_bom_lines(bom_data.get('BomId'))
                if not bom_lines:
                    continue
                
                # 处理零部件信息
                for bom_line in bom_lines:
                    # 将Row对象转换为字典
                    if hasattr(bom_line, 'keys'):
                        bom_line = dict(bom_line)
                    
                    child_item_id = bom_line.get('ChildItemId')
                    if child_item_id:
                        # 获取零部件信息
                        child_item = self.item_service.get_item_by_id(child_item_id)
                        if child_item:
                            # 添加到零部件字典
                            all_components[child_item_id] = {
                                'ItemId': child_item_id,
                                'ItemCode': child_item.get('ItemCode', ''),
                                'CnName': child_item.get('CnName', ''),
                                'ItemSpec': child_item.get('ItemSpec', '')
                            }
                            
                            # 记录用量关系
                            quantity = bom_line.get('QtyPer', 0)
                            bom_matrix[(child_item_id, parent_item_id)] = quantity
            
            # 对成品进行排序：按名称、规格排序
            all_products.sort(key=lambda x: (x.get('CnName', ''), x.get('ItemSpec', '')))
            
            # 对零部件进行排序：按名称、规格排序
            sorted_components = sorted(all_components.items(), key=lambda x: (x[1].get('CnName', ''), x[1].get('ItemSpec', '')))
            
            # 生成Excel表格
            current_row = 1
            
            # 第一行：成品编码（A-C列保持空白）
            for i, product in enumerate(all_products):
                col_letter = openpyxl.utils.get_column_letter(4 + i)  # 从D列开始
                ws[f'{col_letter}1'] = product['ItemCode']
                ws[f'{col_letter}1'].font = normal_font
                ws[f'{col_letter}1'].alignment = center_alignment
                ws[f'{col_letter}1'].border = border
            
            # 第二行：成品名称
            for i, product in enumerate(all_products):
                col_letter = openpyxl.utils.get_column_letter(4 + i)
                ws[f'{col_letter}2'] = product['CnName']
                ws[f'{col_letter}2'].font = normal_font
                ws[f'{col_letter}2'].alignment = center_alignment
                ws[f'{col_letter}2'].border = border
            
            # 第三行：成品规格
            for i, product in enumerate(all_products):
                col_letter = openpyxl.utils.get_column_letter(4 + i)
                ws[f'{col_letter}3'] = product['ItemSpec']
                ws[f'{col_letter}3'].font = normal_font
                ws[f'{col_letter}3'].alignment = center_alignment
                ws[f'{col_letter}3'].border = border
            
            # 第四行：成品品牌
            for i, product in enumerate(all_products):
                col_letter = openpyxl.utils.get_column_letter(4 + i)
                ws[f'{col_letter}4'] = product['Brand']
                ws[f'{col_letter}4'].font = normal_font
                ws[f'{col_letter}4'].alignment = center_alignment
                ws[f'{col_letter}4'].border = border
            
            # 从第5行开始：零部件信息
            current_row = 5
            
            for component_id, component in sorted_components:
                # 零部件信息
                ws[f'A{current_row}'] = component['ItemCode']
                ws[f'A{current_row}'].font = normal_font
                ws[f'A{current_row}'].alignment = center_alignment
                ws[f'A{current_row}'].border = border
                
                ws[f'B{current_row}'] = component['CnName']
                ws[f'B{current_row}'].font = normal_font
                ws[f'B{current_row}'].alignment = center_alignment
                ws[f'B{current_row}'].border = border
                
                ws[f'C{current_row}'] = component['ItemSpec']
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].alignment = center_alignment
                ws[f'C{current_row}'].border = border
                
                # 用量信息
                for i, product in enumerate(all_products):
                    col_letter = openpyxl.utils.get_column_letter(4 + i)
                    quantity = bom_matrix.get((component_id, product['ItemId']), 0)
                    ws[f'{col_letter}{current_row}'] = quantity
                    ws[f'{col_letter}{current_row}'].font = normal_font
                    ws[f'{col_letter}{current_row}'].alignment = center_alignment
                    ws[f'{col_letter}{current_row}'].border = border
                
                current_row += 1
            
            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 零部件编码
            ws.column_dimensions['B'].width = 20  # 零部件名称
            ws.column_dimensions['C'].width = 15  # 零部件规格
            
            # 设置成品列的宽度
            for i in range(len(all_products)):
                col_letter = openpyxl.utils.get_column_letter(4 + i)
                ws.column_dimensions[col_letter].width = 12
            
            # 保存文件
            wb.save(file_path)
            
        except ImportError:
            QMessageBox.warning(self, "警告", "需要安装openpyxl库才能导出Excel文件！\n请运行: pip install openpyxl")
        except Exception as e:
            raise Exception(f"生成Excel文件失败: {str(e)}")
    
    def generate_bom_excel(self, bom_data, bom_lines, file_path):
        """生成BOM的Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM清单"
            
            # 设置样式
            header_font = Font(bold=True, size=12)
            normal_font = Font(size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 获取父产品信息
            parent_item_id = bom_data.get('ParentItemId')
            parent_item = None
            if parent_item_id:
                parent_item = self.item_service.get_item_by_id(parent_item_id)
            
            # 第四列：父产品信息
            if parent_item:
                # 第四列第一行：成品编码
                ws['D1'] = parent_item.get('ItemCode', '')
                ws['D1'].font = header_font
                ws['D1'].alignment = center_alignment
                ws['D1'].border = border
                
                # 第四列第二行：成品名称
                ws['D2'] = parent_item.get('CnName', '')
                ws['D2'].font = header_font
                ws['D2'].alignment = center_alignment
                ws['D2'].border = border
                
                # 第四列第三行：成品规格
                ws['D3'] = parent_item.get('ItemSpec', '')
                ws['D3'].font = header_font
                ws['D3'].alignment = center_alignment
                ws['D3'].border = border
                
                # 第四列第四行：成品品牌
                ws['D4'] = parent_item.get('Brand', '')
                ws['D4'].font = header_font
                ws['D4'].alignment = center_alignment
                ws['D4'].border = border
            
            # 第五行开始：零部件信息
            current_row = 5
            
            for bom_line in bom_lines:
                # 获取零部件信息
                component_item_id = bom_line.get('ComponentItemId')
                component_item = None
                if component_item_id:
                    component_item = self.item_service.get_item_by_id(component_item_id)
                
                if component_item:
                    # 第五行第一列：零部件编码
                    ws[f'A{current_row}'] = component_item.get('ItemCode', '')
                    ws[f'A{current_row}'].font = normal_font
                    ws[f'A{current_row}'].alignment = center_alignment
                    ws[f'A{current_row}'].border = border
                    
                    # 第五行第二列：零部件名称
                    ws[f'B{current_row}'] = component_item.get('CnName', '')
                    ws[f'B{current_row}'].font = normal_font
                    ws[f'B{current_row}'].alignment = center_alignment
                    ws[f'B{current_row}'].border = border
                    
                    # 第五行第三列：零部件规格
                    ws[f'C{current_row}'] = component_item.get('ItemSpec', '')
                    ws[f'C{current_row}'].font = normal_font
                    ws[f'C{current_row}'].alignment = center_alignment
                    ws[f'C{current_row}'].border = border
                    
                    # 第五行第四列开始：数量关系
                    quantity = bom_line.get('Quantity', 0)
                    ws[f'D{current_row}'] = quantity
                    ws[f'D{current_row}'].font = normal_font
                    ws[f'D{current_row}'].alignment = center_alignment
                    ws[f'D{current_row}'].border = border
                
                current_row += 1
            
            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 零部件编码
            ws.column_dimensions['B'].width = 20  # 零部件名称
            ws.column_dimensions['C'].width = 15  # 零部件规格
            ws.column_dimensions['D'].width = 12  # 数量
            
            # 保存文件
            wb.save(file_path)
            
        except ImportError:
            QMessageBox.warning(self, "警告", "需要安装openpyxl库才能导出Excel文件！\n请运行: pip install openpyxl")
        except Exception as e:
            raise Exception(f"生成Excel文件失败: {str(e)}")

    def populate_bom_table(self, boms):
        """填充BOM表格"""
        # 对BOM列表进行排序：有效的排在前面，失效的排在后面
        def sort_key(bom):
            bom_id = bom['BomId'] if 'BomId' in bom.keys() else 0
            bom_status = BomService.get_bom_status(bom_id)
            # 有效=0（排在前面），失效=1（排在后面），未知=2（排在最后）
            if bom_status == "有效":
                return 0
            elif bom_status == "失效":
                return 1
            else:
                return 2
        
        # 排序BOM列表
        sorted_boms = sorted(boms, key=sort_key)
        
        self.bom_table.setRowCount(len(sorted_boms))

        for row, bom in enumerate(sorted_boms):
            # 修复SQLite Row对象的访问方式
            bom_id = bom['BomId'] if 'BomId' in bom.keys() else ''
            bom_name = bom['BomName'] if 'BomName' in bom.keys() else ''
            parent_item_code = bom['ParentItemCode'] if 'ParentItemCode' in bom.keys() else ''
            parent_item_name = bom['ParentItemName'] if 'ParentItemName' in bom.keys() else ''
            parent_item_spec = bom['ParentItemSpec'] if 'ParentItemSpec' in bom.keys() else ''
            rev = bom['Rev'] if 'Rev' in bom.keys() else ''
            effective_date = bom['EffectiveDate'] if 'EffectiveDate' in bom.keys() else ''
            expire_date = bom['ExpireDate'] if 'ExpireDate' in bom.keys() else ''
            remark = bom['Remark'] if 'Remark' in bom.keys() else ''
            
            # BOM ID
            self.bom_table.setItem(row, 0, QTableWidgetItem(str(bom_id)))
            # BOM名称
            self.bom_table.setItem(row, 1, QTableWidgetItem(str(bom_name)))
            # 父产品编码
            self.bom_table.setItem(row, 2, QTableWidgetItem(str(parent_item_code)))
            # 父产品名称
            self.bom_table.setItem(row, 3, QTableWidgetItem(str(parent_item_name)))
            # 父产品规格
            self.bom_table.setItem(row, 4, QTableWidgetItem(str(parent_item_spec)))
            # 版本
            self.bom_table.setItem(row, 5, QTableWidgetItem(str(rev)))
            # 生效日期
            self.bom_table.setItem(row, 6, QTableWidgetItem(str(effective_date)))
            # 失效日期
            expire_date_display = str(expire_date) if expire_date else "未设置"
            self.bom_table.setItem(row, 7, QTableWidgetItem(expire_date_display))
            # 备注
            self.bom_table.setItem(row, 8, QTableWidgetItem(str(remark) if remark else ""))
            # 状态 - 使用新的BOM状态检测逻辑
            bom_status = BomService.get_bom_status(bom_id)
            status_item = QTableWidgetItem(bom_status)
            
            # 设置状态颜色
            if bom_status == "失效":
                status_item.setForeground(QColor("#ff4d4f"))  # 红色
                status_item.setBackground(QColor("#fff2f0"))  # 浅红色背景
            elif bom_status == "有效":
                status_item.setForeground(QColor("#52c41a"))  # 绿色
                status_item.setBackground(QColor("#f6ffed"))  # 浅绿色背景
            else:
                status_item.setForeground(QColor("#8c8c8c"))  # 灰色
                status_item.setBackground(QColor("#fafafa"))  # 浅灰色背景
            
            self.bom_table.setItem(row, 9, status_item)

            # 操作按钮
            view_btn = QPushButton("查看")
            view_btn.setStyleSheet("""
                QPushButton {
                    background: #1890ff;
                    color: white;
                    border: none;
                    padding: 4px 8px;
                    border-radius: 3px;
                    font-size: 11px;
                }
                QPushButton:hover {
                    background: #40a9ff;
                }
            """)
            view_btn.clicked.connect(lambda checked, r=row: self.view_bom(r))

            edit_btn = QPushButton("编辑")
            edit_btn.setStyleSheet("""
                QPushButton {
                    background: #52c41a;
                    color: white;
                    border: none;
                    padding: 4px 8px;
                    border-radius: 3px;
                    font-size: 11px;
                }
                QPushButton:hover {
                    background: #73d13d;
                }
            """)
            edit_btn.clicked.connect(lambda checked, r=row: self.edit_bom(r))

            delete_btn = QPushButton("删除")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background: #ff4d4f;
                    color: white;
                    border: none;
                    padding: 4px 8px;
                    border-radius: 3px;
                    font-size: 11px;
                }
                QPushButton:hover {
                    background: #ff7875;
                }
            """)
            delete_btn.clicked.connect(lambda checked, r=row: self.delete_bom(r))

            btn_layout = QHBoxLayout()
            btn_layout.addWidget(view_btn)
            btn_layout.addWidget(edit_btn)
            btn_layout.addWidget(delete_btn)
            btn_layout.setContentsMargins(4, 2, 4, 2)

            btn_widget = QWidget()
            btn_widget.setLayout(btn_layout)
            self.bom_table.setCellWidget(row, 10, btn_widget)

    def add_bom(self):
        """新增BOM"""
        # 使用新的集成编辑器
        dialog = BomEditorDialog(self)
        if dialog.exec() == QDialog.Accepted:
            self.load_boms()

    def view_bom(self, row):
        """查看BOM详情"""
        try:
            bom_id_item = self.bom_table.item(row, 0)
            if not bom_id_item:
                QMessageBox.warning(self, "警告", "无法获取BOM ID")
                return
                
            bom_id = int(bom_id_item.text())
            bom = BomService.get_bom_by_id(bom_id)
            bom_lines = BomService.get_bom_lines(bom_id)

            if not bom:
                QMessageBox.warning(self, "警告", "未找到BOM信息")
                return

            # 创建BOM查看对话框
            dialog = BomViewDialog(self, bom, bom_lines)
            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"查看BOM详情失败: {str(e)}")

    def edit_bom(self, row):
        """编辑BOM"""
        try:
            bom_id_item = self.bom_table.item(row, 0)
            if not bom_id_item:
                QMessageBox.warning(self, "警告", "无法获取BOM ID")
                return
                
            bom_id = int(bom_id_item.text())
            bom = BomService.get_bom_by_id(bom_id)

            if bom:
                # 使用新的集成编辑器
                dialog = BomEditorDialog(self, bom)
                if dialog.exec() == QDialog.Accepted:
                    self.load_boms()
            else:
                QMessageBox.warning(self, "警告", "未找到BOM信息")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"编辑BOM失败: {str(e)}")

    def delete_bom(self, row):
        """删除BOM"""
        try:
            bom_id_item = self.bom_table.item(row, 0)
            bom_name_item = self.bom_table.item(row, 1)
            
            if not bom_id_item or not bom_name_item:
                QMessageBox.warning(self, "警告", "无法获取BOM信息")
                return
                
            bom_id = int(bom_id_item.text())
            bom_name = bom_name_item.text()

            reply = QMessageBox.question(
                self, "确认删除",
                f"确定要删除BOM '{bom_name}' 吗？\n删除后无法恢复！",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                # 通过服务层删除BOM
                BomService.delete_bom_header(bom_id)
                QMessageBox.information(self, "成功", "BOM删除成功")
                self.load_boms()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"删除BOM失败: {str(e)}")


class BomExpandDialog(QDialog):
    """BOM展开对话框"""

    def __init__(self, parent=None, fg_items=None):
        super().__init__(parent)
        self.fg_items = fg_items or []
        self.setWindowTitle("BOM展开")
        self.resize(800, 650)
        self.setMinimumSize(700, 550)
        self.setMaximumSize(1000, 800)
        self.setModal(True)
        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 标题
        title_label = QLabel("BOM展开 - 计算生产所需物料")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #333;
                padding: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # 选择区域 - 紧凑布局
        select_group = QGroupBox("选择产品和数量")
        select_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-size: 13px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
                color: #262626;
            }
        """)

        select_layout = QFormLayout(select_group)
        select_layout.setSpacing(10)

        # 产品选择
        self.product_combo = QComboBox()
        self.product_combo.setPlaceholderText("请选择要生产的产品")
        for item in self.fg_items:
            # 修复SQLite Row对象的访问方式
            item_code = item['ItemCode'] if 'ItemCode' in item.keys() else ''
            item_name = item['CnName'] if 'CnName' in item.keys() else ''
            item_spec = item['ItemSpec'] if 'ItemSpec' in item.keys() else ''
            item_brand = item['Brand'] if 'Brand' in item.keys() else ''
            item_id = item['ItemId'] if 'ItemId' in item.keys() else 0
            
            # 格式：商品品牌-物资名称-物资规格
            display_text = f"{item_brand} - {item_name} - {item_spec}"
            self.product_combo.addItem(display_text, item_id)
        select_layout.addRow("产品 *:", self.product_combo)

        # 生产数量
        self.qty_spin = QDoubleSpinBox()
        self.qty_spin.setRange(1, 999999)
        self.qty_spin.setDecimals(0)
        self.qty_spin.setValue(100)
        self.qty_spin.setSuffix(" 件")
        select_layout.addRow("生产数量 *:", self.qty_spin)

        # 展开按钮
        expand_btn = QPushButton("展开BOM")
        expand_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        expand_btn.clicked.connect(self.expand_bom)
        select_layout.addRow("", expand_btn)

        layout.addWidget(select_group)

        # 结果区域 - 占用更多空间
        result_group = QGroupBox("展开结果")
        result_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-size: 13px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
                color: #262626;
            }
        """)

        result_layout = QVBoxLayout(result_group)
        result_layout.setContentsMargins(10, 10, 10, 10)

        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(5)
        self.result_table.setHorizontalHeaderLabels([
            "物料编码", "物料名称", "物料类型", "需求数量", "单位"
        ])

        # 设置表格样式
        self.result_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e8e8e8;
                background-color: white;
                alternate-background-color: #fafafa;
                selection-background-color: #e6f7ff;
                selection-color: #262626;
                border: 1px solid #e8e8e8;
                border-radius: 4px;
                font-size: 13px;
            }
            QHeaderView::section {
                background-color: #fafafa;
                color: #262626;
                padding: 8px 6px;
                border: none;
                border-bottom: 1px solid #e8e8e8;
                font-weight: 500;
                font-size: 12px;
            }
        """)

        # 设置表格属性
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.result_table.setMinimumHeight(300)  # 设置最小高度
        self.result_table.verticalHeader().setDefaultSectionSize(30)  # 设置行高

        # 调整列宽
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 物料编码
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # 物料名称
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 物料类型
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 需求数量
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 单位

        result_layout.addWidget(self.result_table)
        layout.addWidget(result_group)

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)

        export_btn = QPushButton("导出结果")
        export_btn.setStyleSheet("""
                QPushButton {
                    background: #52c41a;
                    color: white;
                    border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 100px;
                }
                QPushButton:hover {
                    background: #73d13d;
                }
            """)
        export_btn.clicked.connect(self.export_results)

        button_layout.addWidget(export_btn)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)

        # 设置样式
        self.setStyleSheet("""
            QDialog {
                background: white;
            }
            QLineEdit, QComboBox, QDoubleSpinBox {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 4px;
                font-size: 14px;
                min-width: 200px;
                min-height: 15px;
            }
            QLineEdit:focus, QComboBox:focus, QDoubleSpinBox:focus {
                border-color: #1890ff;
                }
            """)

    def expand_bom(self):
        """展开BOM"""
        try:
            product_id = self.product_combo.currentData()
            if not product_id:
                QMessageBox.warning(self, "警告", "请选择要生产的产品！")
                return

            qty = self.qty_spin.value()
            if qty <= 0:
                QMessageBox.warning(self, "警告", "生产数量必须大于0！")
                return

            # 使用BOM服务展开BOM
            expanded_items = BomService.expand_bom(product_id, qty)
            
            if not expanded_items:
                QMessageBox.information(self, "提示", "该产品没有BOM结构或BOM结构为空")
                return

            # 显示结果
            self.display_expanded_results(expanded_items)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"展开BOM失败: {str(e)}")

    def display_expanded_results(self, expanded_items):
        """显示展开结果"""
        try:
            # 按物料类型分组并汇总数量
            item_summary = {}
            for item in expanded_items:
                # 修复SQLite Row对象的访问方式
                item_id = item['ItemId'] if 'ItemId' in item.keys() else 0
                item_code = item['ItemCode'] if 'ItemCode' in item.keys() else ''
                item_name = item['ItemName'] if 'ItemName' in item.keys() else ''
                item_type = item['ItemType'] if 'ItemType' in item.keys() else ''
                unit = item['Unit'] if 'Unit' in item.keys() else '个'
                actual_qty = item['ActualQty'] if 'ActualQty' in item.keys() else 0
                
                item_key = (item_id, item_code, item_name, item_type, unit)
                if item_key in item_summary:
                    item_summary[item_key] += actual_qty
                else:
                    item_summary[item_key] = actual_qty

            # 填充表格
            self.result_table.setRowCount(len(item_summary))
            
            for row, ((item_id, item_code, item_name, item_type, unit), total_qty) in enumerate(item_summary.items()):
                self.result_table.setItem(row, 0, QTableWidgetItem(str(item_code)))
                self.result_table.setItem(row, 1, QTableWidgetItem(str(item_name)))
                self.result_table.setItem(row, 2, QTableWidgetItem(str(item_type)))
                self.result_table.setItem(row, 3, QTableWidgetItem(f"{total_qty:.3f}"))
                self.result_table.setItem(row, 4, QTableWidgetItem(str(unit)))

        except Exception as e:
            QMessageBox.critical(self, "错误", f"显示展开结果失败: {str(e)}")

    def export_results(self):
        """导出结果"""
        try:
            if self.result_table.rowCount() == 0:
                QMessageBox.information(self, "提示", "没有数据可以导出")
                return

            # 这里可以添加导出到Excel或CSV的功能
            QMessageBox.information(self, "提示", "导出功能开发中...")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")


class MaterialEditDialog(QDialog):
    """物料编辑对话框"""

    def __init__(self, parent=None, current_qty="1.0", current_scrap="0.0%", current_remark=""):
        super().__init__(parent)
        self.current_qty = current_qty
        self.current_scrap = current_scrap.replace("%", "") if "%" in current_scrap else current_scrap
        self.current_remark = current_remark
        
        self.setWindowTitle("编辑物料信息")
        self.resize(400, 320)
        self.setMinimumSize(380, 300)
        self.setModal(True)
        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # 标题
        title_label = QLabel("编辑物料信息")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #333;
                padding: 8px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # 表单区域
        form_group = QGroupBox("物料参数")
        form_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-size: 13px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
                color: #262626;
            }
        """)

        form_layout = QFormLayout(form_group)
        form_layout.setSpacing(15)

        # 用量输入
        self.qty_spin = QDoubleSpinBox()
        self.qty_spin.setRange(0.001, 999999)
        self.qty_spin.setDecimals(3)
        self.qty_spin.setValue(float(self.current_qty) if self.current_qty else 1.0)
        self.qty_spin.setSuffix(" 件")
        form_layout.addRow("用量 *:", self.qty_spin)

        # 损耗率输入
        self.scrap_spin = QDoubleSpinBox()
        self.scrap_spin.setRange(0, 100)
        self.scrap_spin.setDecimals(1)
        self.scrap_spin.setValue(float(self.current_scrap) if self.current_scrap else 0.0)
        self.scrap_spin.setSuffix(" %")
        form_layout.addRow("损耗率:", self.scrap_spin)

        # 备注输入
        self.remark_edit = QTextEdit()
        self.remark_edit.setMaximumHeight(60)
        self.remark_edit.setPlainText(str(self.current_remark))
        self.remark_edit.setPlaceholderText("请输入备注信息")
        form_layout.addRow("备注:", self.remark_edit)

        layout.addWidget(form_group)

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 10, 0, 0)
        button_layout.setSpacing(15)
        button_layout.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: white;
                color: #666;
                border: 1px solid #ccc;
                padding: 8px 20px;
                border-radius: 4px;
                font-size: 14px;
                font-weight: 500;
                min-width: 80px;
            }
            QPushButton:hover {
                border-color: #1890ff;
                color: #1890ff;
            }
        """)
        cancel_btn.clicked.connect(self.reject)

        save_btn = QPushButton("保存")
        save_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-size: 14px;
                font-weight: 500;
                min-width: 80px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        save_btn.clicked.connect(self.validate_and_accept)

        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(save_btn)

        layout.addLayout(button_layout)

        # 设置样式
        self.setStyleSheet("""
            QDialog {
                background: white;
            }
            QLineEdit, QComboBox, QTextEdit, QDoubleSpinBox {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 4px;
                font-size: 14px;
                min-width: 200px;
                min-height: 15px;
            }
            QLineEdit:focus, QComboBox:focus, QTextEdit:focus, QDoubleSpinBox:focus {
                border-color: #1890ff;
            }
        """)

    def validate_and_accept(self):
        """验证数据并接受"""
        try:
            # 验证用量
            qty = self.qty_spin.value()
            if qty <= 0:
                QMessageBox.warning(self, "警告", "用量必须大于0！")
                return

            # 验证损耗率
            scrap = self.scrap_spin.value()
            if scrap < 0 or scrap > 100:
                QMessageBox.warning(self, "警告", "损耗率必须在0-100%之间！")
                return

            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"验证数据失败: {str(e)}")

    def get_edited_data(self):
        """获取编辑后的数据"""
        return (
            self.qty_spin.value(),
            self.scrap_spin.value(),
            self.remark_edit.toPlainText().strip()
        )


class BomViewDialog(QDialog):
    """BOM查看对话框"""

    def __init__(self, parent=None, bom_data=None, bom_lines=None):
        super().__init__(parent)
        self.bom_data = bom_data or {}
        self.bom_lines = bom_lines or []
        self.setWindowTitle("BOM 详情")
        self.resize(800, 600)
        self.setMinimumSize(700, 500)
        self.setModal(True)
        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        # 设置对话框背景色
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 10, 20, 20)  # 减少顶部边距
        layout.setSpacing(15)  # 减少间距

        # 标题
        title_label = QLabel("BOM 详情")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #333;
                padding: 8px;
                background-color: #f8f9fa;
                border-radius: 4px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # BOM基本信息
        info_group = QGroupBox("BOM基本信息")
        info_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-size: 13px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
                color: #262626;
            }
        """)

        info_layout = QFormLayout(info_group)
        info_layout.setSpacing(15)

        # BOM名称
        bom_name_label = QLabel(str(self.bom_data['BomName'] if 'BomName' in self.bom_data.keys() else ''))
        bom_name_label.setStyleSheet("font-size: 14px; color: #262626;")
        info_layout.addRow("BOM名称:", bom_name_label)

        # 版本
        rev_label = QLabel(str(self.bom_data['Rev'] if 'Rev' in self.bom_data.keys() else ''))
        rev_label.setStyleSheet("font-size: 14px; color: #262626;")
        info_layout.addRow("版本:", rev_label)

        # 生效日期
        effective_date = self.bom_data['EffectiveDate'] if 'EffectiveDate' in self.bom_data.keys() else ''
        effective_date_label = QLabel(str(effective_date))
        effective_date_label.setStyleSheet("font-size: 14px; color: #262626;")
        info_layout.addRow("生效日期:", effective_date_label)

        # 失效日期
        expire_date = self.bom_data['ExpireDate'] if 'ExpireDate' in self.bom_data.keys() else ''
        expire_date_label = QLabel(str(expire_date) if expire_date else "无")
        expire_date_label.setStyleSheet("font-size: 14px; color: #262626;")
        info_layout.addRow("失效日期:", expire_date_label)

        # 备注
        remark = self.bom_data['Remark'] if 'Remark' in self.bom_data.keys() else ''
        remark_label = QLabel(str(remark) if remark else "无")
        remark_label.setStyleSheet("font-size: 14px; color: #262626;")
        remark_label.setWordWrap(True)
        info_layout.addRow("备注:", remark_label)

        layout.addWidget(info_group)

        # BOM结构
        structure_group = QGroupBox("BOM结构")
        structure_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-size: 13px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
                color: #262626;
            }
        """)

        structure_layout = QVBoxLayout(structure_group)
        structure_layout.setContentsMargins(10, 15, 10, 10)  # 增加内边距

        # BOM结构表格
        self.structure_table = QTableWidget()
        self.structure_table.setColumnCount(6)
        self.structure_table.setHorizontalHeaderLabels([
            "物料编码", "物料名称", "规格型号", "用量", "损耗率", "备注"
        ])

        # 设置表格样式
        self.structure_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e8e8e8;
                background-color: white;
                alternate-background-color: #f8f9fa;
                selection-background-color: #e6f7ff;
                selection-color: #262626;
                border: 1px solid #e8e8e8;
                border-radius: 4px;
                font-size: 13px;
                min-height: 300px;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                color: #262626;
                padding: 8px 6px;
                border: none;
                border-bottom: 1px solid #e8e8e8;
                font-weight: 500;
                font-size: 12px;
            }
        """)

        # 设置表格属性
        self.structure_table.setAlternatingRowColors(True)
        self.structure_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.structure_table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        # 调整列宽
        header = self.structure_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 物料编码
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # 物料名称
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 规格型号
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 用量
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 损耗率
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 备注

        structure_layout.addWidget(self.structure_table)
        layout.addWidget(structure_group)

        # 填充BOM结构数据
        self.populate_structure_table()

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-size: 14px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        close_btn.clicked.connect(self.accept)

        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

    def populate_structure_table(self):
        """填充BOM结构表格"""
        if not self.bom_lines:
            return

        self.structure_table.setRowCount(len(self.bom_lines))

        for row, line in enumerate(self.bom_lines):
            # 修复SQLite Row对象的访问方式
            child_item_code = line['ChildItemCode'] if 'ChildItemCode' in line.keys() else ''
            child_item_name = line['ChildItemName'] if 'ChildItemName' in line.keys() else ''
            child_item_spec = line['ChildItemSpec'] if 'ChildItemSpec' in line.keys() else ''
            qty_per = line['QtyPer'] if 'QtyPer' in line.keys() else 0
            scrap_factor = line['ScrapFactor'] if 'ScrapFactor' in line.keys() else 0
            remark = line['Remark'] if 'Remark' in line.keys() else ''

            # 物料编码
            self.structure_table.setItem(row, 0, QTableWidgetItem(str(child_item_code)))
            # 物料名称
            self.structure_table.setItem(row, 1, QTableWidgetItem(str(child_item_name)))
            # 规格型号
            self.structure_table.setItem(row, 2, QTableWidgetItem(str(child_item_spec)))
            # 用量
            self.structure_table.setItem(row, 3, QTableWidgetItem(str(qty_per)))
            # 损耗率
            scrap_factor_display = f"{scrap_factor * 100:.1f}%" if scrap_factor else "0.0%"
            self.structure_table.setItem(row, 4, QTableWidgetItem(scrap_factor_display))
            # 备注
            self.structure_table.setItem(row, 5, QTableWidgetItem(str(remark) if remark else ""))


class BomImportDialog(QDialog):
    """BOM导入对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("BOM矩阵导入 - 智能同步更新")
        self.resize(600, 500)
        self.setMinimumSize(550, 450)
        self.setMaximumSize(800, 600)
        self.setModal(True)
        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 标题
        title_label = QLabel("BOM矩阵导入 - 智能同步更新BOM结构")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #333;
                padding: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # 文件选择区域
        file_group = QGroupBox("选择文件")
        file_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-size: 13px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
                color: #262626;
            }
        """)

        file_layout = QVBoxLayout(file_group)

        # 文件路径显示
        self.file_path_label = QLabel("未选择文件")
        self.file_path_label.setStyleSheet("""
            QLabel {
                padding: 8px;
                border: 1px solid #d9d9d9;
                border-radius: 4px;
                background-color: #fafafa;
                color: #666;
                font-size: 13px;
            }
        """)

        # 选择文件按钮
        select_file_btn = QPushButton("选择文件")
        select_file_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: 500;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        select_file_btn.clicked.connect(self.select_file)

        file_layout.addWidget(self.file_path_label)
        file_layout.addWidget(select_file_btn)
        layout.addWidget(file_group)

        # 文件格式说明
        format_group = QGroupBox("文件格式说明和导入规则")
        format_group.setStyleSheet("""
            QGroupBox {
                font-weight: 500;
                border: 1px solid #e8e8e8;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-size: 13px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
                color: #262626;
            }
        """)

        format_layout = QVBoxLayout(format_group)
        
        format_text = QTextEdit()
        format_text.setMaximumHeight(200)
        format_text.setReadOnly(True)
        format_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #d9d9d9;
                border-radius: 4px;
                padding: 8px;
                font-size: 12px;
                line-height: 1.4;
            }
        """)
        
        format_content = """
导入规则说明：
1、导入模板示例请参考导出的表格模板，按照模板格式导入
2. 数量为0表示不使用该零部件（会从BOM中移除）
3. 系统会自动匹配成品和零部件物料（基于品牌和规格）
4. 支持的文件格式：.csv, .xlsx, .xls

智能匹配规则：
• 成品匹配：优先使用品牌字段匹配BOM名称，其次使用其他产品信息
• 零部件匹配：主要基于编码和规格匹配，名称作为辅助参考
• 数据同步：支持增量更新，只更新有变化的BOM关系
• 自动清理：数量为0的零部件会自动从BOM结构中移除
• 新增支持：Excel中新增的零部件会自动添加到BOM中

注意事项：
• 只处理启用状态的物料，禁用的物料会被自动过滤
• 导入前建议先导出现有BOM作为备份
• 支持BOM历史记录，所有操作都会被记录
        """
        format_text.setPlainText(format_content)
        format_layout.addWidget(format_text)
        layout.addWidget(format_group)

        # 导入进度
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)

        self.import_btn = QPushButton("开始导入")
        self.import_btn.setStyleSheet("""
            QPushButton {
                background: #52c41a;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-size: 14px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #73d13d;
            }
        """)
        self.import_btn.clicked.connect(self.start_import)
        self.import_btn.setEnabled(False)

        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(self.import_btn)
        layout.addLayout(button_layout)

    def select_file(self):
        """选择文件"""
        from PySide6.QtWidgets import QFileDialog
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel矩阵文件",
            "",
            "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path = file_path
            self.file_path_label.setText(f"已选择: {file_path}")
            self.import_btn.setEnabled(True)

    def start_import(self):
        """开始导入"""
        if not hasattr(self, 'file_path'):
            QMessageBox.warning(self, "警告", "请先选择文件！")
            return

        try:
            # 显示进度条
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)  # 不确定进度
            self.import_btn.setEnabled(False)
            
            # 导入BOM数据（使用新的矩阵导入服务）
            success_count, errors, warnings = BomMatrixImportService.import_matrix_excel(self.file_path)
            
            # 隐藏进度条
            self.progress_bar.setVisible(False)
            self.import_btn.setEnabled(True)
            
            # 显示结果
            if success_count > 0:
                QMessageBox.information(self, "导入成功", f"成功导入 {success_count} 个BOM关系！")
                self.accept()
            else:
                error_msg = "没有成功导入任何BOM关系。\n\n"
                if errors:
                    error_msg += "错误信息:\n"
                    for error in errors[:5]:  # 只显示前5个错误
                        error_msg += f"• {error}\n"
                    if len(errors) > 5:
                        error_msg += f"... 还有 {len(errors) - 5} 个错误\n"
                
                QMessageBox.warning(self, "导入失败", error_msg)
                
        except Exception as e:
            self.progress_bar.setVisible(False)
            self.import_btn.setEnabled(True)
            QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")


class BomHistoryDialog(QDialog):
    """BOM操作历史对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("BOM操作历史")
        self.setModal(True)
        self.resize(1000, 600)
        self.setMinimumSize(800, 500)
        self.setup_ui()
        self.load_history()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 标题
        title_label = QLabel("BOM操作历史")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #333;
                padding: 10px;
            }
        """)
        layout.addWidget(title_label)

        # 统计信息
        stats_frame = QFrame()
        stats_frame.setStyleSheet("""
            QFrame {
                background: #f5f5f5;
                border: 1px solid #d9d9d9;
                border-radius: 6px;
                padding: 10px;
            }
        """)
        stats_layout = QHBoxLayout(stats_frame)
        
        self.stats_label = QLabel("正在加载统计信息...")
        self.stats_label.setStyleSheet("""
            QLabel {
                color: #666;
                font-size: 13px;
            }
        """)
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()
        
        # 操作按钮
        refresh_btn = QPushButton("刷新")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background: #1890ff;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #40a9ff;
            }
        """)
        refresh_btn.clicked.connect(self.load_history)
        stats_layout.addWidget(refresh_btn)
        
        # 导出按钮
        export_btn = QPushButton("导出")
        export_btn.setStyleSheet("""
            QPushButton {
                background: #52c41a;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #73d13d;
            }
        """)
        export_btn.clicked.connect(self.export_history)
        stats_layout.addWidget(export_btn)
        
        # 清空按钮
        clear_btn = QPushButton("清空")
        clear_btn.setStyleSheet("""
            QPushButton {
                background: #ff4d4f;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #ff7875;
            }
        """)
        clear_btn.clicked.connect(self.clear_history)
        stats_layout.addWidget(clear_btn)
        
        layout.addWidget(stats_frame)

        # 历史记录表格
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(6)
        self.history_table.setHorizontalHeaderLabels([
            "时间", "操作类型", "BOM名称", "操作目标", "操作用户", "备注"
        ])
        
        # 设置表格样式
        self.history_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #d9d9d9;
                border-radius: 6px;
                background: white;
                gridline-color: #f0f0f0;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background: #e6f7ff;
            }
            QHeaderView::section {
                background: #fafafa;
                padding: 10px;
                border: none;
                border-right: 1px solid #f0f0f0;
                border-bottom: 1px solid #d9d9d9;
                font-weight: 500;
                color: #262626;
            }
        """)
        
        # 设置列宽
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 时间
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 操作类型
        header.setSectionResizeMode(2, QHeaderView.Stretch)          # BOM名称
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 操作目标
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 操作用户
        header.setSectionResizeMode(5, QHeaderView.Stretch)          # 备注
        
        layout.addWidget(self.history_table)

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background: #d9d9d9;
                color: #262626;
                border: none;
                padding: 10px 25px;
                border-radius: 5px;
                font-size: 14px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #bfbfbf;
            }
        """)
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)

    def load_history(self):
        """加载历史记录"""
        try:
            # 获取历史记录
            history_records = BomHistoryService.get_all_bom_history(200)
            
            # 更新统计信息
            summary = BomHistoryService.get_operation_summary()
            stats_text = f"总操作数: {summary['total']} | "
            for op_type, count in summary['by_type'].items():
                stats_text += f"{op_type}: {count} | "
            stats_text = stats_text.rstrip(" | ")
            self.stats_label.setText(stats_text)
            
            # 更新表格
            self.history_table.setRowCount(len(history_records))
            
            for row, record in enumerate(history_records):
                # 格式化时间
                time_str = record.get('CreatedDate', '')
                if time_str:
                    try:
                        from datetime import datetime
                        dt = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
                        time_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
                
                # 操作类型
                op_type = record.get('OperationType', '')
                op_target = record.get('OperationTarget', '')
                if op_target == 'HEADER':
                    type_text = f"{op_type} BOM"
                else:
                    type_text = f"{op_type} 零部件"
                
                # BOM名称
                bom_name = record.get('BomName', '未知BOM')
                
                # 操作目标
                if op_target == 'HEADER':
                    target_text = "BOM主表"
                else:
                    child_name = record.get('ChildItemName', '未知零部件')
                    target_text = f"零部件: {child_name}"
                
                # 操作用户
                user = record.get('OperationUser', '系统')
                
                # 备注
                remark = record.get('Remark', '')
                
                # 设置表格项
                self.history_table.setItem(row, 0, QTableWidgetItem(time_str))
                self.history_table.setItem(row, 1, QTableWidgetItem(type_text))
                self.history_table.setItem(row, 2, QTableWidgetItem(bom_name))
                self.history_table.setItem(row, 3, QTableWidgetItem(target_text))
                self.history_table.setItem(row, 4, QTableWidgetItem(user))
                self.history_table.setItem(row, 5, QTableWidgetItem(remark))
                
                # 设置操作类型颜色
                if op_type == 'CREATE':
                    color = QColor(82, 196, 26)  # 绿色
                elif op_type == 'UPDATE':
                    color = QColor(24, 144, 255)  # 蓝色
                elif op_type == 'DELETE':
                    color = QColor(245, 34, 45)  # 红色
                elif op_type == 'IMPORT':
                    color = QColor(114, 46, 209)  # 紫色
                else:
                    color = QColor(140, 140, 140)  # 灰色
                
                type_item = self.history_table.item(row, 1)
                if type_item:
                    type_item.setForeground(color)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载历史记录失败: {str(e)}")

    def export_history(self):
        """导出历史记录到Excel"""
        try:
            # 选择保存文件
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "导出BOM历史记录", 
                f"BOM历史记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel文件 (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM操作历史"
            
            # 设置表头
            headers = ["时间", "操作类型", "BOM名称", "操作目标", "操作用户", "备注", "操作来源"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True, name='宋体', size=10)
                cell.fill = openpyxl.styles.PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # 获取历史记录
            history_records = BomHistoryService.get_all_bom_history(1000)  # 导出更多记录
            
            # 填充数据
            for row, record in enumerate(history_records, 2):
                # 格式化时间
                time_str = record.get('CreatedDate', '')
                if time_str:
                    try:
                        dt = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
                        time_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
                
                # 操作类型
                op_type = record.get('OperationType', '')
                op_target = record.get('OperationTarget', '')
                if op_target == 'HEADER':
                    type_text = f"{op_type} BOM"
                else:
                    type_text = f"{op_type} 零部件"
                
                # BOM名称
                bom_name = record.get('BomName', '未知BOM')
                
                # 操作目标
                if op_target == 'HEADER':
                    target_text = "BOM主表"
                else:
                    child_name = record.get('ChildItemName', '未知零部件')
                    target_text = f"零部件: {child_name}"
                
                # 操作用户
                user = record.get('OperationUser', '系统')
                
                # 备注
                remark = record.get('Remark', '')
                
                # 操作来源
                source = record.get('OperationSource', 'UI')
                
                # 设置数据
                ws.cell(row=row, column=1, value=time_str)
                ws.cell(row=row, column=2, value=type_text)
                ws.cell(row=row, column=3, value=bom_name)
                ws.cell(row=row, column=4, value=target_text)
                ws.cell(row=row, column=5, value=user)
                ws.cell(row=row, column=6, value=remark)
                ws.cell(row=row, column=7, value=source)
                
                # 设置字体
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.font = openpyxl.styles.Font(name='宋体', size=10)
                    cell.alignment = openpyxl.styles.Alignment(vertical='center')
            
            # 设置列宽
            column_widths = [20, 15, 25, 30, 12, 40, 10]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # 设置边框
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(history_records) + 2):
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
            
            # 保存文件
            wb.save(file_path)
            QMessageBox.information(self, "成功", f"历史记录已导出到: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出历史记录失败: {str(e)}")

    def clear_history(self):
        """清空历史记录"""
        try:
            # 确认对话框
            reply = QMessageBox.question(
                self, 
                "确认清空", 
                "确定要清空所有BOM操作历史记录吗？\n此操作不可恢复！",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # 清空历史记录
            from app.db import execute
            execute("DELETE FROM BomOperationHistory")
            
            # 刷新显示
            self.load_history()
            
            QMessageBox.information(self, "成功", "历史记录已清空")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"清空历史记录失败: {str(e)}")


