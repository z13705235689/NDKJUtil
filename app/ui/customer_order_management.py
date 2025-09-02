#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from collections import defaultdict
from typing import Optional, List, Dict, Tuple
from datetime import datetime, timedelta, date as _date

from PySide6.QtCore import Qt, QDate, QThread, Signal, QRect
from PySide6.QtGui import QFont, QColor, QPainter
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QFrame, QLineEdit, QComboBox, QAbstractItemView,
    QMessageBox, QTabWidget, QHeaderView, QDateEdit, QFileDialog,
    QProgressBar, QTextBrowser, QDialog, QAbstractScrollArea, QSizePolicy
)

from app.services.customer_order_service import CustomerOrderService

# Excel导出相关导入
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -------------------- 小工具 --------------------
def _safe_parse_date(s):
    if not s:
        return None
    if isinstance(s, _date):
        return s
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%y", "%m/%d/%Y"):
        try:
            d = datetime.strptime(str(s), fmt).date()
            if d.year < 2000:
                d = d.replace(year=d.year + 2000)
            return d
        except Exception:
            pass
    return None

def _cw(d: _date) -> int:
    return d.isocalendar()[1]

def _norm_int(v, default=0):
    try:
        if v in (None, ""):
            return default
        return int(float(v))
    except Exception:
        return default

def _get(line: dict, *names, default=None):
    for n in names:
        if n in line and line[n] not in (None, ""):
            return line[n]
    return default

def _week_start(d: _date) -> _date:
    return d - timedelta(days=d.weekday())

def _build_week_cols(min_d: _date, max_d: _date) -> Tuple[List[Tuple[str, _date|int]], Dict[int, List[_date]], List[int]]:
    """根据起止日期构造连续周列以及每年的周分组，并在每年后追加“合计”列。"""
    if not (min_d and max_d):
        return [], {}, []
    start = _week_start(min_d)
    end   = _week_start(max_d)
    weeks = []
    cur = start
    while cur <= end:
        weeks.append(cur)
        cur += timedelta(days=7)
    by_year = defaultdict(list)
    for w in weeks:
        by_year[w.isocalendar()[0]].append(w)
    years = sorted(by_year.keys())
    colspec: List[Tuple[str, _date|int]] = []
    for y in years:
        for w in by_year[y]:
            colspec.append(("week", w))
        colspec.append(("sum", y))
    return colspec, by_year, years


# -------------------- 两行表头 --------------------
class TwoRowHeader(QHeaderView):
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setDefaultAlignment(Qt.AlignCenter)
        self._top_font = QFont(); self._top_font.setBold(True)
        self._bottom_font = QFont()
        h = self.fontMetrics().height()
        self.setFixedHeight(int(h * 2.4))

    def sizeHint(self):
        s = super().sizeHint()
        h = self.fontMetrics().height()
        s.setHeight(int(h * 2.4))
        return s

    def paintSection(self, painter: QPainter, rect: QRect, logicalIndex: int):
        if not rect.isValid():
            return
        super().paintSection(painter, rect, logicalIndex)

        table = self.parent()
        item = table.horizontalHeaderItem(logicalIndex) if table else None
        top = item.text() if item else ""
        bottom = item.data(Qt.UserRole) if (item and item.data(Qt.UserRole) is not None) else ""

        painter.save()
        painter.setPen(QColor("#333333"))
        painter.setFont(self._top_font)
        topRect = QRect(rect.left(), rect.top() + 2, rect.width(), rect.height() // 2)
        painter.drawText(topRect, Qt.AlignCenter, str(top))
        painter.setFont(self._bottom_font)
        bottomRect = QRect(rect.left(), rect.top() + rect.height() // 2 - 2, rect.width(), rect.height() // 2)
        painter.drawText(bottomRect, Qt.AlignCenter, str(bottom))
        painter.restore()


# -------------------- 主界面 --------------------
class CustomerOrderManagement(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_version_list()

    def init_ui(self):
        self.setWindowTitle("客户订单管理")
        # 移除最小尺寸设置，让页面适应父容器大小
        # self.setMinimumSize(1200, 800)
        
        # 设置大小策略，让页面适应父容器
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        main_layout = QVBoxLayout()

        # 顶部工具栏
        toolbar_layout = QHBoxLayout()
        import_btn = QPushButton("导入TXT文件")
        import_btn.setStyleSheet("QPushButton{background:#28a745;color:#fff;border:none;padding:8px 16px;border-radius:4px;font-weight:bold;}"
                                 "QPushButton:hover{background:#218838;}")
        import_btn.clicked.connect(self.import_txt_file)
        toolbar_layout.addWidget(import_btn)

        version_btn = QPushButton("版本管理")
        version_btn.setStyleSheet("QPushButton{background:#17a2b8;color:#fff;border:none;padding:8px 16px;border-radius:4px;font-weight:bold;}"
                                  "QPushButton:hover{background:#138496;}")
        version_btn.clicked.connect(self.show_version_management)
        toolbar_layout.addWidget(version_btn)

        toolbar_layout.addStretch()

        refresh_btn = QPushButton("刷新")
        refresh_btn.setStyleSheet("QPushButton{background:#6c757d;color:#fff;border:none;padding:8px 16px;}"
                                  "QPushButton:hover{background:#5a6268;}")
        refresh_btn.clicked.connect(self.refresh_data)
        toolbar_layout.addWidget(refresh_btn)

        main_layout.addLayout(toolbar_layout)

        # 页签
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet(
            "QTabWidget::pane{border:1px solid #dee2e6;background:white;}"
            "QTabBar::tab{background:#f8f9fa;border:1px solid #dee2e6;padding:8px 16px;margin-right:2px;}"
            "QTabBar::tab:selected{background:white;border-bottom:2px solid #007bff;}"
        )

        self.create_kanban_tab()
        self.create_order_details_tab()
        self.create_import_history_tab()

        main_layout.addWidget(self.tab_widget)
        self.setLayout(main_layout)

    # ---------- 看板页 ----------
    def create_kanban_tab(self):
        kanban_widget = QWidget()
        kanban_layout = QVBoxLayout()

        control_panel = QFrame()
        control_panel.setFrameStyle(QFrame.StyledPanel)
        control_panel.setStyleSheet("QFrame{background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;}")
        control_layout = QVBoxLayout()

        # 版本选择 + 日期范围
        version_layout = QHBoxLayout()
        version_layout.addWidget(QLabel("订单版本:"))
        self.version_combo = QComboBox()
        self.version_combo.addItem("全部版本汇总")
        self.version_combo.currentTextChanged.connect(self.on_version_changed)
        self.version_combo.setMinimumWidth(200)  # 设置最小宽度
        self.version_combo.setMinimumHeight(12)   # 设置最小高度
        self.version_combo.setStyleSheet("""
            QComboBox {
                padding: 6px 12px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background: white;
                min-width: 200px;
                max-width: 300px;
                min-height: 12px;
                max-height: 12px;
            }
            QComboBox:focus {
                border-color: #007bff;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #6c757d;
                margin-right: 5px;
            }
        """)
        version_layout.addWidget(self.version_combo)

        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("日期范围:"))
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDate(QDate.currentDate().addDays(-30))
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        date_layout.addWidget(self.start_date_edit)

        date_layout.addWidget(QLabel("至"))
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDate(QDate.currentDate().addDays(30))
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        date_layout.addWidget(self.end_date_edit)

        apply_btn = QPushButton("应用筛选")
        apply_btn.setStyleSheet("QPushButton{background:#007bff;color:#fff;border:none;padding:6px 12px;border-radius:4px;}"
                                "QPushButton:hover{background:#0069d9;}")
        apply_btn.clicked.connect(self.load_kanban_data)

        export_btn = QPushButton("导出Excel")
        export_btn.setStyleSheet("QPushButton{background:#28a745;color:#fff;border:none;padding:6px 12px;border-radius:4px;}"
                                "QPushButton:hover{background:#218838;}")
        export_btn.clicked.connect(self.export_kanban_to_excel)

        version_layout.addLayout(date_layout)
        version_layout.addWidget(apply_btn)
        version_layout.addWidget(export_btn)
        version_layout.addStretch()
        control_layout.addLayout(version_layout)

        # 订单类型过滤
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("订单类型:"))
        self.order_type_combo = QComboBox()
        self.order_type_combo.addItems(["全部", "F(正式)", "P(预测)"])
        self.order_type_combo.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.order_type_combo)
        filter_layout.addStretch()
        control_layout.addLayout(filter_layout)

        control_panel.setLayout(control_layout)
        kanban_layout.addWidget(control_panel)

        # 看板表格
        self.kanban_table = QTableWidget()
        self.kanban_table.setAlternatingRowColors(True)
        self.kanban_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kanban_table.setHorizontalHeader(TwoRowHeader(Qt.Horizontal, self.kanban_table))
        hdr = self.kanban_table.horizontalHeader()
        hdr.setFixedHeight(int(self.fontMetrics().height()*2.4))
        try:
            hdr.repaint()
            hdr.updateGeometry()
        except Exception:
            pass
        try:
            policy = QAbstractScrollArea.SizeAdjustPolicy.AdjustToContentsOnFirstShow
        except AttributeError:
            policy = getattr(QAbstractScrollArea, "AdjustToContentsOnFirstShow", QAbstractScrollArea.AdjustToContents)
        self.kanban_table.setSizeAdjustPolicy(policy)
        self.kanban_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # 设置表格的最小高度，确保在客户端范围内显示
        self.kanban_table.setMinimumHeight(300)

        kanban_layout.addWidget(self.kanban_table)
        kanban_widget.setLayout(kanban_layout)
        self.tab_widget.addTab(kanban_widget, "看板视图")

    # ------- 数据加载 -------
    def on_version_changed(self, version_text: str):
        try:
            if version_text == "全部版本汇总":
                self.load_kanban_data()
                return
            import re
            m = re.search(r'^(\d+) - ', version_text)
            if not m:
                self.load_kanban_data()
                return
            version_id = int(m.group(1))
            data = CustomerOrderService.get_order_lines_by_import_version(version_id)
            all_dates = []
            for ln in data or []:
                d = _safe_parse_date(_get(ln, "DeliveryDate", "due_date"))
                if d:
                    all_dates.append(d)
            if all_dates:
                all_dates.sort()
                self.start_date_edit.setDate(QDate(all_dates[0].year, all_dates[0].month, all_dates[0].day))
                self.end_date_edit.setDate(QDate(all_dates[-1].year, all_dates[-1].month, all_dates[-1].day))
            self.load_kanban_data()
        except Exception as e:
            print(f"版本切换失败: {e}")

    def on_filter_changed(self):
        self.load_kanban_data()

    def load_kanban_data(self):
        try:
            version_text = self.version_combo.currentText()
            sd = self.start_date_edit.date().toString("yyyy-MM-dd")
            ed = self.end_date_edit.date().toString("yyyy-MM-dd")
            if version_text != "全部版本汇总":
                import re
                m = re.search(r'^(\d+) - ', version_text)
                version_id = int(m.group(1)) if m else None
            else:
                version_id = None

            if version_id:
                data = CustomerOrderService.get_order_lines_by_import_version(version_id)
                self.display_kanban_data_by_version(data, sd, ed, self.order_type_combo.currentText())
            else:
                # 汇总口径：把 Firm/Predict 拆成伪明细
                rows = CustomerOrderService.get_ndlutil_kanban_data(start_date=sd, end_date=ed)
                expanded = []
                for r in rows or []:
                    common = {
                        "SupplierCode": r.get("SupplierCode",""),
                        "SupplierName": r.get("SupplierName",""),
                        "ItemNumber":   r.get("ItemNumber",""),
                        "ItemDescription": r.get("ItemDescription",""),
                        "DeliveryDate": r.get("DeliveryDate",""),
                        "CalendarWeek": r.get("CalendarWeek",""),
                        "ReleaseDate":  r.get("ReleaseDate",""),
                        "ReleaseId":    r.get("ReleaseId",""),
                        "PurchaseOrder":r.get("PurchaseOrder",""),
                        "ReceiptQuantity": 0,
                        "CumReceived":  0,
                    }
                    f = _norm_int(r.get("FirmQty"), 0)
                    p = _norm_int(r.get("ForecastQty"), 0)
                    if f:
                        e = dict(common); e["OrderType"]="F"; e["RequiredQty"]=f; expanded.append(e)
                    if p:
                        e = dict(common); e["OrderType"]="P"; e["RequiredQty"]=p; expanded.append(e)
                self.display_kanban_data_by_version(expanded, sd, ed, self.order_type_combo.currentText())
        except Exception as e:
            print(f"加载看板数据失败: {e}")

    # ------- 渲染（行集合固定，列按范围变动）-------
    def display_kanban_data_by_version(self, data: list, start_date: str, end_date: str, order_type: str):
        """
        行集合 = 该版本中检索到的所有 (Supplier, PN)，固定展示；
        列（CW/年合计）随页面日期范围变化；
        数量来自【按日期范围 + 类型过滤】后的数据进行周聚合。
        """
        sd = _safe_parse_date(start_date)
        ed = _safe_parse_date(end_date)
        if sd and ed and sd > ed:
            sd, ed = ed, sd
        page_typ = (order_type or "").upper()

        # 1) 构造列
        colspec, by_year, years = _build_week_cols(sd, ed)

        # 2) 行集合：整版数据收集 (Supplier, PN)
        groups_all = defaultdict(lambda: {"release": {}})
        for ln in (data or []):
            sup = _get(ln, "SupplierCode", "supplier_code", "Supplier", "supplier", "") or ""
            pn = _get(ln, "ItemNumber", "item_number", "Item", "item", "") or ""
            key = (sup, pn)
            ri = groups_all[key]["release"]

            def _set_if_val(k, v, cast_int=False):
                if v in (None, ""): return
                if cast_int: v = _norm_int(v, 0)
                if not ri.get(k): ri[k] = v

            _set_if_val("release_date", _get(ln, "ReleaseDate", "release_date", "release_date_text"))
            _set_if_val("release_id", _get(ln, "ReleaseId", "release_id"))
            _set_if_val("purchase_order", _get(ln, "PurchaseOrder", "OrderNumber", "purchase_order", "order_number"))
            _set_if_val("receipt_qty", _get(ln, "ReceiptQuantity", "receipt_qty"), cast_int=True)
            _set_if_val("cum_received", _get(ln, "CumReceived", "cum_received"), cast_int=True)

        # 3) 过滤后的数据
        lines_filtered = []
        for ln in (data or []):
            d = _safe_parse_date(_get(ln, "DeliveryDate", "due_date"))
            if not d: continue
            if sd and d < sd: continue
            if ed and d > ed: continue
            typ = (_get(ln, "OrderType", "order_type", "FP", default="") or "").upper()
            if page_typ in ("F(正式)", "F") and typ != "F": continue
            if page_typ in ("P(预测)", "P") and typ != "P": continue
            ln = dict(ln)
            ln["__week__"] = _week_start(d)
            ln["__fp__"] = "F" if typ == "F" else "P"
            lines_filtered.append(ln)

        week_qty = defaultdict(lambda: defaultdict(int))
        fp_map = {}
        for ln in lines_filtered:
            sup = _get(ln, "SupplierCode", "supplier_code", "Supplier", "supplier", "")
            pn = _get(ln, "ItemNumber", "item_number", "Item", "item", "")
            wk = ln["__week__"]
            q = _norm_int(_get(ln, "RequiredQty", "req_qty", default=0), 0)
            week_qty[(sup, pn)][wk] += q
            cur = fp_map.get((sup, pn, wk))
            if (cur is None) or (cur == "P" and ln["__fp__"] == "F"):
                fp_map[(sup, pn, wk)] = ln["__fp__"]

        # 4) 表头
        fixed_headers = [
            "Release Date", "Release ID", "PN", "Des", "Project", "Item",
            "Purchase Order", "Receipt Quantity", "Cum Received"
        ]
        headers_count = len(fixed_headers) + len(colspec) + 1
        self.kanban_table.clear()
        self.kanban_table.setColumnCount(headers_count)

        for i, title in enumerate(fixed_headers):
            item = QTableWidgetItem(title)
            self.kanban_table.setHorizontalHeaderItem(i, item)

        base_col = len(fixed_headers)
        for i, (kind, val) in enumerate(colspec):
            if kind == "week":
                cw = f"CW{val.isocalendar()[1]:02d}"
                it = QTableWidgetItem(cw)
                it.setData(Qt.UserRole, val.strftime("%Y/%m/%d"))
            else:
                it = QTableWidgetItem(f"{val}合计")
            self.kanban_table.setHorizontalHeaderItem(base_col + i, it)
        self.kanban_table.setHorizontalHeaderItem(headers_count - 1, QTableWidgetItem("Total"))

        # 5) 行集合（按产品型号规则排序）
        def sort_key(item):
            sup, pn = item
            if not pn:
                return (999, 999, sup)  # 空PN排最后
            
            # 去掉最后一位字母后缀，获取基础产品型号
            if len(pn) > 1 and pn[-1].isalpha():
                base_pn = pn[:-1]  # 去掉最后一位字母
            else:
                base_pn = pn
            
            # 定义排序优先级（基于去掉最后一位字母的基础型号）
            priority_map = {
                "R001H368": 1,  # Passat rear double
                "R001H369": 2,  # Passat rear single
                "R001P320": 3,  # Tiguan L rear double
                "R001P313": 4,  # Tiguan L rear single
                "R001J139": 5,  # A5L rear double
                "R001J140": 6,  # A5L rear single
                "R001J141": 7,  # Lavida rear double
                "R001J142": 8,  # Lavida rear single
            }
            
            # 先尝试完整基础型号匹配
            if base_pn in priority_map:
                priority = priority_map[base_pn]
            else:
                # 如果完整匹配失败，尝试基础项目匹配
                if len(base_pn) > 1 and base_pn[-1].isdigit():
                    base = base_pn[:-1]  # 去掉最后一位数字
                else:
                    base = base_pn
                
                base_priority_map = {
                    "R001H36": 10,  # Passat
                    "R001P32": 20,  # Tiguan L
                    "R001J13": 30,  # A5L
                    "R001J14": 40,  # Lavida
                }
                priority = base_priority_map.get(base, 999)  # 不匹配的排最后
            
            return (priority, sup, base_pn, pn)  # 添加完整PN作为最后的排序依据
        
        keys_all = sorted(groups_all.keys(), key=sort_key)
        
        data_rows = len(keys_all)
        self.kanban_table.setRowCount(data_rows + 1)  # +1 行留给 TOTAL

        def project_name(pn: str) -> str:
            if not pn:
                return "UNKNOWN"
            
            # 去掉最后一位字母后缀，获取基础产品型号
            if len(pn) > 1 and pn[-1].isalpha():
                base_pn = pn[:-1]  # 去掉最后一位字母
            else:
                base_pn = pn
            
            # 完整的产品型号映射（基于去掉最后一位字母的基础型号）
            full_pn_map = {
                "R001H368": "Passat rear double",
                "R001H369": "Passat rear single",
                "R001P320": "Tiguan L rear double", 
                "R001P313": "Tiguan L rear single",
                "R001J139": "A5L rear double",
                "R001J140": "A5L rear single",
                "R001J141": "Lavida rear double",
                "R001J142": "Lavida rear single"
            }
            
            # 先尝试完整基础型号匹配
            if base_pn in full_pn_map:
                return full_pn_map[base_pn]
            
            # 如果完整匹配失败，尝试基础项目匹配
            if len(base_pn) > 1 and base_pn[-1].isdigit():
                base = base_pn[:-1]  # 去掉最后一位数字
            else:
                base = base_pn
            
            base_map = {
                "R001H36": "Passat",
                "R001P32": "Tiguan L", 
                "R001J13": "A5L",
                "R001J14": "Lavida"
            }
            
            project = base_map.get(base, "UNKNOWN")
            if project == "UNKNOWN":
                print(f"警告：产品型号 '{pn}' 没有匹配到项目，默认放到最后")
            return project

        # 6) 填充数据行
        for row_idx, (sup, pn) in enumerate(keys_all):
            ri = groups_all[(sup, pn)]["release"]
            rd_obj = _safe_parse_date(ri.get("release_date"))
            rd_txt = rd_obj.strftime("%Y/%m/%d") if rd_obj else (ri.get("release_date") or "")

            project_result = project_name(pn)

            fixed_vals = [
                rd_txt, str(ri.get("release_id", "") or ""), pn,
                "PEMM ASSY", project_result, "Gross Reqs",
                sup, str(_norm_int(ri.get("receipt_qty"), 0)),
                str(_norm_int(ri.get("cum_received"), 0)),
            ]
            for c, v in enumerate(fixed_vals):
                self.kanban_table.setItem(row_idx, c, QTableWidgetItem(v))

            row_total = 0
            cursor_col = base_col
            for kind, val in colspec:
                if kind == "week":
                    qty = _norm_int(week_qty[(sup, pn)].get(val, 0), 0)
                    row_total += qty
                    cell = QTableWidgetItem(str(qty))
                    fp = fp_map.get((sup, pn, val))
                    if qty > 0 and fp:
                        cell.setBackground(QColor("#C6E0B4") if fp == "F" else QColor("#FFF2CC"))
                    self.kanban_table.setItem(row_idx, cursor_col, cell)
                else:
                    s = sum(_norm_int(week_qty[(sup, pn)].get(w, 0), 0) for w in by_year[val])
                    cell = QTableWidgetItem(str(s))
                    f = cell.font();
                    f.setBold(True);
                    cell.setFont(f)
                    cell.setBackground(QColor("#DDEBF7"))
                    self.kanban_table.setItem(row_idx, cursor_col, cell)
                cursor_col += 1

            self.kanban_table.setItem(row_idx, headers_count - 1, QTableWidgetItem(str(row_total)))

        # 7) TOTAL 行
        total_row = data_rows
        self.kanban_table.setItem(total_row, 0, QTableWidgetItem("TOTAL"))
        # 只从 CW 开始统计（前9列不算）
        for col in range(base_col, headers_count):
            s = 0
            for r in range(0, data_rows):
                it = self.kanban_table.item(r, col)
                try:
                    s += int(float(it.text())) if it and it.text().strip() else 0
                except:
                    pass
            item = QTableWidgetItem(str(s))
            f = item.font();
            f.setBold(True);
            item.setFont(f)
            self.kanban_table.setItem(total_row, col, item)

        # 列宽
        hdr = self.kanban_table.horizontalHeader()
        for i in range(9):
            hdr.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        for c in range(9, headers_count):
            hdr.setSectionResizeMode(c, QHeaderView.Fixed)
            self.kanban_table.setColumnWidth(c, 64)

        self.kanban_table.resizeRowsToContents()
        self.apply_kanban_styling()

    # ====== 样式：通用 ======
    def apply_kanban_styling(self):
        # 表头加粗
        hdr = self.kanban_table.horizontalHeader()
        for c in range(self.kanban_table.columnCount()):
            it = self.kanban_table.horizontalHeaderItem(c)
            if it:
                f = QFont(it.font())
                f.setBold(True)
                it.setFont(f)
        # 交替底色
        self.kanban_table.setAlternatingRowColors(True)
        self.kanban_table.setStyleSheet(
            "QTableWidget{gridline-color:#dee2e6;}"
            "QTableWidget::item{padding:2px;}"
            "QTableWidget::item:selected{background:#cfe2ff;}"
        )
        # TOTAL 行仅加粗
        r = self.kanban_table.rowCount() - 1
        if r >= 0:
            for c in range(self.kanban_table.columnCount()):
                it = self.kanban_table.item(r, c)
                if not it:
                    continue
                f = it.font(); f.setBold(True); it.setFont(f)
                it.setBackground(Qt.transparent)  # 保持透明

    def _get_release_cw_from_kanban(self) -> Optional[dict]:
        """从看板数据中获取Release Date对应的CW信息"""
        try:
            if not hasattr(self, 'kanban_table') or self.kanban_table.rowCount() == 0:
                print("DEBUG: 看板表格为空或不存在")
                return None
            
            print(f"DEBUG: 看板表格行数: {self.kanban_table.rowCount()}")
            
            # 查找第一行数据中的Release Date
            for row in range(self.kanban_table.rowCount()):
                release_date_item = self.kanban_table.item(row, 0)  # Release Date列
                if release_date_item and release_date_item.text().strip():
                    release_date_text = release_date_item.text().strip()
                    print(f"DEBUG: 行 {row} Release Date: '{release_date_text}'")
                    
                    if release_date_text and release_date_text != "TOTAL":
                        # 解析Release Date
                        try:
                            # 尝试解析日期格式 YYYY/MM/DD
                            date_obj = datetime.strptime(release_date_text, "%Y/%m/%d")
                            # 计算对应的CW
                            week_num = date_obj.isocalendar()[1]
                            year = date_obj.year
                            print(f"DEBUG: 成功解析日期: {release_date_text} -> CW{week_num:02d}-{year}")
                            return {"week": week_num, "year": year}
                        except ValueError:
                            # 如果解析失败，尝试其他格式
                            try:
                                date_obj = datetime.strptime(release_date_text, "%Y-%m-%d")
                                week_num = date_obj.isocalendar()[1]
                                year = date_obj.year
                                print(f"DEBUG: 成功解析日期: {release_date_text} -> CW{week_num:02d}-{year}")
                                return {"week": week_num, "year": year}
                            except ValueError:
                                print(f"DEBUG: 无法解析日期格式: {release_date_text}")
                                continue
            
            print("DEBUG: 未找到有效的Release Date")
            return None
        except Exception as e:
            print(f"获取Release Date CW信息失败: {e}")
            return None

    def export_kanban_to_excel(self):
        """导出看板数据到Excel"""
        if not hasattr(self, 'kanban_table') or self.kanban_table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "请先生成看板数据")
            return

                # 获取Release Date对应的CW信息用于文件名
        release_cw = self._get_release_cw_from_kanban()
        if release_cw:
            default_filename = f"CW{release_cw['week']:02d}-{release_cw['year']}.xlsx"
        else:
            default_filename = f"客户订单看板_{QDate.currentDate().toString('yyyyMMdd')}.xlsx"
        
        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "导出看板Excel", 
            default_filename,
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            self._export_kanban_excel(file_path)
            QMessageBox.information(self, "导出成功", f"文件已保存到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误：\n{str(e)}")

    def _export_kanban_excel(self, file_path: str):
        """导出看板数据到Excel文件的具体实现"""
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 获取Release Date对应的CW信息
        release_cw = self._get_release_cw_from_kanban()
        print(f"DEBUG: 导出时获取到的CW信息: {release_cw}")
        
        if release_cw:
            # 设置工作表名称为CW几
            ws_title = f"CW{release_cw['week']:02d}"
            ws.title = ws_title
            print(f"DEBUG: 设置工作表名称为: {ws_title}")
        else:
            ws.title = "客户订单看板"
            print("DEBUG: 未获取到CW信息，使用默认工作表名称: 客户订单看板")

        # 获取当前看板数据
        table = self.kanban_table
        rows_count = table.rowCount()
        cols_count = table.columnCount()

        if rows_count == 0 or cols_count == 0:
            raise ValueError("看板数据为空")

        # 定义样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 获取表头信息
        fixed_headers = []
        week_headers = []
        for col in range(cols_count):
            header_item = table.horizontalHeaderItem(col)
            if header_item:
                header_text = header_item.text()
                if col < 9:  # 前9列是固定列
                    fixed_headers.append(header_text)
                else:
                    week_headers.append(header_text)

        # 写入表头（两行）
        # 第一行：固定列标题 + 周标题
        first_row = fixed_headers + week_headers
        ws.append(first_row)

        # 第二行：空行 + 日期
        second_row = [""] * len(fixed_headers)
        for header in week_headers:
            if "CW" in header:
                # 从表头数据中获取日期
                header_item = table.horizontalHeaderItem(len(fixed_headers) + len(second_row))
                if header_item:
                    date_data = header_item.data(Qt.UserRole)
                    if date_data:
                        try:
                            date_obj = datetime.strptime(date_data, "%Y/%m/%d")
                            second_row.append(date_obj.strftime("%m/%d"))
                        except:
                            second_row.append("")
                    else:
                        second_row.append("")
                else:
                    second_row.append("")
            else:
                second_row.append("")

        ws.append(second_row)

        # 写入数据行
        for row in range(rows_count):
            row_data = []
            for col in range(cols_count):
                item = table.item(row, col)
                if item:
                    # 尝试转换为数字
                    try:
                        value = int(float(item.text())) if item.text().strip() else 0
                    except:
                        value = item.text() or ""
                    row_data.append(value)
                else:
                    row_data.append("")
            ws.append(row_data)

        # 工作表标题已在前面设置，这里不需要重复设置

        # ====== 样式设置 ======

        # 1) 列宽设置
        tight_widths = [11, 14, 12, 10, 16, 10, 14, 12, 12]  # 对应9个固定列
        for i, w in enumerate(tight_widths, start=1):
            if i <= len(fixed_headers):
                ws.column_dimensions[get_column_letter(i)].width = w

        # 周列宽度
        for ci in range(len(fixed_headers) + 1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 9.0

        # 2) 冻结窗格：固定列 + 两行表头
        ws.freeze_panes = get_column_letter(len(fixed_headers) + 1) + "3"

        # 3) 表头样式
        for cell in ws[1] + ws[2]:
            cell.font = Font(name="Arial", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 4) 数据行样式
        first_data_row = 3
        for r in range(first_data_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if c <= len(fixed_headers):
                    # 固定列：居中对齐，自动换行
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center",
                        wrap_text=True, shrink_to_fit=True
                    )
                else:
                    # 数字列：居中对齐
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center",
                        wrap_text=False, shrink_to_fit=True
                    )

                # 设置字体
                cell.font = Font(name="Arial")

                # 设置边框（所有有内容的单元格都有边框）
                if cell.value not in (None, "", 0):
                    cell.border = thin_border
                else:
                    # 对于空单元格，也设置边框以保持表格完整性
                    cell.border = thin_border

        # 5) 数字格式设置
        for r in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row,
                              min_col=len(fixed_headers) + 1, max_col=ws.max_column):
            for c in r:
                if isinstance(c.value, (int, float)) and c.value != 0:
                    c.number_format = "0"

        # 6) F/P 着色（从表格背景色获取）
        fill_f = PatternFill("solid", fgColor="C6E0B4")  # F 绿色
        fill_p = PatternFill("solid", fgColor="FFF2CC")  # P 黄色

        for r in range(rows_count):
            for c in range(len(fixed_headers), cols_count):
                table_item = table.item(r, c)
                if table_item and table_item.background().color().isValid():
                    bg_color = table_item.background().color()
                    excel_cell = ws.cell(r + first_data_row, c + 1)

                    # 根据背景色判断F/P
                    if bg_color.name() == "#c6e0b4":  # 绿色 - F
                        excel_cell.fill = fill_f
                    elif bg_color.name() == "#fff2cc":  # 黄色 - P
                        excel_cell.fill = fill_p

        # 7) 合计列样式
        sum_fill = PatternFill("solid", fgColor="E2F0D9")
        for c in range(len(fixed_headers) + 1, ws.max_column + 1):
            col_letter = get_column_letter(c)
            if "合计" in ws.cell(1, c).value or "Total" in ws.cell(1, c).value:
                for r in range(1, ws.max_row + 1):
                    ws.cell(r, c).fill = sum_fill
                # 数据行加粗
                for r in range(first_data_row, ws.max_row + 1):
                    ws.cell(r, c).font = Font(name="Arial", bold=True)

        # 8) TOTAL行样式
        total_row = ws.max_row
        ws.cell(total_row, 1).font = Font(name="Arial", bold=True)
        for c in range(1, ws.max_column + 1):
            ws.cell(total_row, c).font = Font(name="Arial", bold=True)
            ws.cell(total_row, c).border = thin_border

        # 保存文件
        wb.save(file_path)
        wb.close()
    # ===================== 订单明细页 =====================
    def create_order_details_tab(self):
        details_widget = QWidget()
        layout = QVBoxLayout(details_widget)

        # 过滤区
        filter_panel = QFrame()
        filter_panel.setFrameStyle(QFrame.StyledPanel)
        filter_panel.setStyleSheet("QFrame{background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;}")
        f_layout = QHBoxLayout(filter_panel)

        f_layout.addWidget(QLabel("版本:"))
        self.detail_version_combo = QComboBox()
        self.detail_version_combo.addItem("全部版本汇总")
        self.detail_version_combo.currentTextChanged.connect(lambda _: self.load_order_details())
        f_layout.addWidget(self.detail_version_combo)

        f_layout.addSpacing(12)
        f_layout.addWidget(QLabel("日期范围:"))
        self.detail_start_date_edit = QDateEdit()
        self.detail_start_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.detail_start_date_edit.setDate(QDate.currentDate().addDays(-30))
        f_layout.addWidget(self.detail_start_date_edit)

        f_layout.addWidget(QLabel("至"))
        self.detail_end_date_edit = QDateEdit()
        self.detail_end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.detail_end_date_edit.setDate(QDate.currentDate().addDays(30))
        f_layout.addWidget(self.detail_end_date_edit)

        apply_btn = QPushButton("应用筛选")
        apply_btn.clicked.connect(self.load_order_details)
        f_layout.addWidget(apply_btn)

        f_layout.addStretch()

        export_btn = QPushButton("导出明细CSV")
        export_btn.clicked.connect(self.export_order_data)
        f_layout.addWidget(export_btn)

        layout.addWidget(filter_panel)

        # 明细表
        self.details_table = QTableWidget()
        self.details_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.details_table.setAlternatingRowColors(True)
        hdr = self.details_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        
        # 设置表格的最小高度，确保在客户端范围内显示
        self.details_table.setMinimumHeight(300)
        self.details_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        layout.addWidget(self.details_table)

        self.tab_widget.addTab(details_widget, "订单明细")

    def load_order_details(self):
        try:
            ver_text = self.detail_version_combo.currentText()
            sd = self.detail_start_date_edit.date().toString("yyyy-MM-dd")
            ed = self.detail_end_date_edit.date().toString("yyyy-MM-dd")

            # 表头
            headers = [
                "Release Date","Release ID","Supplier","PN","ItemDesc",
                "DeliveryDate","CalendarWeek","类型","数量",
                "Purchase Order","Receipt Quantity","Cum Received"
            ]
            self.details_table.clear()
            self.details_table.setRowCount(0)
            self.details_table.setColumnCount(len(headers))
            for i, h in enumerate(headers):
                self.details_table.setHorizontalHeaderItem(i, QTableWidgetItem(h))

            # 数据
            if ver_text == "全部版本汇总":
                rows = CustomerOrderService.get_ndlutil_kanban_data(start_date=sd, end_date=ed)
                expanded = []
                for r in rows or []:
                    common = {
                        "ReleaseDate": r.get("ReleaseDate",""),
                        "ReleaseId": r.get("ReleaseId",""),
                        "SupplierCode": r.get("SupplierCode",""),
                        "ItemNumber": r.get("ItemNumber",""),
                        "ItemDescription": r.get("ItemDescription",""),
                        "DeliveryDate": r.get("DeliveryDate",""),
                        "CalendarWeek": r.get("CalendarWeek",""),
                        "PurchaseOrder": r.get("PurchaseOrder",""),
                        "ReceiptQuantity": 0,
                        "CumReceived": 0,
                    }
                    f = _norm_int(r.get("FirmQty"), 0)
                    p = _norm_int(r.get("ForecastQty"), 0)
                    if f:
                        e = dict(common); e["OrderType"]="F"; e["RequiredQty"]=f; expanded.append(e)
                    if p:
                        e = dict(common); e["OrderType"]="P"; e["RequiredQty"]=p; expanded.append(e)
                data = expanded
            else:
                import re
                m = re.search(r'^(\d+)\s-', ver_text)
                import_id = int(m.group(1)) if m else None
                data = CustomerOrderService.get_order_lines_by_import_version(import_id)

            # 填充
            for r, ln in enumerate(data or []):
                self.details_table.insertRow(r)
                vals = [
                    _get(ln, "ReleaseDate",""),
                    _get(ln, "ReleaseId",""),
                    _get(ln, "SupplierCode",""),
                    _get(ln, "ItemNumber",""),
                    _get(ln, "ItemDescription",""),
                    _get(ln, "DeliveryDate",""),
                    _get(ln, "CalendarWeek",""),
                    _get(ln, "OrderType",""),
                    str(_norm_int(_get(ln, "RequiredQty","req_qty", default=0), 0)),
                    _get(ln, "PurchaseOrder",""),
                    str(_norm_int(_get(ln, "ReceiptQuantity", default=0), 0)),
                    str(_norm_int(_get(ln, "CumReceived", default=0), 0)),
                ]
                for c, v in enumerate(vals):
                    self.details_table.setItem(r, c, QTableWidgetItem(v))

            self.details_table.resizeRowsToContents()
        except Exception as e:
            print(f"加载订单明细失败: {e}")

    def export_order_data(self):
        """把当前明细页数据导出为 CSV（简单直导，满足审阅/备份）"""
        try:
            path, _ = QFileDialog.getSaveFileName(self, "导出 CSV", "orders.csv", "CSV Files (*.csv)")
            if not path:
                return
            import csv
            headers = [self.details_table.horizontalHeaderItem(i).text() for i in range(self.details_table.columnCount())]
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                wr = csv.writer(f)
                wr.writerow(headers)
                for r in range(self.details_table.rowCount()):
                    row = []
                    for c in range(self.details_table.columnCount()):
                        it = self.details_table.item(r, c)
                        row.append(it.text() if it else "")
                    wr.writerow(row)
            QMessageBox.information(self, "导出完成", f"已导出到：\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", str(e))

    # ===================== 导入历史页 =====================
    def create_import_history_tab(self):
        history_widget = QWidget()
        layout = QVBoxLayout(history_widget)

        top = QHBoxLayout()
        refresh_btn = QPushButton("刷新历史")
        refresh_btn.clicked.connect(self.load_import_history)
        top.addWidget(refresh_btn)

        delete_btn = QPushButton("删除所选版本")
        delete_btn.clicked.connect(self.delete_selected_import)
        top.addWidget(delete_btn)

        top.addStretch()
        layout.addLayout(top)

        self.history_table = QTableWidget()
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.history_table.setAlternatingRowColors(True)
        
        # 设置表格的最小高度，确保在客户端范围内显示
        self.history_table.setMinimumHeight(300)
        self.history_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        layout.addWidget(self.history_table)

        self.tab_widget.addTab(history_widget, "导入历史")
        self.load_import_history()

    def load_import_history(self):
        try:
            rows = CustomerOrderService.get_import_history()
            headers = ["ImportId","FileName","ImportDate","OrderCount","LineCount","ImportStatus","ImportedBy"]
            self.history_table.clear()
            self.history_table.setRowCount(0)
            self.history_table.setColumnCount(len(headers))
            for i, h in enumerate(headers):
                self.history_table.setHorizontalHeaderItem(i, QTableWidgetItem(h))
            for r, row in enumerate(rows or []):
                self.history_table.insertRow(r)
                vals = [str(row.get(k,"")) for k in ["ImportId","FileName","ImportDate","OrderCount","LineCount","ImportStatus","ImportedBy"]]
                for c, v in enumerate(vals):
                    self.history_table.setItem(r, c, QTableWidgetItem(v))
            self.history_table.resizeColumnsToContents()

            # 同步两个版本下拉
            self.load_version_list()
        except Exception as e:
            print(f"加载导入历史失败: {e}")

    def delete_selected_import(self):
        r = self.history_table.currentRow()
        if r < 0:
            QMessageBox.information(self, "提示", "请先选中要删除的版本。")
            return
        import_id_item = self.history_table.item(r, 0)
        if not import_id_item:
            return
        import_id = int(import_id_item.text())
        ok = QMessageBox.question(self, "确认删除", f"确定删除版本 {import_id} 及其数据？", QMessageBox.Yes | QMessageBox.No)
        if ok != QMessageBox.Yes:
            return
        ok, msg = CustomerOrderService.delete_import(import_id)
        if not ok:
            QMessageBox.warning(self, "删除失败", msg or "未知错误")
            return
        QMessageBox.information(self, "删除成功", f"版本 {import_id} 已删除。")
        self.load_import_history()
        self.refresh_data()

    # ===================== 顶部动作 =====================
    def load_version_list(self):
        """刷新两个版本下拉框"""
        rows = CustomerOrderService.get_import_history()
        items = ["全部版本汇总"]
        for r in rows:
            items.append(f"{r['ImportId']} - {r['FileName']} ({r['ImportDate']})")

        # 看板页下拉
        cur = self.version_combo.currentText() if self.version_combo.count() else None
        self.version_combo.blockSignals(True)
        self.version_combo.clear()
        self.version_combo.addItems(items)
        self.version_combo.blockSignals(False)
        if cur and cur in items:
            self.version_combo.setCurrentText(cur)

        # 明细页下拉
        cur2 = self.detail_version_combo.currentText() if self.detail_version_combo.count() else None
        self.detail_version_combo.blockSignals(True)
        self.detail_version_combo.clear()
        self.detail_version_combo.addItems(items)
        self.detail_version_combo.blockSignals(False)
        if cur2 and cur2 in items:
            self.detail_version_combo.setCurrentText(cur2)

    def refresh_data(self):
        """刷新当前页数据"""
        # 先刷新版本列表（可能有新增/删除）
        self.load_version_list()
        # 触发两页数据加载
        self.load_kanban_data()
        self.load_order_details()

    def import_txt_file(self):
        """从 TXT 导入"""
        path, _ = QFileDialog.getOpenFileName(self, "选择客户订单 TXT", "", "Text Files (*.txt);;All Files (*)")
        if not path:
            return
        ok, msg, import_id = CustomerOrderService.import_orders_from_txt(path, import_user="UI")
        if not ok:
            QMessageBox.warning(self, "导入失败", msg or "未知错误")
        else:
            # 检查是否有不匹配的产品型号
            unmatched_items = self._check_unmatched_items(import_id)
            if unmatched_items:
                warning_msg = f"导入成功！\n\n发现以下产品型号没有匹配到项目，将默认放到最后：\n{unmatched_items}\n\n这些产品将按照默认规则排序。"
                QMessageBox.information(self, "导入成功（含警告）", warning_msg)
            else:
                QMessageBox.information(self, "导入成功", msg)
        # 刷新
        self.refresh_data()
        
        # 通知主窗口刷新MRP看板的订单版本列表
        self.notify_mrp_refresh()

    def _check_unmatched_items(self, import_id: int) -> str:
        """检查导入数据中是否有不匹配的产品型号"""
        try:
            data = CustomerOrderService.get_order_lines_by_import_version(import_id)
            unmatched = set()
            priority_bases = {
                "R001H368", "R001H369",  # Passat
                "R001P320", "R001P313",  # Tiguan L
                "R001J139", "R001J140",  # A5L
                "R001J141", "R001J142"   # Lavida
            }
            
            for line in data or []:
                pn = line.get("ItemNumber", "")
                if pn and len(pn) > 1:
                    # 使用新的匹配规则：去掉最后一位字母
                    base = pn[:-1]
                    if base not in priority_bases:
                        # 如果完整匹配失败，尝试基础项目匹配
                        if len(base) > 1 and base[-1].isdigit():
                            base_project = base[:-1]  # 去掉最后一位数字
                        else:
                            base_project = base
                        
                        base_project_bases = {"R001H36", "R001P32", "R001J13", "R001J14"}
                        if base_project not in base_project_bases:
                            unmatched.add(pn)
            
            if unmatched:
                return "\n".join(sorted(unmatched))
            return ""
        except Exception as e:
            print(f"检查不匹配项目失败: {e}")
            return ""

    def notify_mrp_refresh(self):
        """通知主窗口刷新MRP看板的订单版本列表"""
        try:
            # 通过父窗口链向上查找主窗口
            parent = self.parent()
            while parent:
                if hasattr(parent, 'content_area') and hasattr(parent.content_area, 'mrp_widget'):
                    parent.content_area.mrp_widget.refresh_order_versions()
                    break
                parent = parent.parent()
        except Exception as e:
            print(f"通知MRP刷新失败: {e}")

    def show_version_management(self):
        dlg = VersionManagementDialog(self)
        dlg.exec()
        # 关闭后刷新
        self.refresh_data()


# ===================== 版本管理对话框 =====================
class VersionManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("版本管理")
        self.resize(800, 480)

        layout = QVBoxLayout(self)

        tip = QLabel("提示：删除将同时清除该版本的头表和明细数据。")
        tip.setStyleSheet("color:#666;")
        layout.addWidget(tip)

        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        btns = QHBoxLayout()
        delete_btn = QPushButton("删除所选版本")
        delete_btn.clicked.connect(self.delete_selected)
        btns.addWidget(delete_btn)
        btns.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        btns.addWidget(close_btn)
        layout.addLayout(btns)

        self.load()

    def load(self):
        rows = CustomerOrderService.get_import_history()
        headers = ["ImportId","FileName","ImportDate","OrderCount","LineCount","ImportStatus","ImportedBy"]
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(len(headers))
        for i, h in enumerate(headers):
            self.table.setHorizontalHeaderItem(i, QTableWidgetItem(h))
        for r, row in enumerate(rows or []):
            self.table.insertRow(r)
            vals = [str(row.get(k,"")) for k in ["ImportId","FileName","ImportDate","OrderCount","LineCount","ImportStatus","ImportedBy"]]
            for c, v in enumerate(vals):
                self.table.setItem(r, c, QTableWidgetItem(v))
        self.table.resizeColumnsToContents()

    def delete_selected(self):
        r = self.table.currentRow()
        if r < 0:
            QMessageBox.information(self, "提示", "请先选中要删除的版本。")
            return
        import_id_item = self.table.item(r, 0)
        if not import_id_item:
            return
        import_id = int(import_id_item.text())
        ok = QMessageBox.question(self, "确认删除", f"确定删除版本 {import_id} 及其数据？", QMessageBox.Yes | QMessageBox.No)
        if ok != QMessageBox.Yes:
            return
        ok, msg = CustomerOrderService.delete_import(import_id)
        if not ok:
            QMessageBox.warning(self, "删除失败", msg or "未知错误")
            return
        QMessageBox.information(self, "删除成功", f"版本 {import_id} 已删除。")
        self.load()


