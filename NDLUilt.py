# NDLUilt.py —— 客户订单 TXT —— Excel 看板
# 依赖：openpyxl  (pip install openpyxl)
# 打包：pyinstaller NDLUilt.py -n CustomerOrderExtractor --onefile --windowed --clean

# ========== 可选调试 ==========
DEBUG = False
LOG = []
def dbg(msg: str):
    if DEBUG:
        LOG.append(msg)

# -------- 高 DPI（让界面清晰） --------
import ctypes
def _enable_dpi_awareness():
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)   # Win 8.1+
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()    # Win 7
        except Exception:
            pass
_enable_dpi_awareness()

import os
import re
import sys
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

# GUI
import tkinter as tk
from tkinter import filedialog, messagebox

# Excel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# -------- 解析规则（关键：头字段“行内匹配”，允许同一行多字段） --------
RE_SUPPLIER   = re.compile(r"^\s*Supplier:\s*([A-Za-z0-9\-]+)")
RE_SHIPTO     = re.compile(r"^\s*Ship-To:")
RE_ITEM       = re.compile(r"^\s*Item Number:\s*([A-Z0-9\-]+)", re.I)

RE_PO         = re.compile(r"Purchase Order:\s*([A-Z0-9\-]+)", re.I)
RE_RELID      = re.compile(r"Release ID:\s*([\w\-]+)", re.I)
RE_RELD       = re.compile(r"Release Date:\s*([0-9/]+)", re.I)
RE_RECEIPT_Q  = re.compile(r"Receipt Quantity:\s*([0-9][0-9,]*\.\d)", re.I)
RE_CUM_RECV   = re.compile(r"Cum Received:\s*([0-9][0-9,]*\.\d)", re.I)

RE_LINE       = re.compile(
    r"^\s*(?:Daily|Weekly|Monthly)?\s*([0-9]{2}/[0-9]{2}/[0-9]{2})\s+([FPfp])\s+([0-9][0-9,]*\.\d)(?:\s+.*)?$"
)

# 行顺序优先的 8 个 PN（保持原顺序）
DESIRED_PN_ORDER = [
    "R001H368E","R001H369E","R001P320B","R001P313B",
    "R001J139B","R001J140B","R001J141B","R001J142B"
]

# —— 项目映射按“去掉 PN 最后一个字母”规则
PN_TO_PROJECT_BASE = {
    "R001H368": "Passat rear double",
    "R001H369": "Passat rear single",
    "R001P320": "Tiguan L rear double",
    "R001P313": "Tiguan L rear single",
    "R001J139": "A5L rear double",
    "R001J140": "A5L rear single",
    "R001J141": "Lavida rear double",
    "R001J142": "Lavida rear single",
}

# -------- 解析 TXT --------
def parse_txt(paths):
    """
    返回 (lines, fp_map, suppliers, items, dates, warnings, release_info)

    lines: 明细需求行（含 due_date/FP/req_qty 等）
    fp_map: (supplier, item, date)->F/P（着色用）
    release_info: (supplier, item)-> {release_date(date或str), release_id, purchase_order, receipt_qty(float), cum_received(float)}
    """
    lines, fp_map, suppliers, items, dates = [], {}, set(), set(), set()
    today = date.today()
    warnings = {"unknown_pn": set(), "bad_fp": 0, "missing_core": 0}
    release_info = {}

    def parse_date_safe(s):
        try:
            return datetime.strptime(s, "%m/%d/%y").date()
        except Exception:
            return s  # 保底返回原字符串

    for p in map(Path, paths):
        raw = p.read_text(encoding="utf-8", errors="ignore")
        dbg(f"[FILE] {p.name}")

        # 当前Supplier & PN
        sup_code = sup_name = None
        item = None

        # Supplier级 header（先出现时暂存）
        header_po = None
        header_rel_id = None
        header_rel_date_txt = None
        header_receipt_qty = None
        header_cum_received = None

        # 当前 PN 的实际头字段值（作为导出基准）
        po = None
        rel_id = None
        rel_date_txt = None
        receipt_qty = None
        cum_received = None

        capture_name = False

        def flush_release_info():
            if sup_code and item:
                key = (sup_code, item)
                info = release_info.get(key, {})
                if rel_date_txt:
                    info["release_date"] = parse_date_safe(rel_date_txt)
                if rel_id:
                    info["release_id"] = rel_id
                if po:
                    info["purchase_order"] = po
                if receipt_qty is not None:
                    info["receipt_qty"] = float(str(receipt_qty).replace(",", ""))
                if cum_received is not None:
                    info["cum_received"] = float(str(cum_received).replace(",", ""))
                release_info[key] = info
                dbg(f"  [FLUSH] {sup_code}-{item}  RD={info.get('release_date')}  RID={info.get('release_id')}  PO={info.get('purchase_order')}  RQ={info.get('receipt_qty')}  CR={info.get('cum_received')}")

        for ln in raw.splitlines():
            # Supplier 段落（结构性行：独占处理并 continue）
            m = RE_SUPPLIER.search(ln)
            if m:
                flush_release_info()
                sup_code = m.group(1)
                sup_name = None
                capture_name = True
                # reset header & 当前 PN 值
                header_po = header_rel_id = header_rel_date_txt = None
                header_receipt_qty = header_cum_received = None
                item = None
                po = rel_id = rel_date_txt = None
                receipt_qty = cum_received = None
                dbg(f"[SUPPLIER] {sup_code}")
                continue

            if capture_name:
                if RE_SHIPTO.search(ln):
                    capture_name = False
                    continue
                t = ln.strip()
                if t:
                    sup_name = sup_name or t
                    capture_name = False
                continue

            # ===== 头字段：允许同一行多字段同时匹配（不使用 continue）=====
            m = RE_PO.search(ln)
            if m:
                if item:
                    po = m.group(1)
                else:
                    header_po = m.group(1)
                dbg(f"    PO -> {po or header_po}")

            m = RE_RELID.search(ln)
            if m:
                if item:
                    rel_id = m.group(1)
                else:
                    header_rel_id = m.group(1)
                dbg(f"    ReleaseID -> {rel_id or header_rel_id}")

            m = RE_RELD.search(ln)
            if m:
                if item:
                    rel_date_txt = m.group(1)
                else:
                    header_rel_date_txt = m.group(1)
                dbg(f"    ReleaseDate -> {rel_date_txt or header_rel_date_txt}")

            m = RE_RECEIPT_Q.search(ln)
            if m:
                if item:
                    receipt_qty = m.group(1)
                else:
                    header_receipt_qty = m.group(1)
                dbg(f"    ReceiptQty -> {receipt_qty or header_receipt_qty}")

            m = RE_CUM_RECV.search(ln)
            if m:
                if item:
                    cum_received = m.group(1)
                else:
                    header_cum_received = m.group(1)
                dbg(f"    CumReceived -> {cum_received or header_cum_received}")

            # Item Number（结构性行：处理后 continue）
            m = RE_ITEM.search(ln)
            if m:
                flush_release_info()
                item = m.group(1)
                # 用 header_* 初始化当前 PN（关键）
                po = header_po
                rel_id = header_rel_id
                rel_date_txt = header_rel_date_txt
                receipt_qty = header_receipt_qty
                cum_received = header_cum_received
                dbg(f"[ITEM] {item}  init: RD={rel_date_txt}  RID={rel_id}  PO={po}  RQ={receipt_qty}  CR={cum_received}")
                continue

            # 计划行
            m = RE_LINE.match(ln)
            if m:
                if not (sup_code and item):
                    warnings["missing_core"] += 1
                    continue
                d_s, fp, qty_s = m.groups()
                d = datetime.strptime(d_s, "%m/%d/%y").date()
                fp = fp.upper() if fp and fp.upper() in ("F", "P") else "P"
                qty = float(qty_s.replace(",", ""))

                rec = {
                    "supplier_code": sup_code,
                    "supplier_name": sup_name,
                    "purchase_order": po,
                    "item_number": item,
                    "due_date": d,
                    "FP": fp,
                    "req_qty": qty,
                    "release_id": rel_id,
                    "release_date_text": rel_date_txt,  # 原样保留
                    "file_date": today,
                }
                lines.append(rec)
                suppliers.add(sup_code); items.add(item); dates.add(d)

                key = (sup_code, item, d)
                cur = fp_map.get(key)
                if (cur is None) or (cur == "P" and fp == "F"):
                    fp_map[key] = fp

                if item not in DESIRED_PN_ORDER:
                    warnings["unknown_pn"].add(item)

        flush_release_info()

    lines.sort(key=lambda r: (r["supplier_code"], r["item_number"], r["due_date"]))
    return lines, fp_map, sorted(suppliers), sorted(items), sorted(dates), warnings, release_info


def build_board(lines, dates):
    board = defaultdict(lambda: defaultdict(float))
    for r in lines:
        key = (r["supplier_code"], r["item_number"])
        board[key][r["due_date"]] += r["req_qty"]
    for key in board:
        for d in dates:
            board[key].setdefault(d, 0.0)
    return board


# -------- 导出 Excel --------
def export_excel(lines, board, dates, fp_map, out_path, release_info):
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active

    # 日期按年分组
    by_year = defaultdict(list)
    dates = sorted(dates)
    for d in dates:
        y = d.isocalendar()[0]
        by_year[y].append(d)
    for y in by_year:
        by_year[y].sort()
    years_sorted = sorted(by_year.keys())

    # 每年后加合计
    col_specs = []
    for y in years_sorted:
        for d in by_year[y]:
            col_specs.append(("date", d))
        col_specs.append(("sum", y))

    # 固定列
    fixed_headers = [
        "Release Date", "Release ID", "PN", "Des", "Project", "Item",
        "Purchase Order", "Receipt Quantity", "Cum Received"
    ]

    # 表头
    top = fixed_headers[:] + [
        (f"CW{val.isocalendar()[1]:02d}" if kind == "date" else f"{val}合计")
        for kind, val in col_specs
    ]
    ws.append(top)
    second = [""] * len(fixed_headers) + [
        (val.strftime("%m/%d") if kind == "date" else "")
        for kind, val in col_specs
    ]
    ws.append(second)
    release_dates = [v.get("release_date") for v in release_info.values() if v.get("release_date")]
    if release_dates:
        first_rel_date = release_dates[0]
        if isinstance(first_rel_date, date):
            cw = first_rel_date.isocalendar()[1]
            ws.title = f"CW{cw:02d}"
        else:
            # 如果解析不到 date 对象，就退回默认
            ws.title = "board"
    else:
        ws.title = "board"

    # 行顺序
    keys = list(board.keys())
    known_keys_by_order = []
    for pn in DESIRED_PN_ORDER:
        group = sorted([k for k in keys if k[1] == pn])
        known_keys_by_order.extend(group)
    unknown_keys = sorted([k for k in keys if k[1] not in DESIRED_PN_ORDER],
                          key=lambda k: (k[1], k[0]))
    ordered_keys = known_keys_by_order + unknown_keys

    # 数据行
    row_index = {}
    for (supplier_code, item_number) in ordered_keys:
        head = release_info.get((supplier_code, item_number), {})

        rel_date_val = head.get("release_date", "")
        if isinstance(rel_date_val, date):
            rel_date_str = rel_date_val.strftime("%Y/%m/%d")
        else:
            rel_date_str = str(rel_date_val or "")

        rel_id        = head.get("release_id", "") or ""
        po            = head.get("purchase_order", "") or ""
        receipt_qty   = int(head.get("receipt_qty", 0.0) or 0.0)
        cum_received  = int(head.get("cum_received", 0.0) or 0.0)

        pn_base = item_number[:-1] if item_number else ""
        project = PN_TO_PROJECT_BASE.get(pn_base, "UNKNOWN")

        fixed_vals = [
            rel_date_str,
            rel_id,
            item_number,
            "PEMM ASSY",
            project,
            "Gross Reqs",
            po,
            receipt_qty,
            cum_received,
        ]

        print(item_number)
        cols = board[(supplier_code, item_number)]
        tail_vals = []
        for kind, val in col_specs:
            if kind == "date":
                tail_vals.append(int(cols[val]))
            else:
                tail_vals.append(sum(int(cols[d]) for d in by_year[val]))

        ws.append(fixed_vals + tail_vals)
        row_index[(supplier_code, item_number)] = ws.max_row

    # TOTAL 行
    totals = ["TOTAL"] + [""]*(len(fixed_headers)-1) + [0]*len(col_specs)
    ws.append(totals)

    # ====== 样式 & 版式 ======

    # 1) 列宽：固定列紧凑；数字列更窄
    tight_widths = [11, 14, 12, 10, 16, 10, 14, 12, 12]  # 对应 9 个固定列
    for i, w in enumerate(tight_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for ci in range(len(fixed_headers)+1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(ci)].width = 9.0

    # 2) 冻结窗格：固定列 + 两行表头
    ws.freeze_panes = get_column_letter(len(fixed_headers)+1) + "3"

    # 3) 表头字体加粗、边框
    for cell in ws[1] + ws[2]:
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 4) 内容区样式（固定列自动换行/缩放；数字列右对齐/缩放；仅“有数据”画边框）
    def has_value(c):
        v = c.value
        return not (v is None or (isinstance(v, str) and v.strip() == ""))

    first_data_row = 3
    for r in range(first_data_row, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell = ws.cell(r, c)
            if c <= len(fixed_headers):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center",
                    wrap_text=True, shrink_to_fit=True
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center",
                    wrap_text=False, shrink_to_fit=True
                )
            cell.font = Font(name="Arial", bold=bool(cell.font and cell.font.bold))
            if has_value(cell):  # 如需全表边框，把这行改成：cell.border = border
                cell.border = border

    # 5) 日期区域数字格式为整数
    for r in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row,
                          min_col=len(fixed_headers)+1, max_col=ws.max_column):
        for c in r:
            c.number_format = "0"

    # 6) F/P 着色（仅日期列）
    date_col_idx = {}
    col = len(fixed_headers) + 1
    for kind, val in col_specs:
        if kind == "date":
            date_col_idx[val] = col
        col += 1
    fill_f = PatternFill("solid", fgColor="C6E0B4")   # F 绿色
    fill_p = PatternFill("solid", fgColor="FFF2CC")   # P 黄色
    for (supplier, item, d), fp in fp_map.items():
        r = row_index.get((supplier, item))
        c = date_col_idx.get(d)
        if r and c:
            cell = ws.cell(r, c)
            if cell.value and cell.value != 0:
                cell.fill = fill_f if fp == "F" else fill_p

    # 7) 当前周头部蓝色 + 第一个 CW 蓝色
    _, week_now, _ = date.today().isocalendar()
    blue = PatternFill("solid", fgColor="BDD7EE")
    first_cw_done = False
    col = len(fixed_headers)+1
    for kind, val in col_specs:
        if kind == "date":
            cw = val.isocalendar()[1]
            if cw == week_now:
                ws.cell(1, col).fill = blue
            if not first_cw_done:
                ws.cell(1, col).fill = blue
                first_cw_done = True
        col += 1

    # 8) 合计列底色 + 数值加粗
    sum_fill = PatternFill("solid", fgColor="E2F0D9")
    col = len(fixed_headers)+1
    for kind, val in col_specs:
        if kind == "sum":
            for r in range(1, ws.max_row+1):
                ws.cell(r, col).fill = sum_fill
            for r in range(first_data_row, ws.max_row+1):
                ws.cell(r, col).font = Font(name="Arial", bold=True)
        col += 1

    # 9) TOTAL 求和
    total_row = ws.max_row
    for c in range(len(fixed_headers)+1, ws.max_column+1):
        s = 0
        for r in range(first_data_row, total_row):
            s += int(ws.cell(r, c).value or 0)
        ws.cell(total_row, c).value = s
        ws.cell(total_row, c).font = Font(name="Arial", bold=True)

    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 1).font = Font(name="Arial", bold=True)
    for c in range(1, ws.max_column + 1):
        ws.cell(total_row, c).border = border
    wb.save(out_path)


# -------- GUI --------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NDKJ客户订单TXT转换")
        self.geometry("730x240")
        self.minsize(730, 240)
        self.resizable(True, False)

        frm = tk.Frame(self); frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        tk.Label(frm, text="选择客户订单 TXT 文件：").grid(row=0, column=0, sticky="w")
        self.files_var = tk.StringVar()
        e1 = tk.Entry(frm, textvariable=self.files_var)
        e1.grid(row=1, column=0, sticky="we", padx=(0,10))
        tk.Button(frm, text="Browse", width=12, command=self.browse_files).grid(row=1, column=1)

        tk.Label(frm, text="输出 Excel 路径：").grid(row=2, column=0, sticky="w", pady=(10,0))
        self.out_var = tk.StringVar(value="customer_order_extract.xlsx")
        e2 = tk.Entry(frm, textvariable=self.out_var)
        e2.grid(row=3, column=0, sticky="we", padx=(0,10))
        tk.Button(frm, text="Save As...", width=12, command=self.save_as).grid(row=3, column=1)

        btn_frame = tk.Frame(frm); btn_frame.grid(row=4, column=0, sticky="w", pady=(12,0))
        tk.Button(btn_frame, text="开始提取", width=14, command=self.run).pack(side=tk.LEFT, padx=(0,12))
        tk.Button(btn_frame, text="退出", width=12, command=self.destroy).pack(side=tk.LEFT)

        frm.grid_columnconfigure(0, weight=1)
        frm.grid_columnconfigure(1, weight=0)

    def browse_files(self):
        paths = filedialog.askopenfilenames(title="选择 TXT 文件", filetypes=[("Text","*.txt"),("All","*.*")])
        if not paths: return
        self.files_var.set(";".join(paths))

    def save_as(self):
        out = filedialog.asksaveasfilename(title="保存 Excel", defaultextension=".xlsx",
                                           filetypes=[("Excel","*.xlsx")],
                                           initialfile="customer_order_extract.xlsx")
        if out:
            self.out_var.set(out)

    def run(self):
        files = [f for f in self.files_var.get().split(";") if f.strip()]
        out = self.out_var.get().strip() or "customer_order_extract.xlsx"
        if not files:
            messagebox.showerror("提示","请先选择一个或多个 TXT 文件")
            return
        try:
            lines, fp_map, _, _, dates, warnings, release_info = parse_txt(files)
            if not lines:
                messagebox.showerror("提示","没有解析到任何订单行，请检查TXT格式")
                return
            board = build_board(lines, dates)
            export_excel(lines, board, dates, fp_map, out, release_info)

            warn_lines = []
            if warnings["unknown_pn"]:
                warn_lines.append("未知 PN（已置底）：\n  - " + "\n  - ".join(sorted(warnings["unknown_pn"])))
            if warnings["bad_fp"]:
                warn_lines.append(f"异常 F/P 已按 P 处理：{warnings['bad_fp']} 行")
            if warnings["missing_core"]:
                warn_lines.append(f"缺少关键字段（Supplier/Item）行数：{warnings['missing_core']}")

            # 写调试日志（若开启）
            if DEBUG and LOG:
                log_path = os.path.splitext(out)[0] + "_parse_debug.txt"
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(LOG))
                warn_lines.append(f"(Debug) 解析日志已生成：{log_path}")

            if warn_lines:
                messagebox.showinfo("成功", f"已导出：\n{out}\n\n" + "\n\n".join(warn_lines))
            else:
                messagebox.showinfo("成功", f"已导出：\n{out}")

            try:
                os.startfile(out)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("错误", f"解析失败：\n{e}")

# -------- 主程序 --------
def main():
    global DEBUG
    if len(sys.argv) > 1:
        args = sys.argv[1:]
        out = "customer_order_extract.xlsx"
        if "--debug" in args:
            DEBUG = True
            args = [a for a in args if a != "--debug"]
        if "-o" in args:
            i = args.index("-o")
            out = args[i+1]
            args = args[:i] + args[i+2:]
        lines, fp_map, _, _, dates, warnings, release_info = parse_txt(args)
        if not lines:
            print("Nothing parsed. Check your TXT format."); return
        board = build_board(lines, dates)
        export_excel(lines, board, dates, fp_map, out, release_info)
        print(f"[OK] {out}")
        if warnings["unknown_pn"] or warnings["bad_fp"] or warnings["missing_core"]:
            print("Warnings:", warnings)
        if DEBUG and LOG:
            log_path = os.path.splitext(out)[0] + "_parse_debug.txt"
            with open(log_path, "w", encoding="utf-8") as f:
                f.write("\n".join(LOG))
            print(f"[DEBUG] Log written -> {log_path}")
    else:
        App().mainloop()

if __name__ == "__main__":
    main()
